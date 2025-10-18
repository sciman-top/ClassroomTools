# -*- coding: utf-8 -*-
from __future__ import annotations

import base64
import configparser
import contextlib
import ctypes
from ctypes import wintypes
import importlib
import io
import json
import logging
import math
import os
import random
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import traceback
import hashlib
import hmac
from collections import OrderedDict, deque
from queue import Empty, Queue
from dataclasses import dataclass
from enum import Enum
from typing import (
    TYPE_CHECKING,
    Any,
    Callable,
    Dict,
    Iterable,
    List,
    Mapping,
    Optional,
    Set,
    Tuple,
    Union,
    cast,
)

logger = logging.getLogger(__name__)

if sys.platform == "win32":
    try:  # pragma: no cover - Windows 专用依赖
        import win32api
        import win32con
        import win32gui
    except ImportError:  # pragma: no cover - 部分环境未安装 pywin32
        win32api = None  # type: ignore[assignment]
        win32con = None  # type: ignore[assignment]
        win32gui = None  # type: ignore[assignment]
    try:
        _USER32 = ctypes.windll.user32  # type: ignore[attr-defined]
    except Exception:  # pragma: no cover - 某些环境可能限制 Win32 API
        _USER32 = None  # type: ignore[assignment]
else:
    win32api = None  # type: ignore[assignment]
    win32con = None  # type: ignore[assignment]
    win32gui = None  # type: ignore[assignment]
    _USER32 = None  # type: ignore[assignment]

VK_UP = getattr(win32con, "VK_UP", 0x26)
VK_DOWN = getattr(win32con, "VK_DOWN", 0x28)
VK_LEFT = getattr(win32con, "VK_LEFT", 0x25)
VK_RIGHT = getattr(win32con, "VK_RIGHT", 0x27)
VK_PRIOR = getattr(win32con, "VK_PRIOR", 0x21)
VK_NEXT = getattr(win32con, "VK_NEXT", 0x22)
VK_SPACE = getattr(win32con, "VK_SPACE", 0x20)
VK_RETURN = getattr(win32con, "VK_RETURN", 0x0D)
KEYEVENTF_EXTENDEDKEY = getattr(win32con, "KEYEVENTF_EXTENDEDKEY", 0x0001)
KEYEVENTF_KEYUP = getattr(win32con, "KEYEVENTF_KEYUP", 0x0002)
_NAVIGATION_EXTENDED_KEYS = {VK_UP, VK_DOWN, VK_LEFT, VK_RIGHT}
WHEEL_DELTA = getattr(win32con, "WHEEL_DELTA", 120)

_NAVIGATION_WHEEL_DELTAS: Dict[int, int] = {
    VK_UP: WHEEL_DELTA,
    VK_PRIOR: WHEEL_DELTA,
    VK_DOWN: -WHEEL_DELTA,
    VK_NEXT: -WHEEL_DELTA,
}

NAVIGATION_RESTORE_TIMEOUT = 0.45


class ToolMode(str, Enum):
    BRUSH = "brush"
    SHAPE = "shape"
    ERASER = "eraser"
    CURSOR = "cursor"


class InteractionMode(str, Enum):
    DRAWING = "drawing"
    NAVIGATION = "navigation"
    CURSOR = "cursor"


@dataclass
class NavigationOriginState:
    mode: ToolMode
    shape: Optional[str]
    inside_toolbar: bool


@dataclass(frozen=True)
class ModeState:
    tool: ToolMode
    interaction: InteractionMode
    passthrough: bool
    canvas_hidden: bool
    keyboard_grabbed: bool = False


@dataclass(frozen=True)
class InputContext:
    global_pos: Optional["QPoint"]
    inside_toolbar: bool
    tool: ToolMode

if _USER32 is not None:
    _WNDENUMPROC = ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
else:  # pragma: no cover - 非 Windows 平台不会调用
    _WNDENUMPROC = None  # type: ignore[assignment]


def _user32_window_rect(hwnd: int) -> Optional[Tuple[int, int, int, int]]:
    if _USER32 is None or hwnd == 0:
        return None
    rect = wintypes.RECT()
    try:
        ok = bool(_USER32.GetWindowRect(wintypes.HWND(hwnd), ctypes.byref(rect)))
    except Exception:
        return None
    if not ok:
        return None
    return rect.left, rect.top, rect.right, rect.bottom


def _user32_is_window(hwnd: int) -> bool:
    if _USER32 is None or hwnd == 0:
        return False
    try:
        return bool(_USER32.IsWindow(wintypes.HWND(hwnd)))
    except Exception:
        return False


def _user32_is_window_visible(hwnd: int) -> bool:
    if _USER32 is None or hwnd == 0:
        return False
    try:
        return bool(_USER32.IsWindowVisible(wintypes.HWND(hwnd)))
    except Exception:
        return False


def _user32_is_window_iconic(hwnd: int) -> bool:
    if _USER32 is None or hwnd == 0:
        return False
    try:
        return bool(_USER32.IsIconic(wintypes.HWND(hwnd)))
    except Exception:
        return False


def _user32_window_class_name(hwnd: int) -> str:
    if _USER32 is None or hwnd == 0:
        return ""
    buffer = ctypes.create_unicode_buffer(256)
    try:
        length = int(_USER32.GetClassNameW(wintypes.HWND(hwnd), buffer, len(buffer)))
    except Exception:
        return ""
    if length <= 0:
        return ""
    return buffer.value.strip().lower()


def _user32_get_foreground_window() -> int:
    if _USER32 is None:
        return 0
    try:
        return int(_USER32.GetForegroundWindow())
    except Exception:
        return 0


def _user32_get_parent(hwnd: int) -> int:
    if _USER32 is None or hwnd == 0:
        return 0
    try:
        return int(_USER32.GetParent(wintypes.HWND(hwnd)))
    except Exception:
        return 0


def _user32_top_level_hwnd(hwnd: int) -> int:
    if _USER32 is None or hwnd == 0:
        return hwnd
    try:
        ga_root = getattr(win32con, "GA_ROOT", 2) if win32con is not None else 2
    except Exception:
        ga_root = 2
    try:
        ancestor = int(_USER32.GetAncestor(wintypes.HWND(hwnd), ga_root))
    except Exception:
        ancestor = 0
    if ancestor:
        return ancestor
    parent = _user32_get_parent(hwnd)
    return parent or hwnd


def _user32_focus_window(hwnd: int) -> bool:
    if _USER32 is None or hwnd == 0:
        return False
    focused = False
    try:
        focused = bool(_USER32.SetForegroundWindow(wintypes.HWND(hwnd)))
    except Exception:
        focused = False
    if not focused:
        try:
            focused = bool(_USER32.SetActiveWindow(wintypes.HWND(hwnd)))
        except Exception:
            focused = False
    focus_ok = False
    try:
        focus_ok = bool(_USER32.SetFocus(wintypes.HWND(hwnd)))
    except Exception:
        focus_ok = False
    return focused or focus_ok

from PyQt6.QtCore import (
    QByteArray,
    QPoint,
    QPointF,
    QRect,
    QRectF,
    QSize,
    Qt,
    QTimer,
    QEvent,
    pyqtSignal,
    QObject,
)
from PyQt6.QtGui import (
    QBrush,
    QColor,
    QCursor,
    QFont,
    QFontDatabase,
    QFontMetrics,
    QIcon,
    QPainter,
    QPainterPath,
    QPainterPathStroker,
    QPen,
    QPixmap,
    QKeyEvent,
    QResizeEvent,
    QScreen,
    QWheelEvent,
    QAction,
)
from PyQt6.QtWidgets import (
    QApplication,
    QButtonGroup,
    QCheckBox,
    QDialog,
    QDialogButtonBox,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMenu,
    QInputDialog,
    QMessageBox,
    QPushButton,
    QSpacerItem,
    QSizePolicy,
    QSlider,
    QSpinBox,
    QStackedWidget,
    QToolButton,
    QVBoxLayout,
    QWidget,
    QToolTip,
)

# ---------- 运行环境准备 ----------

def _prepare_windows_tts_environment() -> None:
    """确保 Windows 打包环境下的语音依赖可以写入缓存。"""

    if sys.platform != "win32":
        return
    cache_dir = os.environ.get("COMTYPES_CACHE_DIR", "").strip()
    if cache_dir and os.path.isdir(cache_dir):
        return
    try:
        base = os.environ.get("LOCALAPPDATA")
        if not base:
            base = os.path.join(os.path.expanduser("~"), "AppData", "Local")
        cache_dir = os.path.join(base, "ClassroomTools", "comtypes_cache")
        os.makedirs(cache_dir, exist_ok=True)
        os.environ["COMTYPES_CACHE_DIR"] = cache_dir
    except Exception:
        # 打包环境下若目录创建失败，也不要阻塞主程序。
        pass


# 兼容早期 Python 版本缺失的 ULONG_PTR 定义，供 Win32 输入结构体使用。
if not hasattr(wintypes, "ULONG_PTR"):
    if ctypes.sizeof(ctypes.c_void_p) == ctypes.sizeof(ctypes.c_ulonglong):
        wintypes.ULONG_PTR = ctypes.c_ulonglong  # type: ignore[attr-defined]
    else:
        wintypes.ULONG_PTR = ctypes.c_ulong  # type: ignore[attr-defined]

_prepare_windows_tts_environment()

# ---------- 图标 ----------
class IconManager:
    """集中管理浮动工具条的 SVG 图标，方便后续统一换肤。"""
    _icons: Dict[str, str] = {
        "cursor": "PHN2ZyB4bWxucz0naHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmcnIHZpZXdCb3g9JzAgMCAyNCAyNCc+CiAgICA8cGF0aCBmaWxsPScjZjFmM2Y0JyBkPSdNNCAzLjMgMTEuNCAyMWwxLjgtNS44IDYuMy0yLjF6Jy8+CiAgICA8cGF0aCBmaWxsPScjOGFiNGY4JyBkPSdtMTIuNiAxNC40IDQuOCA0LjgtMi4xIDIuMS00LjItNC4yeicvPgo8L3N2Zz4=",
        "shape": "PHN2ZyB4bWxucz0naHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmcnIHZpZXdCb3g9JzAgMCAyNCAyNCc+CiAgICA8cmVjdCB4PSczLjUnIHk9JzMuNScgd2lkdGg9JzknIGhlaWdodD0nOScgcng9JzInIGZpbGw9JyNmMWYzZjQnLz4KICAgIDxjaXJjbGUgY3g9JzE2LjUnIGN5PScxNi41JyByPSc1LjUnIGZpbGw9J25vbmUnIHN0cm9rZT0nI2YxZjNmNCcgc3Ryb2tlLXdpZHRoPScxLjgnLz4KICAgIDxjaXJjbGUgY3g9JzE2LjUnIGN5PScxNi41JyByPSczLjUnIGZpbGw9JyM4YWI0ZjgnIGZpbGwtb3BhY2l0eT0nMC4zNScvPgo8L3N2Zz4=",
        "eraser": "PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHZpZXdCb3g9IjAgMCAyNCAyNCI+CiAgPHBhdGggZD0iTTQuNiAxNC4yIDExLjMgNy40YTIgMiAwIDAgMSAyLjggMGwzLjUgMy41YTIgMiAwIDAgMSAwIDIuOGwtNC44IDQuOEg5LjRhMiAyIDAgMCAxLTEuNC0uNmwtMy0zYTIgMiAwIDAgMSAwLTIuOHoiIGZpbGw9IiNmNGE5YjciLz4KICA8cGF0aCBkPSJNOS4yIDE5LjZoNi4xYy42IDAgMS4xLS4yIDEuNS0uNmwxLjctMS43IiBmaWxsPSJub25lIiBzdHJva2U9IiM1ZjYzNjgiIHN0cm9rZS13aWR0aD0iMS42IiBzdHJva2UtbGluZWNhcD0icm91bmQiLz4KICA8cGF0aCBkPSJtNy4yIDEyLjMgNC41IDQuNSIgZmlsbD0ibm9uZSIgc3Ryb2tlPSIjZmZmZmZmIiBzdHJva2Utd2lkdGg9IjEuNiIgc3Rya2UtbGluZWNhcD0icm91bmQiLz4KICA8cGF0aCBkPSJNMy42IDE4LjZoNiIgc3Ryb2tlPSIjNWY2MzY4IiBzdHJva2Utd2lkdGg9IjEuNiIgc3Rya2UtbGluZWNhcD0icm91bmQiLz4KPC9zdmc+",
        "clear_all": "PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHZpZXdCb3g9IjAgMCAyNCAyNCI+CiAgPGRlZnM+CiAgICA8bGluZWFyR3JhZGllbnQgaWQ9ImciIHgxPSIwIiB4Mj0iMCIgeTE9IjAiIHkyPSIxIj4KICAgICAgPHN0b3Agb2Zmc2V0PSIwIiBzdG9wLWNvbG9yPSIjOGFiNGY4Ii8+CiAgICAgIDxzdG9wIG9mZnNldD0iMSIgc3RvcC1jb2xvcj0iIzFhNzNlOCIvPgogICAgPC9saW5lYXJHcmFkaWVudD4KICA8L2RlZnM+CiAgPHBhdGggZD0iTTUuNSA4aDEzbC0uOSAxMS4yQTIgMiAwIDAgMSAxNS42IDIxSDguNGEyIDIgMCAwIDEtMS45LTEuOEw1LjUgOHoiIGZpbGw9InVybCgjZykiIHN0cm9rZT0iIzFhNzNlOCIgc3Rya2Utd2lkdGg9IjEuMiIvPgogIDxwYXRoIGQ9Ik05LjUgNS41IDEwLjMgNGgzLjRsLjggMS41aDQuNSIgZmlsbD0ibm9uZSIgc3Rya2U9IiM1ZjYzNjgiIHN0cm9rZS13aWR0aD0iMS42IiBzdHJva2UtbGluZWNhcD0icm91bmQiIHN0cm9rZS1saW5lam9pbj0icm91bmQiLz4KICA8cGF0aCBkPSJNNSA1LjVoNCIgc3Ryb2tlPSIjNWY2MzY4IiBzdHJva2Utd2lkdGg9IjEuNiIgc3Rya2UtbGluZWNhcD0icm91bmQiLz4KICA8cGF0aCBkPSJNMTAgMTEuMnY2LjFNMTQgMTEuMnY2LjEiIHN0cm9rZT0iI2ZmZmZmZiIgc3Rya2Utd2lkdGg9IjEuNCIgc3Rya2UtbGluZWNhcD0icm91bmQiLz4KICA8cGF0aCBkPSJNOC4yIDExLjJ2Ni4xIiBzdHJva2U9IiMzYjc4ZTciIHN0cm9rZS13aWR0aD0iMS40IiBzdHJva2UtbGluZWNhcD0icm91bmQiIG9wYWNpdHk9Ii43Ii8+CiAgPHBhdGggZD0iTTE1LjggMTEuMnY2LjEiIHN0cm9rZT0iIzNiNzhlNyIgc3Rya2Utd2lkdGg9IjEuNCIgc3Rya2UtbGluZWNhcD0icm91bmQiIG9wYWNpdHk9Ii43Ii8+CiAgPHBhdGggZD0iTTYuMiAzLjYgNy40IDIuNCIgc3Ryb2tlPSIjZmJiYzA0IiBzdHJva2Utd2lkdGg9IjEuNCIgc3Rya2UtbGluZWNhcD0icm91bmQiLz4KICA8cGF0aCBkPSJtMTguNCAzLjQgMS40LTEuNCIgc3Rya2U9IiMzNGE4NTMiIHN0cm9rZS13aWR0aD0iMS40IiBzdHJva2UtbGluZWNhcD0icm91bmQiLz4KPC9zdmc+",
        "settings": "PHN2ZyB4bWxucz0naHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmcnIHZpZXdCb3g9JzAgMCAyNCAyNCc+CiAgICA8Y2lyY2xlIGN4PScxMicgY3k9JzEyJyByPSczLjUnIGZpbGw9JyM4YWI0ZjgnLz4KICAgIDxwYXRoIGZpbGw9J25vbmUnIHN0cm9rZT0nI2YxZjNmNCcgc3Ryb2tlLXdpZHRoPScxLjYnIHN0cm9rZS1saW5lY2FwPSdyb3VuZCcgc3Ryb2tlLWxpbmVqb2luPSdyb3VuZCcKICAgICAgICBkPSdNMTIgNC41VjIuOG0wIDE4LjR2LTEuN203LjEtNy41SDIwbS0xOCAwaDEuNk0xNy42IDZsMS4yLTEuMk01LjIgMTguNCA2LjQgMTcuMk02LjQgNiA1LjIgNC44bTEzLjYgMTMuNi0xLjItMS4yJy8+Cjwvc3ZnPg==",
        "whiteboard": "PHN2ZyB4bWxucz0naHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmcnIHZpZXdCb3g9JzAgMCAyNCAyNCc+CiAgICA8cmVjdCB4PSczJyB5PSc0JyB3aWR0aD0nMTgnIGhlaWdodD0nMTInIHJ4PScyJyByeT0nMicgZmlsbD0nI2YxZjNmNCcgZmlsbC1vcGFjaXR5PScwLjEyJyBzdHJva2U9JyNmMWYzZjQnIHN0cm9rZS13aWR0aD0nMS42Jy8+CiAgICA8cGF0aCBkPSdtNyAxOCA1LTUgNSA1JyBmaWxsPSdub25lJyBzdHJva2U9JyM4YWI0ZjgnIHN0cm9rZS13aWR0aD0nMS44JyBzdHJva2UtbGluZWNhcD0ncm91bmQnIHN0cm9rZS1saW5lam9pbj0ncm91bmQnLz4KICAgIDxwYXRoIGQ9J004IDloOG0tOCAzaDUnIHN0cm9rZT0nI2YxZjNmNCcgc3Ryb2tlLXdpZHRoPScxLjYnIHN0cm9rZS1saW5lY2FwPSdyb3VuZCcvPgo8L3N2Zz4=",
        "undo": "PHN2ZyB4bWxucz0naHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmcnIHZpZXdCb3g9JzAgMCAyNCAyNCc+CiAgPHBhdGggZmlsbD0nI2YxZjNmNCcgZD0nTTguNCA1LjJMMyAxMC42bDUuNCA1LjQgMS40LTEuNC0yLjMtMi4zaDUuNWMzLjIgMCA1LjggMi42IDUuOCA1LjggMCAuNS0uMSAxLS4yIDEuNWwyLjEuNmMuMi0uNy4zLTEuNC4zLTIuMSAwLTQuNC0zLjYtOC04LThINy41bDIuMy0yLjMtMS40LTEuNHonLz4KPC9zdmc+",
    }
    _cache: Dict[str, QIcon] = {}
    _icons.update(
        {
            "arrow_down": "PHN2ZyB4bWxucz0naHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmcnIHZpZXdCb3g9JzAgMCAyNCAyNCc+PHBhdGggZmlsbD0nI2YxZjNmNCcgZD0nTTEyIDE2LjUgNSA5LjVoMTR6Jy8+PHBhdGggZmlsbD0nbm9uZScgc3Ryb2tlPScjMWE3M2U4JyBzdHJva2Utd2lkdGg9JzEuNScgc3Rya2UtbGluZWpvaW49J3JvdW5kJyBkPSdNMTIgMTYuNSA1IDkuNWgxNHonLz48L3N2Zz4=",
            "arrow_up": "PHN2ZyB4bWxucz0naHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmcnIHZpZXdCb3g9JzAgMCAyNCAyNCc+PHBhdGggZmlsbD0nI2YxZjNmNCcgZD0nTTEyIDcuNSAxOSAxNC41SDV6Jy8+PHBhdGggZmlsbD0nbm9uZScgc3Ryb2tlPScjMWE3M2U4JyBzdHJva2Utd2lkdGg9JzEuNScgc3Ryb2tlLWxpbmVqb2luPSdyb3VuZCcgZD0nTTEyIDcuNSAxOSAxNC41SDV6Jy8+PC9zdmc+",
        }
    )

    @classmethod
    def get_brush_icon(cls, color_hex: str) -> QIcon:
        key = f"brush_{color_hex.lower()}"
        if key in cls._cache:
            return cls._cache[key]
        pixmap = QPixmap(28, 28)
        pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        brush_color = QColor(color_hex)
        if not brush_color.isValid():
            brush_color = QColor("#999999")
        painter.setBrush(QBrush(brush_color))
        painter.setPen(QPen(QColor(0, 0, 0, 140), 1.4))
        painter.drawEllipse(5, 6, 18, 18)
        painter.setPen(QPen(QColor(255, 255, 255, 230), 3, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap))
        painter.drawLine(9, 10, 18, 19)
        painter.setPen(QPen(QColor(0, 0, 0, 90), 2, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap))
        painter.drawLine(10, 9, 19, 18)
        painter.end()
        icon = QIcon(pixmap)
        cls._cache[key] = icon
        return icon

    @classmethod
    def get_icon(cls, name: str) -> QIcon:
        """返回缓存的图标，如果未缓存则即时加载。"""
        if name == "clear":
            name = "clear_all"  # 兼容旧配置
        if name in cls._cache:
            return cls._cache[name]
        data = cls._icons.get(name)
        if not data:
            return QIcon()
        try:
            pixmap = QPixmap()
            pixmap.loadFromData(QByteArray.fromBase64(data.encode("ascii")), "SVG")
            icon = QIcon(pixmap)
            cls._cache[name] = icon
            return icon
        except Exception:
            return QIcon()


# ---------- 可选依赖 ----------
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    pd = None
    PANDAS_AVAILABLE = False

PANDAS_READY = PANDAS_AVAILABLE and pd is not None

try:
    import openpyxl  # noqa: F401
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pyttsx3
    PYTTSX3_AVAILABLE = True
except ImportError:
    pyttsx3 = None
    PYTTSX3_AVAILABLE = False

try:
    import sounddevice as sd
    import numpy as np
    SOUNDDEVICE_AVAILABLE = True
except ImportError:
    sd = None
    np = None
    SOUNDDEVICE_AVAILABLE = False

if TYPE_CHECKING:
    from pandas import DataFrame as PandasDataFrame
else:  # pragma: no cover - runtime fallback for typing
    PandasDataFrame = Any  # type: ignore[misc, assignment]

if sys.platform == "win32":
    try:
        import winreg
        WINREG_AVAILABLE = True
    except ImportError:
        WINREG_AVAILABLE = False
else:
    WINREG_AVAILABLE = False


_SESSION_STUDENT_PASSWORD: Optional[str] = None
_SESSION_STUDENT_FILE_ENCRYPTED: bool = False


def _set_session_student_encryption(encrypted: bool, password: Optional[str]) -> None:
    global _SESSION_STUDENT_PASSWORD, _SESSION_STUDENT_FILE_ENCRYPTED
    _SESSION_STUDENT_FILE_ENCRYPTED = bool(encrypted)
    _SESSION_STUDENT_PASSWORD = password if encrypted else None


def _get_session_student_encryption() -> tuple[bool, Optional[str]]:
    return _SESSION_STUDENT_FILE_ENCRYPTED, _SESSION_STUDENT_PASSWORD


# ---------- 缓存 ----------
_SPEECH_ENV_CACHE: tuple[float, str, List[str]] = (0.0, "", [])


# ---------- DPI ----------
def ensure_high_dpi_awareness() -> None:
    if sys.platform != "win32":
        return
    os.environ.setdefault("QT_ENABLE_HIGHDPI_SCALING", "1")
    os.environ.setdefault("QT_SCALE_FACTOR_ROUNDING_POLICY", "PassThrough")
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


# ---------- 工具 ----------
def geometry_to_text(widget: QWidget) -> str:
    """以不含边框的尺寸记录窗口几何信息，避免重复放大。"""

    frame = widget.frameGeometry()
    inner = widget.geometry()
    width = inner.width() or frame.width()
    height = inner.height() or frame.height()
    return f"{width}x{height}+{frame.x()}+{frame.y()}"


def apply_geometry_from_text(widget: QWidget, geometry: str) -> None:
    if not geometry:
        return
    parts = geometry.split("+")
    if len(parts) != 3:
        return
    size_part, x_str, y_str = parts
    if "x" not in size_part:
        return
    width_str, height_str = size_part.split("x", 1)
    try:
        width = int(width_str)
        height = int(height_str)
        x = int(x_str)
        y = int(y_str)
    except ValueError:
        return

    base_min_width = getattr(widget, "_base_minimum_width", widget.minimumWidth())
    base_min_height = getattr(widget, "_base_minimum_height", widget.minimumHeight())

    custom_min_width = getattr(widget, "_ensure_min_width", 160)
    custom_min_height = getattr(widget, "_ensure_min_height", 120)

    min_width = max(base_min_width, custom_min_width)
    min_height = max(base_min_height, custom_min_height)

    screen = QApplication.screenAt(QPoint(x, y))
    if screen is None:
        try:
            screen = widget.screen() or QApplication.primaryScreen()
        except Exception:
            screen = QApplication.primaryScreen()
    if screen is not None:
        available = screen.availableGeometry()
        max_width = max(min_width, 320, int(available.width() * 0.9))
        max_height = max(min_height, 240, int(available.height() * 0.9))
        width = max(min_width, min(width, max_width))
        height = max(min_height, min(height, max_height))
        x = max(available.left(), min(x, available.right() - width))
        y = max(available.top(), min(y, available.bottom() - height))
    target_width = max(min_width, width)
    target_height = max(min_height, height)
    widget.resize(target_width, target_height)
    widget.move(x, y)


def ensure_widget_within_screen(widget: QWidget) -> None:
    screen = None
    try:
        screen = widget.screen()
    except Exception:
        screen = None
    if screen is None:
        screen = QApplication.primaryScreen()
    if screen is None:
        return
    base_min_width = getattr(widget, "_base_minimum_width", widget.minimumWidth())
    base_min_height = getattr(widget, "_base_minimum_height", widget.minimumHeight())

    custom_min_width = getattr(widget, "_ensure_min_width", 160)
    custom_min_height = getattr(widget, "_ensure_min_height", 120)

    min_width = max(base_min_width, custom_min_width)
    min_height = max(base_min_height, custom_min_height)

    available = screen.availableGeometry()
    geom = widget.frameGeometry()
    width = widget.width() or geom.width() or widget.sizeHint().width()
    height = widget.height() or geom.height() or widget.sizeHint().height()
    max_width = min(available.width(), max(min_width, int(available.width() * 0.9)))
    max_height = min(available.height(), max(min_height, int(available.height() * 0.9)))
    width = max(min_width, min(width, max_width))
    height = max(min_height, min(height, max_height))
    left_limit = available.x()
    top_limit = available.y()
    right_limit = max(left_limit, available.x() + available.width() - width)
    bottom_limit = max(top_limit, available.y() + available.height() - height)
    x = geom.x() if geom.width() else widget.x()
    y = geom.y() if geom.height() else widget.y()
    x = max(left_limit, min(x, right_limit))
    y = max(top_limit, min(y, bottom_limit))
    widget.resize(width, height)
    widget.move(x, y)


def str_to_bool(value: str, default: bool = False) -> bool:
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "on"}
    return default


def bool_to_str(value: bool) -> str:
    return "True" if value else "False"


def dedupe_strings(values: List[str]) -> List[str]:
    seen: set[str] = set()
    unique: List[str] = []
    for value in values:
        normalized = value.strip() if isinstance(value, str) else ""
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        unique.append(normalized)
    return unique


def _try_import_module(module: str) -> bool:
    try:
        importlib.import_module(module)
        return True
    except Exception:
        return False


def _count_windows_voice_tokens() -> tuple[int, Optional[str]]:
    if not WINREG_AVAILABLE:
        return -1, "无法访问 Windows 注册表"
    token_names: set[str] = set()
    path = r"SOFTWARE\\Microsoft\\Speech\\Voices\\Tokens"
    flags = {0}
    for attr in ("KEY_WOW64_32KEY", "KEY_WOW64_64KEY"):
        flag = getattr(winreg, attr, 0) if WINREG_AVAILABLE else 0  # type: ignore[name-defined]
        if flag:
            flags.add(flag)
    try:
        for hive in (winreg.HKEY_LOCAL_MACHINE, winreg.HKEY_CURRENT_USER):  # type: ignore[name-defined]
            for flag in flags:
                access = getattr(winreg, "KEY_READ", 0) | flag  # type: ignore[name-defined]
                try:
                    handle = winreg.OpenKey(hive, path, 0, access)  # type: ignore[name-defined]
                except FileNotFoundError:
                    continue
                except OSError as exc:
                    return 0, str(exc)
                with contextlib.closing(handle):
                    index = 0
                    while True:
                        try:
                            name = winreg.EnumKey(handle, index)  # type: ignore[name-defined]
                        except OSError:
                            break
                        token_names.add(str(name))
                        index += 1
    except Exception as exc:
        return 0, str(exc)
    return len(token_names), None


def _find_powershell_executable() -> Optional[str]:
    if sys.platform != "win32":
        return None
    path = shutil.which("pwsh") or shutil.which("powershell")
    if path:
        return path
    system_root = os.environ.get("SystemRoot") or os.environ.get("WINDIR")
    candidate_paths: List[str] = []
    if system_root:
        candidate_paths.extend(
            [
                os.path.join(system_root, "System32", "WindowsPowerShell", "v1.0", "pwsh.exe"),
                os.path.join(system_root, "System32", "WindowsPowerShell", "v1.0", "powershell.exe"),
                os.path.join(system_root, "SysWOW64", "WindowsPowerShell", "v1.0", "powershell.exe"),
            ]
        )
    candidate_paths.extend(
        [
            os.path.join("C:\\Program Files\\PowerShell\\7", "pwsh.exe"),
            os.path.join("C:\\Program Files\\PowerShell\\6", "pwsh.exe"),
        ]
    )
    for candidate in candidate_paths:
        if candidate and os.path.exists(candidate):
            return candidate
    return None


def _probe_powershell_speech_runtime(executable: Optional[str]) -> tuple[bool, Optional[str]]:
    if sys.platform != "win32" or not executable:
        return True, None
    script = (
        "try { "
        "Add-Type -AssemblyName System.Speech; "
        "[void][System.Speech.Synthesis.SpeechSynthesizer]::new().GetInstalledVoices().Count; "
        "\"OK\" } catch { $_.Exception.Message }"
    )
    startupinfo = None
    if os.name == "nt":
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    try:
        result = subprocess.run(
            [executable, "-NoLogo", "-NonInteractive", "-NoProfile", "-Command", script],
            capture_output=True,
            text=True,
            timeout=8,
            startupinfo=startupinfo,
        )
    except Exception as exc:
        return False, str(exc)
    output = (result.stdout or "").strip()
    if result.returncode != 0:
        message = output or (result.stderr or "").strip()
        return False, message or f"PowerShell exited with code {result.returncode}"
    if "OK" in output:
        return True, None
    if output:
        return False, output
    return True, None


def _detect_pyttsx3_driver_issue() -> Optional[str]:
    if pyttsx3 is None or sys.platform != "win32":
        return None
    try:
        drivers_spec = importlib.util.find_spec("pyttsx3.drivers")
        sapi_spec = importlib.util.find_spec("pyttsx3.drivers.sapi5")
    except Exception:
        return None
    if drivers_spec is None or sapi_spec is None:
        return "pyttsx3 �Ĳ��� sapi5 ����û���ý���ɵط���ϵͳ�޷�ʹ�� pyttsx3 ��������"
    return None


def detect_speech_environment_issues(
    force_refresh: bool = False,
    cache_seconds: float = 30.0,
) -> tuple[str, List[str]]:
    global _SPEECH_ENV_CACHE
    now = time.time()
    cached_at, cached_reason, cached_suggestions = _SPEECH_ENV_CACHE
    if not force_refresh and cached_at and now - cached_at < cache_seconds:
        return cached_reason, list(cached_suggestions)

    issues: List[str] = []
    suggestions: List[str] = []
    if sys.platform == "win32":
        missing: List[str] = []
        module_hints = (
            ("pyttsx3", "请安装 pyttsx3（命令：pip install pyttsx3）"),
            ("comtypes.client", "请安装 comtypes（命令：pip install comtypes）"),
            ("win32com.client", "请安装 pywin32（命令：pip install pywin32）"),
        )
        for module_name, hint in module_hints:
            if not _try_import_module(module_name):
                base_name = module_name.split(".")[0]
                if base_name not in missing:
                    missing.append(base_name)
                if hint not in suggestions:
                    suggestions.append(hint)
        if missing:
            issues.append(f"缺少依赖包{'、'.join(sorted(missing))}")
        token_count, token_error = _count_windows_voice_tokens()
        if token_error:
            issues.append(f"无法读取语音库信息：{token_error}")
        elif token_count == 0:
            issues.append("系统未检测到任何语音包")
            suggestions.append("请在 Windows 设置 -> 时间和语言 -> 语音 中下载并启用语音包")
        driver_issue = _detect_pyttsx3_driver_issue()
        if driver_issue:
            issues.append(driver_issue)
            suggestions.append("请确认 pyttsx3 的 SAPI5 驱动已随程序打包，或在命令提示窗运行：python -m pip install pyttsx3 comtypes pywin32")
        powershell_path = _find_powershell_executable()
        if not powershell_path:
            issues.append("未检测到 PowerShell，可用语音回退不可用")
            suggestions.append("请确保系统安装了 PowerShell 5+ 或 PowerShell 7，并能在 PATH 中访问")
        else:
            ps_ok, ps_reason = _probe_powershell_speech_runtime(powershell_path)
            if not ps_ok:
                detail = (ps_reason or "").strip()
                if detail:
                    issues.append(f"PowerShell 初始化语音失败：{detail}")
                else:
                    issues.append("PowerShell 初始化语音失败")
                suggestions.append("请在 PowerShell 中执行 Add-Type -AssemblyName System.Speech 检查错误，必要时启用 RemoteSigned 策略并安装最新 .NET 组件")
        if getattr(sys, "frozen", False):
            suggestions.append("如使用打包版，请确保 pyttsx3、comtypes、pywin32 被包含或在目标机器单独安装")
        suggestions.append("建议在命令提示窗执行：python -m pip install pyttsx3 comtypes pywin32（需管理员权限）")
        suggestions.append("若依旧失败，可尝试以管理员身份首次启动并重启系统")
    else:
        suggestions.append("请确认系统已安装可用的语音引擎（如 espeak 或系统自带语音）。")
    reason = "；".join(issues)
    deduped = dedupe_strings(suggestions)
    _SPEECH_ENV_CACHE = (now, reason, list(deduped))
    return reason, list(deduped)


class QuietInfoPopup(QWidget):
    """提供一个静音的小型提示窗口，避免系统提示音干扰课堂。"""

    _active_popups: List["QuietInfoPopup"] = []

    def __init__(self, parent: Optional[QWidget], text: str, title: str) -> None:
        flags = (
            Qt.WindowType.Tool
            | Qt.WindowType.WindowTitleHint
            | Qt.WindowType.WindowCloseButtonHint
            | Qt.WindowType.CustomizeWindowHint
        )
        super().__init__(parent, flags)
        self.setWindowTitle(title)
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, True)
        self.setWindowModality(Qt.WindowModality.ApplicationModal)
        self.setMinimumWidth(240)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(18, 18, 18, 12)
        layout.setSpacing(12)

        self.message_label = QLabel(text, self)
        self.message_label.setWordWrap(True)
        self.message_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
        layout.addWidget(self.message_label)

        button_row = QHBoxLayout()
        button_row.addStretch(1)
        self.ok_button = QPushButton("确定", self)
        self.ok_button.setDefault(True)
        apply_button_style(
            self.ok_button,
            ButtonStyles.PRIMARY,
            height=recommended_control_height(self.ok_button.font(), extra=14, minimum=36),
        )
        self.ok_button.clicked.connect(self.close)
        button_row.addWidget(self.ok_button)
        layout.addLayout(button_row)

        QuietInfoPopup._active_popups.append(self)
        self.destroyed.connect(self._cleanup)

    def showEvent(self, event) -> None:  # type: ignore[override]
        super().showEvent(event)
        self.adjustSize()
        self._relocate()
        self.activateWindow()
        self.ok_button.setFocus(Qt.FocusReason.ActiveWindowFocusReason)

    def _relocate(self) -> None:
        target_rect: Optional[QRect] = None
        parent = self.parentWidget()
        if parent and parent.isVisible():
            target_rect = parent.frameGeometry()
        else:
            screen = QApplication.primaryScreen()
            if screen:
                target_rect = screen.availableGeometry()
        if not target_rect:
            return
        geo = self.frameGeometry()
        geo.moveCenter(target_rect.center())
        self.move(geo.topLeft())

    def _cleanup(self, *_args) -> None:
        try:
            QuietInfoPopup._active_popups.remove(self)
        except ValueError:
            pass


def show_quiet_information(parent: Optional[QWidget], text: str, title: str = "提示") -> None:
    popup = QuietInfoPopup(parent, text, title)
    popup.show()


class QuietQuestionDialog(QDialog):
    """静音确认对话框，避免系统默认的提示音。"""

    def __init__(self, parent: Optional[QWidget], text: str, title: str) -> None:
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.setWindowFlag(Qt.WindowType.WindowStaysOnTopHint, True)
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, True)
        self.setMinimumWidth(320)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 18, 24, 18)
        layout.setSpacing(12)

        label = QLabel(text, self)
        label.setWordWrap(True)
        label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        layout.addWidget(label)

        buttons = QHBoxLayout()
        buttons.addStretch(1)

        cancel = QPushButton("取消", self)
        control_height = recommended_control_height(cancel.font(), extra=14, minimum=36)
        apply_button_style(cancel, ButtonStyles.TOOLBAR, height=control_height)
        cancel.clicked.connect(self.reject)
        buttons.addWidget(cancel)

        ok = QPushButton("确定", self)
        ok.setDefault(True)
        apply_button_style(ok, ButtonStyles.PRIMARY, height=control_height)
        ok.clicked.connect(self.accept)
        buttons.addWidget(ok)

        target_width = max(cancel.sizeHint().width(), ok.sizeHint().width())
        cancel.setFixedWidth(target_width)
        ok.setFixedWidth(target_width)

        layout.addLayout(buttons)
        self._ok_button = ok

    def showEvent(self, event) -> None:  # type: ignore[override]
        super().showEvent(event)
        self._ok_button.setFocus(Qt.FocusReason.ActiveWindowFocusReason)


class PasswordPromptDialog(QDialog):
    """统一样式的密码输入窗口，确保按钮始终可见且尺寸一致。"""

    def __init__(self, parent: Optional[QWidget], title: str, prompt: str) -> None:
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, True)
        self.setWindowFlag(Qt.WindowType.WindowStaysOnTopHint, True)
        self._captured_text: str = ""

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 18, 24, 18)
        layout.setSpacing(12)

        label = QLabel(prompt, self)
        label.setWordWrap(True)
        label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        layout.addWidget(label)

        self.line_edit = QLineEdit(self)
        self.line_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.line_edit.setMinimumWidth(220)
        layout.addWidget(self.line_edit)

        self.error_label = QLabel("", self)
        self.error_label.setWordWrap(True)
        self.error_label.setObjectName("passwordPromptErrorLabel")
        self.error_label.setStyleSheet(
            "#passwordPromptErrorLabel { color: #d93025; font-size: 12px; margin-top: 4px; }"
        )
        self.error_label.hide()
        layout.addWidget(self.error_label)

        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel,
            Qt.Orientation.Horizontal,
            self,
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        style_dialog_buttons(
            button_box,
            {
                QDialogButtonBox.StandardButton.Ok: ButtonStyles.PRIMARY,
                QDialogButtonBox.StandardButton.Cancel: ButtonStyles.TOOLBAR,
            },
            extra_padding=12,
            minimum_height=34,
            uniform_width=True,
        )
        layout.addWidget(button_box)

        self.line_edit.returnPressed.connect(self.accept)
        self.line_edit.textChanged.connect(self._clear_error)

    @classmethod
    def get_password(
        cls,
        parent: Optional[QWidget],
        title: str,
        prompt: str,
        *,
        allow_empty: bool = True,
    ) -> tuple[str, bool]:
        dialog = cls(parent, title, prompt)
        dialog._allow_empty = allow_empty  # type: ignore[attr-defined]
        accepted = dialog.exec() == QDialog.DialogCode.Accepted
        value = dialog._captured_text if accepted else ""
        return value, accepted

    def showEvent(self, event) -> None:  # type: ignore[override]
        super().showEvent(event)
        self.line_edit.setFocus(Qt.FocusReason.ActiveWindowFocusReason)

    def accept(self) -> None:  # type: ignore[override]
        text = self.line_edit.text()
        if not getattr(self, "_allow_empty", True) and not text.strip():
            self._show_error("密码不能为空，请重新输入。")
            return
        self._captured_text = text
        super().accept()

    def reject(self) -> None:  # type: ignore[override]
        self._captured_text = ""
        super().reject()

    def _show_error(self, message: str) -> None:
        self.error_label.setText(message)
        self.error_label.show()
        self.line_edit.setFocus(Qt.FocusReason.ActiveWindowFocusReason)

    def _clear_error(self, _: str) -> None:
        if self.error_label.isVisible():
            self.error_label.hide()


class PasswordSetupDialog(QDialog):
    """一次性采集两次输入的新密码，避免外部循环引发界面阻塞。"""

    def __init__(self, parent: Optional[QWidget], title: str, prompt: str, confirm_prompt: str) -> None:
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, True)
        self.setWindowFlag(Qt.WindowType.WindowStaysOnTopHint, True)
        self._password: str = ""

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 18, 24, 18)
        layout.setSpacing(12)

        prompt_label = QLabel(prompt, self)
        prompt_label.setWordWrap(True)
        prompt_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        layout.addWidget(prompt_label)

        self.password_edit = QLineEdit(self)
        self.password_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_edit.setMinimumWidth(220)
        layout.addWidget(self.password_edit)

        confirm_label = QLabel(confirm_prompt, self)
        confirm_label.setWordWrap(True)
        confirm_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        layout.addWidget(confirm_label)

        self.confirm_edit = QLineEdit(self)
        self.confirm_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.confirm_edit.setMinimumWidth(220)
        layout.addWidget(self.confirm_edit)

        self.error_label = QLabel("", self)
        self.error_label.setWordWrap(True)
        self.error_label.setObjectName("passwordSetupErrorLabel")
        self.error_label.setStyleSheet(
            "#passwordSetupErrorLabel { color: #d93025; font-size: 12px; margin-top: 4px; }"
        )
        self.error_label.hide()
        layout.addWidget(self.error_label)

        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel,
            Qt.Orientation.Horizontal,
            self,
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        style_dialog_buttons(
            button_box,
            {
                QDialogButtonBox.StandardButton.Ok: ButtonStyles.PRIMARY,
                QDialogButtonBox.StandardButton.Cancel: ButtonStyles.TOOLBAR,
            },
            extra_padding=12,
            minimum_height=34,
            uniform_width=True,
        )
        layout.addWidget(button_box)

        self.password_edit.returnPressed.connect(self._focus_confirm)
        self.confirm_edit.returnPressed.connect(self.accept)
        self.password_edit.textChanged.connect(self._clear_error)
        self.confirm_edit.textChanged.connect(self._clear_error)

    @classmethod
    def get_new_password(
        cls,
        parent: Optional[QWidget],
        title: str,
        prompt: str,
        confirm_prompt: str,
    ) -> tuple[str, bool]:
        dialog = cls(parent, title, prompt, confirm_prompt)
        accepted = dialog.exec() == QDialog.DialogCode.Accepted
        value = dialog._password if accepted else ""
        return value, accepted

    def showEvent(self, event) -> None:  # type: ignore[override]
        super().showEvent(event)
        self.password_edit.setFocus(Qt.FocusReason.ActiveWindowFocusReason)

    def accept(self) -> None:  # type: ignore[override]
        password = self.password_edit.text().strip()
        confirm = self.confirm_edit.text().strip()
        if not password:
            self._show_error("密码不能为空，请重新输入。", focus_widget=self.password_edit)
            return
        if password != confirm:
            self._show_error("两次输入的密码不一致，请重新设置。", focus_widget=self.confirm_edit)
            return
        self._password = password
        super().accept()

    def reject(self) -> None:  # type: ignore[override]
        self._password = ""
        super().reject()

    def _show_error(self, message: str, *, focus_widget: Optional[QWidget] = None) -> None:
        self.error_label.setText(message)
        self.error_label.show()
        target = focus_widget or self.password_edit
        target.setFocus(Qt.FocusReason.ActiveWindowFocusReason)

    def _clear_error(self, _: str) -> None:
        if self.error_label.isVisible():
            self.error_label.hide()

    def _focus_confirm(self) -> None:
        self.confirm_edit.setFocus(Qt.FocusReason.TabFocusReason)


def ask_quiet_confirmation(parent: Optional[QWidget], text: str, title: str = "确认") -> bool:
    dialog = QuietQuestionDialog(parent, text, title)
    return dialog.exec() == QDialog.DialogCode.Accepted


def recommended_control_height(font: QFont, *, extra: int = 12, minimum: int = 32) -> int:
    """Return a DPI-aware button height based on the supplied font metrics."""

    metrics = QFontMetrics(font)
    text_height = metrics.boundingRect("Ag").height()
    line_height = metrics.height()
    base_height = max(text_height, line_height)
    return max(minimum, int(math.ceil(base_height + extra)))


class ButtonStyles:
    """Centralised QPushButton样式，避免各窗口重复定义造成视觉不一致。"""

    TOOLBAR = (
        "QPushButton {\n"
        "    padding: 4px 12px;\n"
        "    border-radius: 12px;\n"
        "    border: 1px solid #c4c8d0;\n"
        "    background-color: #ffffff;\n"
        "    color: #202124;\n"
        "}\n"
        "QPushButton:disabled {\n"
        "    color: rgba(32, 33, 36, 0.45);\n"
        "    background-color: #f3f5f9;\n"
        "    border-color: #d9dde3;\n"
        "}\n"
        "QPushButton:hover {\n"
        "    border-color: #1a73e8;\n"
        "    background-color: #eaf2ff;\n"
        "}\n"
        "QPushButton:pressed {\n"
        "    background-color: #d7e7ff;\n"
        "}\n"
        "QPushButton:checked {\n"
        "    background-color: #1a73e8;\n"
        "    border-color: #1a73e8;\n"
        "    color: #ffffff;\n"
        "}\n"
    )

    GRID = (
        "QPushButton {\n"
        "    padding: 6px 12px;\n"
        "    border-radius: 10px;\n"
        "    border: 1px solid #c4c8d0;\n"
        "    background-color: #ffffff;\n"
        "    color: #202124;\n"
        "}\n"
        "QPushButton:hover {\n"
        "    border-color: #1a73e8;\n"
        "    background-color: #eaf2ff;\n"
        "}\n"
        "QPushButton:pressed {\n"
        "    background-color: #d7e7ff;\n"
        "}\n"
    )

    PRIMARY = (
        "QPushButton {\n"
        "    padding: 6px 20px;\n"
        "    border-radius: 20px;\n"
        "    background-color: #1a73e8;\n"
        "    color: #ffffff;\n"
        "    border: 1px solid #1a73e8;\n"
        "}\n"
        "QPushButton:hover {\n"
        "    background-color: #185abc;\n"
        "    border-color: #185abc;\n"
        "}\n"
        "QPushButton:pressed {\n"
        "    background-color: #174ea6;\n"
        "}\n"
        "QPushButton:disabled {\n"
        "    background-color: #aac6ff;\n"
        "    border-color: #aac6ff;\n"
        "    color: rgba(255, 255, 255, 0.8);\n"
        "}\n"
    )

    ORDER_TOGGLE = (
        "QPushButton {\n"
        "    padding: 4px 18px;\n"
        "    border-radius: 22px;\n"
        "    border: 1px solid rgba(16, 61, 115, 0.28);\n"
        "    background-color: rgba(255, 255, 255, 0.96);\n"
        "    color: #0b3d91;\n"
        "}\n"
        "QPushButton:hover {\n"
        "    border-color: #1a73e8;\n"
        "    background-color: rgba(26, 115, 232, 0.16);\n"
        "}\n"
        "QPushButton:checked {\n"
        "    background-color: #1a73e8;\n"
        "    border-color: #1a73e8;\n"
        "    color: #ffffff;\n"
        "}\n"
    )


def apply_button_style(button: QPushButton, style: str, *, height: Optional[int] = None) -> None:
    """Apply a reusable QPushButton stylesheet and pointer cursor."""

    button.setCursor(Qt.CursorShape.PointingHandCursor)
    if height is not None:
        button.setMinimumHeight(height)
        button.setMaximumHeight(height)
        button.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Fixed)
    button.setStyleSheet(style)


def style_dialog_buttons(
    button_box: QDialogButtonBox,
    styles: Mapping[QDialogButtonBox.StandardButton, str],
    *,
    extra_padding: int = 10,
    minimum_height: int = 34,
    uniform_width: bool = False,
) -> None:
    """Apply shared styling to all buttons contained in a QDialogButtonBox."""

    styled_buttons: list[QPushButton] = []
    for standard_button, style in styles.items():
        button = button_box.button(standard_button)
        if button is None:
            continue
        control_height = recommended_control_height(
            button.font(), extra=extra_padding, minimum=minimum_height
        )
        apply_button_style(button, style, height=control_height)
        styled_buttons.append(button)
    if uniform_width and styled_buttons:
        target_width = max(button.sizeHint().width() for button in styled_buttons)
        for button in styled_buttons:
            button.setFixedWidth(target_width)


# ---------- 配置 ----------
class SettingsManager:
    """负责读取/写入配置文件的轻量封装。"""

    def __init__(self, filename: str = "settings.ini") -> None:
        # 统一维护配置文件的存放路径，优先使用用户配置目录，保证跨次启动仍能读取到历史点名状态。
        self._mirror_targets: set[str] = set()
        self.filename = self._prepare_storage_path(filename)
        self._mirror_targets.add(self.filename)
        self.config = configparser.ConfigParser()
        self.config.optionxform = str
        self._settings_cache: Optional[Dict[str, Dict[str, str]]] = None
        self.defaults: Dict[str, Dict[str, str]] = {
            "Launcher": {
                "x": "120",
                "y": "120",
                "minimized": "False",
                "bubble_x": "120",
                "bubble_y": "120",
            },
            "Startup": {"autostart_enabled": "False"},
            "RollCallTimer": {
                "geometry": "480x280+180+180",
                "show_id": "True",
                "show_name": "True",
                "show_photo": "False",
                "photo_duration_seconds": "0",
                "speech_enabled": "False",
                "speech_voice_id": "",
                "current_group": "全部",
                "timer_countdown_minutes": "5",
                "timer_countdown_seconds": "0",
                "timer_sound_enabled": "True",
                "mode": "roll_call",
                "timer_mode": "countdown",
                "timer_seconds_left": "300",
                "timer_stopwatch_seconds": "0",
                "timer_running": "False",
                "id_font_size": "48",
                "name_font_size": "60",
                "timer_font_size": "56",
                "scoreboard_order": "rank",
            },
            "Paint": {"x": "260", "y": "260", "brush_size": "12", "brush_color": "#ff0000"},
        }

    def _get_config_dir(self) -> str:
        """返回当前系统下建议的配置目录。"""

        home = os.path.expanduser("~")
        if sys.platform.startswith("win"):
            base = os.environ.get("APPDATA") or os.path.join(home, "AppData", "Roaming")
            return os.path.join(base, "ClassroomTools")
        if sys.platform == "darwin":
            return os.path.join(home, "Library", "Application Support", "ClassroomTools")
        base = os.environ.get("XDG_CONFIG_HOME") or os.path.join(home, ".config")
        return os.path.join(base, "ClassroomTools")

    def _prepare_storage_path(self, filename: str) -> str:
        """根据平台选择合适的设置文件路径，并在需要时迁移旧文件。"""

        base_name = os.path.basename(filename) or "settings.ini"
        legacy_path = os.path.abspath(filename)
        config_dir = self._get_config_dir()
        try:
            os.makedirs(config_dir, exist_ok=True)
        except OSError:
            config_dir = os.path.dirname(legacy_path) or os.getcwd()
        target_path = os.path.join(config_dir, base_name)

        if os.path.exists(legacy_path):
            same_file = False
            try:
                same_file = os.path.samefile(legacy_path, target_path)
            except (OSError, FileNotFoundError):
                same_file = False
            if not same_file and not os.path.exists(target_path):
                try:
                    shutil.copy2(legacy_path, target_path)
                except Exception:
                    target_path = legacy_path
            if not same_file:
                legacy_dir = os.path.dirname(legacy_path) or os.getcwd()
                if os.access(legacy_dir, os.W_OK):
                    self._mirror_targets.add(legacy_path)

        try:
            with open(target_path, "a", encoding="utf-8"):
                pass
        except OSError:
            target_path = legacy_path

        return target_path

    def get_defaults(self) -> Dict[str, Dict[str, str]]:
        return {section: values.copy() for section, values in self.defaults.items()}

    def load_settings(self) -> Dict[str, Dict[str, str]]:
        if self._settings_cache is not None:
            return {section: values.copy() for section, values in self._settings_cache.items()}

        settings = self.get_defaults()
        if os.path.exists(self.filename):
            try:
                self.config.read(self.filename, encoding="utf-8")
                for section in self.config.sections():
                    if section not in settings:
                        settings[section] = {}
                    for key, value in self.config.items(section):
                        settings[section][key] = value
            except configparser.Error:
                settings = self.get_defaults()

        self._settings_cache = {section: values.copy() for section, values in settings.items()}
        return {section: values.copy() for section, values in self._settings_cache.items()}

    def _write_atomic(self, path: str, data: str) -> None:
        directory = os.path.dirname(path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)
        target_dir = directory or os.getcwd()
        fd, tmp_path = tempfile.mkstemp(prefix="ctools_", suffix=".tmp", dir=target_dir)
        try:
            with os.fdopen(fd, "w", encoding="utf-8") as handle:
                handle.write(data)
            os.replace(tmp_path, path)
        except Exception:
            try:
                os.remove(tmp_path)
            except OSError:
                pass
            raise

    def save_settings(self, settings: Dict[str, Dict[str, str]]) -> None:
        config = configparser.ConfigParser()
        config.optionxform = str
        snapshot: Dict[str, Dict[str, str]] = {}
        for section, options in settings.items():
            snapshot[section] = {key: str(value) for key, value in options.items()}
            config[section] = snapshot[section]

        buffer = io.StringIO()
        config.write(buffer)
        data = buffer.getvalue()
        buffer.close()

        failed: List[str] = []
        for path in sorted(self._mirror_targets):
            try:
                self._write_atomic(path, data)
            except Exception:
                failed.append(path)
        if failed and self.filename not in failed:
            print(f"无法写入备用配置文件: {failed}")

        self._settings_cache = {section: values.copy() for section, values in snapshot.items()}

    def clear_roll_call_history(self) -> None:
        """清除点名历史信息，仅在用户主动重置时调用。"""

        settings = self.load_settings()
        section = settings.get("RollCallTimer", {})
        removed = False
        for key in ("group_remaining", "group_last", "global_drawn"):
            if key in section:
                section.pop(key, None)
                removed = True
        if removed:
            settings["RollCallTimer"] = section
            self.save_settings(settings)


# ---------- 自绘置顶 ToolTip ----------
class TipWindow(QWidget):
    """一个轻量的自绘 ToolTip，确保位于所有置顶窗之上。"""
    def __init__(self) -> None:
        super().__init__(None, Qt.WindowType.Tool | Qt.WindowType.FramelessWindowHint | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, True)
        self.setAttribute(Qt.WidgetAttribute.WA_ShowWithoutActivating, True)
        self.setStyleSheet(
            """
            QLabel {
                color: #f1f3f4;
                background: rgba(32, 33, 36, 230);
                border: 1px solid rgba(255, 255, 255, 45);
                border-radius: 6px;
                padding: 4px 8px;
                font-size: 12px;
            }
            """
        )
        self._label = QLabel("", self)
        self._label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self._hide_timer = QTimer(self); self._hide_timer.setSingleShot(True); self._hide_timer.timeout.connect(self.hide)

    def show_tip(self, text: str, pos: QPoint, duration_ms: int = 2500) -> None:
        self._label.setText(text or "")
        self._label.adjustSize()
        self.resize(self._label.size())
        self.move(pos + QPoint(12, 16))
        self.show()
        self.raise_()
        self._hide_timer.start(duration_ms)

    def hide_tip(self) -> None:
        self._hide_timer.stop()
        self.hide()


# ---------- 对话框 ----------
class PenSettingsDialog(QDialog):
    """画笔粗细与颜色选择对话框。"""
    COLORS = {
        "#FFFF00": "黄",
        "#FFA500": "橙",
        "#24B47E": "绿",
        "#800080": "紫",
        "#FFFFFF": "白",
    }

    def __init__(self, parent: Optional[QWidget] = None, initial_size: int = 12, initial_color: str = "#FF0000") -> None:
        super().__init__(parent)
        self.setWindowTitle("画笔设置")
        self.setWindowFlag(Qt.WindowType.WindowStaysOnTopHint, True)
        self.pen_color = QColor(initial_color)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        size_layout = QHBoxLayout()
        size_label = QLabel("粗细:")
        self.size_slider = QSlider(Qt.Orientation.Horizontal)
        self.size_slider.setRange(2, 40)
        self.size_slider.setValue(int(initial_size))
        self.size_slider.setMinimumWidth(120)
        self.size_value = QLabel(str(initial_size))
        self.size_slider.valueChanged.connect(lambda value: self.size_value.setText(str(value)))

        size_layout.addWidget(size_label)
        size_layout.addWidget(self.size_slider, 1)
        size_layout.addWidget(self.size_value)
        layout.addLayout(size_layout)

        layout.addWidget(QLabel("颜色:"))
        color_layout = QGridLayout()
        color_layout.setContentsMargins(0, 0, 0, 0)
        color_layout.setSpacing(6)
        for index, (color_hex, name) in enumerate(self.COLORS.items()):
            button = QPushButton()
            button.setFixedSize(24, 24)
            button.setCursor(Qt.CursorShape.PointingHandCursor)
            button.setStyleSheet(
                f"background-color: {color_hex}; border: 1px solid rgba(0, 0, 0, 60); border-radius: 12px;"
            )
            button.setToolTip(name)
            button.clicked.connect(lambda _checked, c=color_hex: self._select_color(c))
            color_layout.addWidget(button, index // 3, index % 3)
        layout.addLayout(color_layout)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        style_dialog_buttons(
            buttons,
            {
                QDialogButtonBox.StandardButton.Ok: ButtonStyles.PRIMARY,
                QDialogButtonBox.StandardButton.Cancel: ButtonStyles.TOOLBAR,
            },
            extra_padding=10,
            minimum_height=32,
        )
        layout.addWidget(buttons)

        self.setFixedSize(self.sizeHint())

    def _select_color(self, color_hex: str) -> None:
        self.pen_color = QColor(color_hex)

    def get_settings(self) -> tuple[int, QColor]:
        return self.size_slider.value(), self.pen_color

    def showEvent(self, event) -> None:  # type: ignore[override]
        super().showEvent(event)
        ensure_widget_within_screen(self)


class ShapeSettingsDialog(QDialog):
    """图形工具的快捷选择窗口。"""
    SHAPES = {"line": "直线", "dashed_line": "虚线", "rect": "矩形", "circle": "圆形"}

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("选择图形")
        self.setWindowFlag(Qt.WindowType.WindowStaysOnTopHint, True)
        self.selected_shape: Optional[str] = None

        layout = QGridLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setHorizontalSpacing(6)
        layout.setVerticalSpacing(6)

        for index, (shape_key, name) in enumerate(self.SHAPES.items()):
            button = QPushButton(name)
            button.setMinimumWidth(68)
            apply_button_style(
                button,
                ButtonStyles.TOOLBAR,
                height=recommended_control_height(button.font(), extra=10, minimum=32),
            )
            button.clicked.connect(lambda _checked, key=shape_key: self._select_shape(key))
            layout.addWidget(button, index // 2, index % 2)

        self.setFixedSize(self.sizeHint())

    def _select_shape(self, shape: str) -> None:
        self.selected_shape = shape
        self.accept()

    def get_shape(self) -> Optional[str]:
        return self.selected_shape

    def showEvent(self, event) -> None:  # type: ignore[override]
        super().showEvent(event)
        ensure_widget_within_screen(self)


class BoardColorDialog(QDialog):
    """白板背景颜色选择对话框。"""
    COLORS = {"#FFFFFF": "白板", "#000000": "黑板", "#0E4020": "绿板"}

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("选择颜色")
        self.setWindowFlag(Qt.WindowType.WindowStaysOnTopHint, True)
        self.selected_color: Optional[QColor] = None

        layout = QHBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(6)

        for color_hex, name in self.COLORS.items():
            button = QPushButton(name)
            button.setCursor(Qt.CursorShape.PointingHandCursor)
            button.setFixedSize(72, 28)
            button.clicked.connect(lambda _checked, c=color_hex: self._select_color(c))
            layout.addWidget(button)

        self.setFixedSize(self.sizeHint())

    def _select_color(self, color_hex: str) -> None:
        self.selected_color = QColor(color_hex)
        self.accept()

    def get_color(self) -> Optional[QColor]:
        return self.selected_color

    def showEvent(self, event) -> None:  # type: ignore[override]
        super().showEvent(event)
        ensure_widget_within_screen(self)


# ---------- 标题栏 ----------
class TitleBar(QWidget):
    """浮动工具条的标题栏，负责拖拽移动。"""

    def __init__(self, toolbar: "FloatingToolbar") -> None:
        super().__init__(toolbar)
        self.toolbar = toolbar
        self._dragging = False
        self._drag_offset = QPoint()
        self.setCursor(Qt.CursorShape.OpenHandCursor)

        self.setAutoFillBackground(True)
        palette = self.palette()
        palette.setColor(self.backgroundRole(), QColor(36, 37, 41, 235))
        self.setPalette(palette)
        self.setFixedHeight(22)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(6, 0, 6, 0)
        title = QLabel("屏幕画笔")
        font = title.font()
        font.setBold(True)
        title.setFont(font)
        title.setStyleSheet("color: #f1f3f4; font-size: 10.5px;")
        layout.addWidget(title)
        layout.addStretch()

    def mousePressEvent(self, event) -> None:
        if event.button() == Qt.MouseButton.LeftButton:
            self._dragging = True
            self._drag_offset = event.globalPosition().toPoint() - self.toolbar.pos()
            self.setCursor(Qt.CursorShape.ClosedHandCursor)
            event.accept()
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event) -> None:
        if self._dragging:
            self.toolbar.move(event.globalPosition().toPoint() - self._drag_offset)
            event.accept()
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event) -> None:
        if self._dragging:
            self.toolbar.overlay.save_window_position()
        self._dragging = False
        self.setCursor(Qt.CursorShape.OpenHandCursor)
        try:
            self.toolbar.overlay.raise_toolbar()
        except Exception:
            pass
        event.accept()
        super().mouseReleaseEvent(event)


# ---------- 浮动工具条（画笔/白板） ----------
class FloatingToolbar(QWidget):
    """悬浮工具条：提供画笔、图形、白板等常用按钮。"""

    def __init__(self, overlay: "OverlayWindow", settings_manager: SettingsManager) -> None:
        super().__init__(
            None,
            Qt.WindowType.Tool | Qt.WindowType.FramelessWindowHint | Qt.WindowType.WindowStaysOnTopHint,
        )
        self.overlay = overlay
        self.settings_manager = settings_manager
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setAttribute(Qt.WidgetAttribute.WA_AlwaysShowToolTips, True)
        self.setWindowFlag(Qt.WindowType.WindowDoesNotAcceptFocus, True)
        self._tip = TipWindow()
        self._build_ui()

        settings = self.settings_manager.load_settings().get("Paint", {})
        self.move(int(settings.get("x", "260")), int(settings.get("y", "260")))
        self.adjustSize()
        self.setFixedSize(self.sizeHint())
        self._base_minimum_width = self.width()
        self._base_minimum_height = self.height()
        self._ensure_min_width = self.width()
        self._ensure_min_height = self.height()

    def _build_ui(self) -> None:
        self.setStyleSheet(
            """
            #container {
                background-color: rgba(28, 29, 32, 235);
                border-radius: 10px;
                border: 1px solid rgba(255, 255, 255, 45);
            }
            QPushButton {
                color: #f1f3f4;
                background: rgba(60, 64, 67, 240);
                border: 1px solid rgba(255, 255, 255, 45);
                border-radius: 6px;
                padding: 3px;
                min-width: 28px;
                min-height: 28px;
            }
            QPushButton:hover {
                background: rgba(138, 180, 248, 245);
                border-color: rgba(138, 180, 248, 255);
                color: #202124;
            }
            QPushButton:checked {
                background: rgba(138, 180, 248, 255);
                color: #202124;
            }
            QPushButton#eraserButton {
                background: rgba(241, 243, 244, 235);
                color: #3c4043;
                border-color: rgba(138, 180, 248, 170);
            }
            QPushButton#eraserButton:hover,
            QPushButton#eraserButton:checked {
                background: rgba(202, 225, 255, 255);
                border-color: #1a73e8;
                color: #174ea6;
            }
            QPushButton#clearButton {
                background: rgba(255, 236, 232, 240);
                color: #a03a1e;
                border-color: rgba(255, 173, 153, 230);
            }
            QPushButton#clearButton:hover,
            QPushButton#clearButton:checked {
                background: rgba(255, 210, 204, 255);
                border-color: rgba(255, 138, 101, 255);
                color: #5f2121;
            }
            #whiteboardButtonActive {
                background: rgba(255, 214, 102, 240);
                border-color: rgba(251, 188, 5, 255);
                color: #202124;
            }
            """
        )

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        container = QWidget(self)
        container.setObjectName("container")
        root.addWidget(container)

        layout = QVBoxLayout(container)
        layout.setContentsMargins(6, 5, 6, 6)
        layout.setSpacing(5)
        self.title_bar = TitleBar(self)
        layout.addWidget(self.title_bar)

        self.btn_cursor = QPushButton(IconManager.get_icon("cursor"), "")
        self.brush_color_buttons: Dict[str, QPushButton] = {}
        brush_configs = [
            ("#000000", "黑色画笔"),
            ("#FF0000", "红色画笔"),
            ("#1E90FF", "蓝色画笔"),
        ]  # 预设的高频画笔颜色
        brush_buttons = []
        for color_hex, name in brush_configs:
            button = QPushButton(IconManager.get_brush_icon(color_hex), "")
            button.setToolTip(name)
            self.brush_color_buttons[color_hex.lower()] = button
            brush_buttons.append(button)
        self.btn_settings = QPushButton(IconManager.get_icon("settings"), "")
        self.btn_shape = QPushButton(IconManager.get_icon("shape"), "")
        self.btn_slide_down = QPushButton(IconManager.get_icon("arrow_down"), "")
        self.btn_slide_up = QPushButton(IconManager.get_icon("arrow_up"), "")
        self.btn_undo = QPushButton(IconManager.get_icon("undo"), "")
        self.btn_eraser = QPushButton(IconManager.get_icon("eraser"), "")
        self.btn_eraser.setObjectName("eraserButton")
        self.btn_clear_all = QPushButton(IconManager.get_icon("clear_all"), "")
        self.btn_clear_all.setObjectName("clearButton")
        self.btn_whiteboard = QPushButton(IconManager.get_icon("whiteboard"), "")

        row_top = QHBoxLayout()
        row_top.setContentsMargins(0, 0, 0, 0)
        row_top.setSpacing(3)
        row_bottom = QHBoxLayout()
        row_bottom.setContentsMargins(0, 0, 0, 0)
        row_bottom.setSpacing(3)

        top_buttons = [
            self.btn_cursor,
            *brush_buttons,
            self.btn_settings,
            self.btn_shape,
        ]
        bottom_buttons = [
            self.btn_slide_down,
            self.btn_slide_up,
            self.btn_undo,
            self.btn_eraser,
            self.btn_clear_all,
            self.btn_whiteboard,
        ]

        def _configure_toolbar_button(btn: QPushButton) -> None:
            btn.setIconSize(QSize(18, 18))
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)

        for btn in top_buttons + bottom_buttons:
            _configure_toolbar_button(btn)
        for btn in top_buttons:
            row_top.addWidget(btn)
        for btn in bottom_buttons:
            row_bottom.addWidget(btn)
        layout.addLayout(row_top)
        layout.addLayout(row_bottom)

        tooltip_text = {
            self.btn_cursor: "光标",
            self.btn_slide_down: "下一页",
            self.btn_slide_up: "上一页",
            self.btn_shape: "图形",
            self.btn_undo: "撤销",
            self.btn_eraser: "橡皮（再次点击恢复画笔）",
            self.btn_clear_all: "清除（并恢复画笔）",
            self.btn_whiteboard: "白板（单击开关 / 双击换色）",
            self.btn_settings: "画笔设置",
        }
        for button in brush_buttons:
            tooltip_text[button] = button.toolTip() or "画笔"
        for btn, tip_text in tooltip_text.items():
            btn.setToolTip(tip_text)
            btn.installEventFilter(self)

        self.tool_buttons = QButtonGroup(self)
        for btn in (self.btn_cursor, *brush_buttons, self.btn_shape, self.btn_eraser):
            btn.setCheckable(True)
            self.tool_buttons.addButton(btn)
        self.tool_buttons.setExclusive(True)

        self.btn_cursor.clicked.connect(self.overlay.toggle_cursor_mode)
        for color_hex, button in zip([c for c, _ in brush_configs], brush_buttons):
            button.clicked.connect(lambda _checked, c=color_hex: self.overlay.use_brush_color(c))
        self.btn_shape.clicked.connect(self._select_shape)
        self.btn_slide_down.pressed.connect(lambda: self.overlay.handle_navigation_button_press("down"))
        self.btn_slide_down.released.connect(self.overlay.handle_navigation_button_release)
        self.btn_slide_up.pressed.connect(lambda: self.overlay.handle_navigation_button_press("up"))
        self.btn_slide_up.released.connect(self.overlay.handle_navigation_button_release)
        self.btn_undo.clicked.connect(self.overlay.undo_last_action)
        self.btn_eraser.clicked.connect(self.overlay.toggle_eraser_mode)
        self.btn_clear_all.clicked.connect(self.overlay.clear_all)
        self.btn_settings.clicked.connect(self.overlay.open_pen_settings)
        self.btn_whiteboard.clicked.connect(self._handle_whiteboard_click)

        self._wb_click_timer = QTimer(self)
        self._wb_click_timer.setInterval(QApplication.instance().doubleClickInterval())
        self._wb_click_timer.setSingleShot(True)
        # 使用单次定时器来区分白板按钮的单击与双击行为
        self._wb_click_timer.timeout.connect(self.overlay.toggle_whiteboard)

        self.btn_undo.setEnabled(False)

        for widget in (self, container, self.title_bar):
            widget.installEventFilter(self)

    def update_tool_states(
        self,
        mode: str,
        interaction_mode: str,
        pen_color: QColor,
        *,
        navigation_from_cursor: bool = False,
        highlight_mode: Optional[str] = None,
    ) -> None:
        color_key = pen_color.name().lower()
        active_mode: Optional[str] = mode
        if highlight_mode is not None:
            active_mode = highlight_mode
        elif interaction_mode == "cursor":
            active_mode = "cursor"
        elif interaction_mode == "navigation" and navigation_from_cursor:
            active_mode = ""
        elif interaction_mode == "navigation":
            active_mode = mode
        prev_exclusive = self.tool_buttons.exclusive()
        if prev_exclusive:
            self.tool_buttons.setExclusive(False)
        try:
            for hex_key, button in self.brush_color_buttons.items():
                prev = button.blockSignals(True)
                button.setChecked(active_mode == "brush" and hex_key == color_key)
                button.blockSignals(prev)
            for tool, button in (
                ("cursor", self.btn_cursor),
                ("shape", self.btn_shape),
                ("eraser", self.btn_eraser),
            ):
                prev = button.blockSignals(True)
                button.setChecked(active_mode == tool)
                button.blockSignals(prev)
            if active_mode == "brush" and color_key not in self.brush_color_buttons:
                for button in (self.btn_cursor, self.btn_shape, self.btn_eraser):
                    prev = button.blockSignals(True)
                    button.setChecked(False)
                    button.blockSignals(prev)
        finally:
            if prev_exclusive:
                self.tool_buttons.setExclusive(True)

    def update_undo_state(self, enabled: bool) -> None:
        self.btn_undo.setEnabled(enabled)

    def eventFilter(self, obj, event):
        event_type = event.type()
        if event_type == QEvent.Type.Wheel:
            wheel_event = cast(QWheelEvent, event)
            if self.overlay.handle_toolbar_navigation_wheel(wheel_event):
                return True
        if event_type == QEvent.Type.KeyPress:
            key_event = cast(QKeyEvent, event)
            if self.overlay.handle_toolbar_navigation_key_event(key_event, pressed=True):
                return True
        if event_type == QEvent.Type.KeyRelease:
            key_event = cast(QKeyEvent, event)
            if self.overlay.handle_toolbar_navigation_key_event(key_event, pressed=False):
                return True
        if isinstance(obj, QPushButton) and event_type == QEvent.Type.ToolTip:
            try:
                self.overlay.raise_toolbar()
            except Exception:
                pass
            self._tip.show_tip(obj.toolTip(), QCursor.pos())
            return True
        if event_type in (
            QEvent.Type.Leave,
            QEvent.Type.MouseButtonPress,
            QEvent.Type.MouseButtonDblClick,
        ):
            self._tip.hide_tip()
        return super().eventFilter(obj, event)

    def _select_shape(self) -> None:
        dialog = ShapeSettingsDialog(self)
        if dialog.exec():
            shape = dialog.get_shape()
            if shape:
                self.overlay.set_mode("shape", shape_type=shape)
        else:
            self.overlay.update_toolbar_state()

    def _handle_whiteboard_click(self) -> None:
        if self._wb_click_timer.isActive():
            self._wb_click_timer.stop()
            self.overlay.open_board_color_dialog()
        else:
            self._wb_click_timer.start()

    def update_whiteboard_button_state(self, active: bool) -> None:
        self.btn_whiteboard.setObjectName("whiteboardButtonActive" if active else "")
        self.style().polish(self.btn_whiteboard)

    def enterEvent(self, event) -> None:
        self.overlay.raise_toolbar()
        try:
            self.overlay.on_toolbar_mouse_enter()
        except Exception:
            pass
        super().enterEvent(event)

    def leaveEvent(self, event) -> None:
        super().leaveEvent(event)
        QTimer.singleShot(0, self.overlay.on_toolbar_mouse_leave)

    def showEvent(self, event) -> None:  # type: ignore[override]
        super().showEvent(event)
        ensure_widget_within_screen(self)


# ---------- 叠加层（画笔/白板） ----------


class _PresentationForwarder:
    """在绘图模式下将特定输入事件转发给下层演示窗口。"""

    __slots__ = (
        "overlay",
        "_last_target_hwnd",
        "_last_target_role",
        "_last_detected_role",
        "_child_buffer",
        "_self_pid",
        "_injecting_wheel",
    )

    _SMTO_ABORTIFHUNG = 0x0002
    _MAX_CHILD_FORWARDS = 32
    _INPUT_MOUSE = 0
    _INPUT_KEYBOARD = 1
    _KEYEVENTF_EXTENDEDKEY = 0x0001
    _KEYEVENTF_KEYUP = 0x0002

    if _USER32 is not None:

        class _GuiThreadInfo(ctypes.Structure):
            _fields_ = [
                ("cbSize", wintypes.DWORD),
                ("flags", wintypes.DWORD),
                ("hwndActive", wintypes.HWND),
                ("hwndFocus", wintypes.HWND),
                ("hwndCapture", wintypes.HWND),
                ("hwndMenuOwner", wintypes.HWND),
                ("hwndMoveSize", wintypes.HWND),
                ("hwndCaret", wintypes.HWND),
                ("rcCaret", wintypes.RECT),
            ]

        _KeyboardInput = None  # type: ignore[assignment]
        _InputUnion = None  # type: ignore[assignment]
        _Input = None  # type: ignore[assignment]
        try:

            _KeyboardInput = type(
                "_KeyboardInput",
                (ctypes.Structure,),
                {
                    "_fields_": [
                        ("wVk", wintypes.WORD),
                        ("wScan", wintypes.WORD),
                        ("dwFlags", wintypes.DWORD),
                        ("time", wintypes.DWORD),
                        ("dwExtraInfo", wintypes.ULONG_PTR),
                    ]
                },
            )
            _MouseInput = type(
                "_MouseInput",
                (ctypes.Structure,),
                {
                    "_fields_": [
                        ("dx", wintypes.LONG),
                        ("dy", wintypes.LONG),
                        ("mouseData", wintypes.DWORD),
                        ("dwFlags", wintypes.DWORD),
                        ("time", wintypes.DWORD),
                        ("dwExtraInfo", wintypes.ULONG_PTR),
                    ]
                },
            )
            _InputUnion = type(
                "_InputUnion",
                (ctypes.Union,),
                {"_fields_": [("ki", _KeyboardInput), ("mi", _MouseInput)]},
            )
            _Input = type(
                "_Input",
                (ctypes.Structure,),
                {
                    "_fields_": [
                        ("type", wintypes.DWORD),
                        ("data", _InputUnion),
                    ]
                },
            )
        except Exception:
            _KeyboardInput = None  # type: ignore[assignment]
            _InputUnion = None  # type: ignore[assignment]
            _Input = None  # type: ignore[assignment]

    else:

        class _GuiThreadInfo(ctypes.Structure):  # type: ignore[misc,override]
            _fields_: List[Tuple[str, Any]] = []

        _KeyboardInput = None  # type: ignore[assignment]
        _MouseInput = None  # type: ignore[assignment]
        _InputUnion = None  # type: ignore[assignment]
        _Input = None  # type: ignore[assignment]

    _SLIDESHOW_PRIORITY_CLASSES: Set[str] = {"screenclass"}
    _SLIDESHOW_SECONDARY_CLASSES: Set[str] = {
        "pptviewwndclass",
        "kwppshowframeclass",
        "kwppshowframe",
        "kwppshowwndclass",
        "kwpsshowframe",
    }
    _SLIDESHOW_PREFIXES: Tuple[str, ...] = ("kwppshow", "kwpsshow", "pptslideshow", "mspptslideshow")
    _DOCUMENT_WINDOW_CLASSES: Set[str] = {
        "opusapp",
        "kwpsmainframe",
        "kwpsframeclass",
        "wpsframeclass",
        "wpsmainframe",
        "wpsframewindow",
    }
    _DOCUMENT_WINDOW_PREFIXES: Tuple[str, ...] = ("opusapp", "kwps", "wpsframe", "wpsmain")
    _FORCE_WHEEL_INJECTION_CLASSES: Set[str] = set()
    _KNOWN_PRESENTATION_CLASSES: Set[str] = (
        _SLIDESHOW_PRIORITY_CLASSES
        | _SLIDESHOW_SECONDARY_CLASSES
        | _DOCUMENT_WINDOW_CLASSES
        if win32gui is not None
        else set()
    )
    _KNOWN_PRESENTATION_PREFIXES: Tuple[str, ...] = (
        _SLIDESHOW_PREFIXES + _DOCUMENT_WINDOW_PREFIXES if win32gui is not None else tuple()
    )
    _KEY_FORWARD_MAP: Dict[int, int] = (
        {
            int(Qt.Key.Key_PageUp): win32con.VK_PRIOR,
            int(Qt.Key.Key_PageDown): win32con.VK_NEXT,
            int(Qt.Key.Key_Up): win32con.VK_UP,
            int(Qt.Key.Key_Down): win32con.VK_DOWN,
            int(Qt.Key.Key_Left): win32con.VK_LEFT,
            int(Qt.Key.Key_Right): win32con.VK_RIGHT,
            int(Qt.Key.Key_Space): win32con.VK_SPACE,
            int(Qt.Key.Key_Return): win32con.VK_RETURN,
            int(Qt.Key.Key_Enter): win32con.VK_RETURN,
        }
        if win32con is not None
        else {}
    )
    _DOCUMENT_NAVIGATION_KEYS: Dict[int, int] = (
        {
            VK_UP: WHEEL_DELTA,
            VK_PRIOR: WHEEL_DELTA,
            VK_DOWN: -WHEEL_DELTA,
            VK_NEXT: -WHEEL_DELTA,
        }
        if win32con is not None
        else {}
    )
    _EXTENDED_KEY_CODES: Set[int] = (
        {
            win32con.VK_UP,
            win32con.VK_DOWN,
            win32con.VK_LEFT,
            win32con.VK_RIGHT,
        }
        if win32con is not None
        else set()
    )

    @staticmethod
    def is_supported() -> bool:
        return bool(win32api and win32con and win32gui)

    def __init__(self, overlay: "OverlayWindow") -> None:
        self.overlay = overlay
        self._last_target_hwnd: Optional[int] = None
        self._last_target_role: Optional[str] = None
        self._last_detected_role: Optional[str] = None
        self._child_buffer: List[int] = []
        self._self_pid = os.getpid()
        self._injecting_wheel = False

    def _log_debug(self, message: str, *args: Any) -> None:
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug(message, *args)

    def clear_cached_target(self) -> None:
        self._last_target_hwnd = None
        self._last_target_role = None
        self._last_detected_role = None

    def _window_class_name(self, hwnd: int) -> str:
        if hwnd == 0:
            return ""
        if win32gui is not None:
            try:
                return win32gui.GetClassName(hwnd).strip().lower()
            except Exception:
                return ""
        return _user32_window_class_name(hwnd)

    def _is_slideshow_class(self, class_name: str) -> bool:
        if not class_name:
            return False
        if class_name in self._SLIDESHOW_PRIORITY_CLASSES:
            return True
        if class_name in self._SLIDESHOW_SECONDARY_CLASSES:
            return True
        if any(class_name.startswith(prefix) for prefix in self._SLIDESHOW_PREFIXES):
            return True
        return False

    def _is_document_class(self, class_name: str) -> bool:
        if not class_name:
            return False
        if class_name in self._DOCUMENT_WINDOW_CLASSES:
            return True
        return any(class_name.startswith(prefix) for prefix in self._DOCUMENT_WINDOW_PREFIXES)

    def _classify_role_from_class_name(self, class_name: str) -> Optional[str]:
        if not class_name:
            return None
        if self._is_slideshow_class(class_name):
            return "slideshow"
        if self._is_document_class(class_name):
            return "document"
        return None

    def _should_refresh_cached_target(self, hwnd: int) -> bool:
        if self._last_target_role == "document":
            return True
        class_name = self._window_class_name(hwnd)
        return not self._is_slideshow_class(class_name)

    def _set_cached_target(self, hwnd: int, role: Optional[str]) -> None:
        self._last_target_hwnd = hwnd
        resolved_role = role or self._classify_role_from_class_name(self._window_class_name(hwnd))
        self._last_target_role = resolved_role
        self._last_detected_role = resolved_role

    def get_presentation_target(self) -> Optional[int]:
        hwnd = self._resolve_presentation_target()
        if hwnd:
            return hwnd
        return self._detect_presentation_window()

    def bring_target_to_foreground(self, hwnd: int) -> bool:
        if hwnd == 0:
            return False
        activated = False
        attach_pair = self._attach_to_target_thread(hwnd)
        try:
            activated = self._activate_window_for_input(hwnd)
            if win32gui is not None:
                try:
                    win32gui.SetForegroundWindow(hwnd)
                    activated = True
                except Exception:
                    pass
            if activated and not self._wait_for_target_foreground(hwnd):
                activated = False
        finally:
            self._detach_from_target_thread(attach_pair)
        if activated:
            role = self._last_target_role or self._classify_role_from_class_name(
                self._window_class_name(hwnd)
            )
            self._set_cached_target(hwnd, role)
        return activated

    # ---- 公共接口 ----
    def focus_presentation_window(self) -> bool:
        if not self.is_supported():
            return False
        target = self._resolve_presentation_target()
        if not target:
            target = self._detect_presentation_window()
        if not target:
            return False
        attach_pair = self._attach_to_target_thread(target)
        try:
            activated = self._activate_window_for_input(target)
        finally:
            self._detach_from_target_thread(attach_pair)
        if activated:
            self._set_cached_target(
                target,
                self._last_detected_role or self._classify_role_from_class_name(self._window_class_name(target)),
            )
        return activated

    def forward_wheel(self, event: QWheelEvent) -> bool:
        if not self._can_forward():
            self.clear_cached_target()
            return False
        if self._injecting_wheel:
            self._injecting_wheel = False
            return False
        delta_vec = event.angleDelta()
        delta = int(delta_vec.y() or delta_vec.x())
        if delta == 0:
            pixel_vec = event.pixelDelta()
            delta = int(pixel_vec.y() or pixel_vec.x())
        if delta == 0:
            return False
        target = self._resolve_presentation_target()
        if not target:
            self.clear_cached_target()
            return False
        role = self._last_target_role or self._classify_role_from_class_name(
            self._window_class_name(target)
        )
        if role == "document":
            modifiers = self._translate_mouse_modifiers(event) if win32con is not None else 0
            delivered = self._send_document_wheel(
                target,
                delta,
                modifiers=modifiers,
                source_event=event,
            )
            if not delivered:
                self.clear_cached_target()
            return delivered
        modifiers = self._translate_mouse_modifiers(event) if win32con is not None else 0
        delivered = self._send_navigation_wheel_impl(
            target,
            delta,
            role_hint=role or "slideshow",
            modifiers=modifiers,
            source_event=event,
        )
        if not delivered:
            self.clear_cached_target()
        return delivered

    def forward_key(self, event: QKeyEvent, *, is_press: bool) -> bool:
        if not self._can_forward():
            self.clear_cached_target()
            return False
        vk_code = self._resolve_vk_code(event)
        if vk_code is None:
            return False
        target = self._resolve_presentation_target()
        if not target:
            self._log_debug("forward_key: target window not found for key=%s", event.key())
            self.clear_cached_target()
            return False
        role = self._last_target_role or self._classify_role_from_class_name(
            self._window_class_name(target)
        )
        if role == "document":
            delta = self._document_nav_delta(vk_code)
            if delta:
                if not is_press:
                    return True
                delivered = self._send_document_wheel(target, delta)
                if delivered:
                    return True
                self.clear_cached_target()
                return False
        for hwnd, update_cache in self._iter_key_targets(target):
            if self._send_key_to_window(
                hwnd, vk_code, event, is_press=is_press, update_cache=update_cache
            ):
                self._log_debug(
                    "forward_key: delivered to hwnd=%s key=%s is_press=%s",
                    hwnd,
                    vk_code,
                    is_press,
                )
                return True
        self._log_debug("forward_key: delivery failed for key=%s", vk_code)
        self.clear_cached_target()
        return False

    def send_virtual_key(self, vk_code: int) -> bool:
        if not self.is_supported() or vk_code == 0:
            return False
        target = self._resolve_presentation_target()
        if not target:
            target = self._detect_presentation_window()
        if not target:
            self._log_debug("send_virtual_key: target not found vk=%s", vk_code)
            return False
        role = self._last_target_role or self._classify_role_from_class_name(
            self._window_class_name(target)
        )
        if role == "document":
            delta = self._document_nav_delta(vk_code)
            if delta:
                success = self._send_document_wheel(target, delta)
                if not success:
                    self.clear_cached_target()
                return success
        press = release = False
        with self._keyboard_capture_guard():
            attach_pair = self._attach_to_target_thread(target)
            try:
                if not self._activate_window_for_input(target):
                    self._log_debug("send_virtual_key: activate failed hwnd=%s", target)
                    return False
                press = self._send_input_event(vk_code, is_press=True)
                release = self._send_input_event(vk_code, is_press=False)
            finally:
                self._detach_from_target_thread(attach_pair)
        success = press and release
        if success:
            self._set_cached_target(target, role)
        else:
            self._log_debug(
                "send_virtual_key: send input failed vk=%s press=%s release=%s",
                vk_code,
                press,
                release,
            )
        return success

    def send_navigation_wheel_message(self, delta: int) -> bool:
        if not self.is_supported() or delta == 0:
            return False
        target = self._resolve_presentation_target()
        if not target:
            target = self._detect_presentation_window()
        if not target:
            return False
        role = self._last_target_role or self._classify_role_from_class_name(
            self._window_class_name(target)
        )
        if self._should_force_wheel_injection(target, role):
            return False
        delivered = self._deliver_wheel_messages(target, delta)
        if delivered:
            self._set_cached_target(target, role)
        else:
            self.clear_cached_target()
        return delivered

    def send_navigation_wheel(self, delta: int, *, injection_only: bool = False) -> bool:
        if not self.is_supported() or delta == 0:
            return False
        target = self._resolve_presentation_target()
        if not target:
            target = self._detect_presentation_window()
        if not target:
            return False
        role = self._last_target_role or self._classify_role_from_class_name(
            self._window_class_name(target)
        )
        role_hint = role or "slideshow"
        delivered = self._send_navigation_wheel_impl(
            target,
            delta,
            role_hint=role_hint,
            injection_only=injection_only,
        )
        if not delivered and role_hint != "slideshow":
            delivered = self._send_navigation_wheel_impl(
                target,
                delta,
                role_hint="slideshow",
                injection_only=injection_only,
            )
        if not delivered:
            self.clear_cached_target()
        return delivered

    def get_last_target_role(self) -> Optional[str]:
        return self._last_target_role

    def send_document_scroll_for_vk(self, vk_code: int) -> bool:
        delta = self._document_nav_delta(vk_code)
        if not delta:
            return False
        target = self._resolve_presentation_target()
        if not target:
            return False
        role = self._last_target_role or self._classify_role_from_class_name(
            self._window_class_name(target)
        )
        if role != "document":
            return False
        return self._send_document_wheel(target, delta)

    # ---- 内部工具方法 ----
    def _can_forward(self) -> bool:
        if not self.is_supported():
            return False
        mode = getattr(self.overlay, "mode", "cursor")
        if mode == "cursor" and not getattr(
            self.overlay, "navigation_passthrough_active", False
        ):
            return False
        if getattr(self.overlay, "whiteboard_active", False):
            return False
        return True

    def _translate_mouse_modifiers(self, event: QWheelEvent) -> int:
        keys = 0
        modifiers = event.modifiers()
        if modifiers & Qt.KeyboardModifier.ShiftModifier:
            keys |= win32con.MK_SHIFT
        if modifiers & Qt.KeyboardModifier.ControlModifier:
            keys |= win32con.MK_CONTROL
        buttons = event.buttons()
        if buttons & Qt.MouseButton.LeftButton:
            keys |= win32con.MK_LBUTTON
        if buttons & Qt.MouseButton.RightButton:
            keys |= win32con.MK_RBUTTON
        if buttons & Qt.MouseButton.MiddleButton:
            keys |= win32con.MK_MBUTTON
        return keys

    def _resolve_vk_code(self, event: QKeyEvent) -> Optional[int]:
        native_getter = getattr(event, "nativeVirtualKey", None)
        vk_code = 0
        if callable(native_getter):
            try:
                vk_code = int(native_getter())
            except Exception:
                vk_code = 0
        if vk_code:
            return vk_code
        vk_code = self._KEY_FORWARD_MAP.get(event.key(), 0)
        return vk_code or None

    def _send_key_to_window(
        self,
        hwnd: int,
        vk_code: int,
        event: QKeyEvent,
        *,
        is_press: bool,
        update_cache: bool,
    ) -> bool:
        delivered = False
        if self._inject_key_event(hwnd, vk_code, event, is_press=is_press):
            delivered = True
        elif win32con is not None:
            message = win32con.WM_KEYDOWN if is_press else win32con.WM_KEYUP
            l_param = self._build_key_lparam(vk_code, event, is_press)
            delivered = self._deliver_key_message(hwnd, message, vk_code, l_param)
        if delivered and update_cache:
            self._last_target_hwnd = hwnd
        if not delivered:
            self._log_debug(
                "_send_key_to_window: failed hwnd=%s vk=%s is_press=%s",
                hwnd,
                vk_code,
                is_press,
            )
        return delivered

    def _inject_key_event(
        self,
        hwnd: int,
        vk_code: int,
        event: QKeyEvent,
        *,
        is_press: bool,
    ) -> bool:
        if (
            _USER32 is None
            or self._Input is None
            or self._KeyboardInput is None
            or hwnd == 0
            or vk_code == 0
        ):
            return False
        success = False
        with self._keyboard_capture_guard():
            attach_pair = self._attach_to_target_thread(hwnd)
            try:
                if not self._activate_window_for_input(hwnd):
                    self._log_debug("_inject_key_event: activate failed hwnd=%s", hwnd)
                    return False
                success = self._send_input_event(vk_code, is_press=is_press)
            finally:
                self._detach_from_target_thread(attach_pair)
        return success

    @contextlib.contextmanager
    def _keyboard_capture_guard(self) -> Iterable[None]:
        release = getattr(self.overlay, "_release_keyboard_capture", None)
        capture = getattr(self.overlay, "_ensure_keyboard_capture", None)
        try:
            if callable(release):
                release()
        except Exception:
            pass
        try:
            yield
        finally:
            if callable(capture):
                def _restore_focus() -> None:
                    try:
                        if getattr(self.overlay, "mode", "") != "cursor":
                            capture()
                        else:
                            if getattr(self.overlay, "_keyboard_grabbed", False):
                                self.overlay._release_keyboard_capture()
                    except Exception:
                        return
                    try:
                        self.overlay.raise_toolbar()
                    except Exception:
                        pass
                    try:
                        QApplication.processEvents()
                    except Exception:
                        pass

                try:
                    QTimer.singleShot(10, _restore_focus)
                except Exception:
                    try:
                        _restore_focus()
                    except Exception:
                        pass

    def _attach_to_target_thread(self, hwnd: int) -> Optional[Tuple[int, int]]:
        if _USER32 is None:
            return None
        target_thread = self._window_thread_id(hwnd)
        if not target_thread:
            return None
        try:
            current_thread = int(_USER32.GetCurrentThreadId())
        except Exception:
            current_thread = 0
        if not current_thread or current_thread == target_thread:
            return None
        try:
            attached = bool(_USER32.AttachThreadInput(current_thread, target_thread, True))
        except Exception:
            attached = False
        return (current_thread, target_thread) if attached else None

    def _detach_from_target_thread(self, pair: Optional[Tuple[int, int]]) -> None:
        if _USER32 is None or not pair:
            return
        src, dst = pair
        if not src or not dst or src == dst:
            return
        try:
            _USER32.AttachThreadInput(src, dst, False)
        except Exception:
            pass

    def _window_thread_id(self, hwnd: int) -> int:
        if _USER32 is None or hwnd == 0:
            return 0
        pid = wintypes.DWORD()
        try:
            thread_id = int(_USER32.GetWindowThreadProcessId(wintypes.HWND(hwnd), ctypes.byref(pid)))
        except Exception:
            thread_id = 0
        return thread_id

    def _wait_for_target_foreground(self, hwnd: int, *, timeout: float = 0.3) -> bool:
        if hwnd == 0:
            return False
        class_name = self._window_class_name(hwnd)
        if not self._is_slideshow_class(class_name):
            return True
        deadline = time.time() + max(0.0, float(timeout))
        target_top = self._top_level_hwnd(hwnd)
        while time.time() < deadline:
            foreground = 0
            if win32gui is not None:
                try:
                    fg = win32gui.GetForegroundWindow()
                    if fg:
                        foreground = int(fg)
                except Exception:
                    foreground = 0
            if not foreground:
                foreground = _user32_get_foreground_window()
            if foreground:
                if foreground == hwnd or foreground == target_top:
                    return True
                fg_class = self._window_class_name(foreground)
                if self._is_slideshow_class(fg_class) and self._top_level_hwnd(foreground) == target_top:
                    return True
            time.sleep(0.01)
        return False

    def _activate_window_for_input(self, hwnd: int) -> bool:
        if _USER32 is None:
            return False
        root_hwnd = self._top_level_hwnd(hwnd)
        activated = False
        try:
            activated = bool(_USER32.SetActiveWindow(wintypes.HWND(root_hwnd)))
        except Exception:
            activated = False
        if not activated:
            try:
                activated = bool(_USER32.SetForegroundWindow(wintypes.HWND(root_hwnd)))
            except Exception:
                activated = False
        focus_ok = False
        try:
            focus_ok = bool(_USER32.SetFocus(wintypes.HWND(hwnd)))
        except Exception:
            focus_ok = False
        if not focus_ok and root_hwnd != hwnd:
            try:
                focus_ok = bool(_USER32.SetFocus(wintypes.HWND(root_hwnd)))
            except Exception:
                focus_ok = False
        success = activated or focus_ok
        if success and not self._wait_for_target_foreground(hwnd):
            success = False
        return success

    def _top_level_hwnd(self, hwnd: int) -> int:
        if win32gui is None or hwnd == 0:
            return hwnd
        try:
            ga_root = getattr(win32con, "GA_ROOT", 2) if win32con is not None else 2
        except Exception:
            ga_root = 2
        try:
            root = win32gui.GetAncestor(hwnd, ga_root)
            if root:
                return int(root)
        except Exception:
            pass
        try:
            parent = win32gui.GetParent(hwnd)
            if parent:
                return int(parent)
        except Exception:
            pass
        return hwnd

    def _send_input_event(self, vk_code: int, *, is_press: bool) -> bool:
        if _USER32 is None or self._Input is None or self._KeyboardInput is None:
            return False
        keyboard_input = self._KeyboardInput()
        keyboard_input.wVk = vk_code & 0xFFFF
        keyboard_input.wScan = self._map_virtual_key(vk_code)
        flags = 0
        if vk_code in self._EXTENDED_KEY_CODES:
            flags |= self._KEYEVENTF_EXTENDEDKEY
        if not is_press:
            flags |= self._KEYEVENTF_KEYUP
        keyboard_input.dwFlags = flags
        keyboard_input.time = 0
        try:
            keyboard_input.dwExtraInfo = 0
        except Exception:
            pass
        input_record = self._Input()
        input_record.type = self._INPUT_KEYBOARD
        input_record.data.ki = keyboard_input
        try:
            sent = int(_USER32.SendInput(1, ctypes.byref(input_record), ctypes.sizeof(self._Input)))
        except Exception:
            sent = 0
        return bool(sent)

    def _send_wheel_input(self, delta: int) -> bool:
        if (
            _USER32 is None
            or self._Input is None
            or getattr(self, "_MouseInput", None) is None
            or self._MouseInput is None
        ):
            return False
        mouse_input = self._MouseInput()
        mouse_input.dx = 0
        mouse_input.dy = 0
        try:
            mouse_input.mouseData = ctypes.c_int(delta).value & 0xFFFFFFFF
        except Exception:
            mouse_input.mouseData = ctypes.c_int(delta).value
        mouse_input.dwFlags = getattr(win32con, "MOUSEEVENTF_WHEEL", 0x0800)
        mouse_input.time = 0
        try:
            mouse_input.dwExtraInfo = 0
        except Exception:
            pass
        input_record = self._Input()
        input_record.type = self._INPUT_MOUSE
        try:
            input_record.data.mi = mouse_input
        except Exception:
            return False
        try:
            sent = int(_USER32.SendInput(1, ctypes.byref(input_record), ctypes.sizeof(self._Input)))
        except Exception:
            sent = 0
        return bool(sent)

    def _get_cursor_pos(self) -> Optional[Tuple[int, int]]:
        if win32api is not None:
            try:
                pos = win32api.GetCursorPos()
                if isinstance(pos, tuple) and len(pos) == 2:
                    return int(pos[0]), int(pos[1])
            except Exception:
                pass
        if _USER32 is not None:
            pt = wintypes.POINT()
            try:
                if _USER32.GetCursorPos(ctypes.byref(pt)):
                    return int(pt.x), int(pt.y)
            except Exception:
                pass
        return None

    def _set_cursor_pos(self, point: Tuple[int, int]) -> bool:
        x, y = point
        if win32api is not None:
            try:
                win32api.SetCursorPos((int(x), int(y)))
                return True
            except Exception:
                pass
        if _USER32 is not None:
            try:
                return bool(_USER32.SetCursorPos(int(x), int(y)))
            except Exception:
                return False
        return False

    def _set_overlay_passthrough(self, enable: bool) -> bool:
        overlay = getattr(self, "overlay", None)
        if overlay is None:
            return False
        method = getattr(overlay, "_apply_input_passthrough", None)
        if not callable(method):
            return False
        try:
            method(enable)
        except Exception:
            return False
        return True

    def _restore_overlay_interaction(self) -> None:
        overlay = getattr(self, "overlay", None)
        if overlay is None:
            return
        try:
            overlay._apply_input_passthrough(False)
        except Exception:
            pass
        try:
            overlay._ensure_keyboard_capture()
        except Exception:
            pass
        try:
            QApplication.processEvents()
        except Exception:
            pass

    def _map_virtual_key(self, vk_code: int) -> int:
        map_vk = getattr(win32api, "MapVirtualKey", None) if win32api is not None else None
        if callable(map_vk):
            try:
                return int(map_vk(vk_code, 0)) & 0xFFFF
            except Exception:
                return 0
        return 0

    def _iter_key_targets(self, target: int) -> Iterable[Tuple[int, bool]]:
        seen: Set[int] = set()

        def _push(
            hwnd: int,
            *,
            cache: bool,
            require_visible: bool,
        ) -> Iterable[Tuple[int, bool]]:
            if hwnd in seen:
                return ()
            if not self._is_keyboard_target(hwnd, require_visible=require_visible):
                return ()
            seen.add(hwnd)
            return ((hwnd, cache),)

        for focus_hwnd in self._gather_thread_focus_handles(target):
            for candidate in _push(focus_hwnd, cache=False, require_visible=False):
                yield candidate
        for candidate in _push(target, cache=True, require_visible=True):
            yield candidate
        for child_hwnd in self._collect_descendant_windows(target):
            for candidate in _push(child_hwnd, cache=False, require_visible=False):
                yield candidate
    
    def _iter_wheel_targets(self, target: int) -> Iterable[int]:
        seen: Set[int] = set()

        for focus_hwnd in self._gather_thread_focus_handles(target):
            if focus_hwnd in seen:
                continue
            if self._is_keyboard_target(focus_hwnd, require_visible=False):
                seen.add(focus_hwnd)
                yield focus_hwnd

        if self._is_keyboard_target(target, require_visible=True) and target not in seen:
            seen.add(target)
            yield target

        for child_hwnd in self._collect_descendant_windows(target):
            if child_hwnd in seen:
                continue
            if self._is_keyboard_target(child_hwnd, require_visible=False):
                seen.add(child_hwnd)
                yield child_hwnd

    def _build_key_lparam(self, vk_code: int, event: QKeyEvent, is_press: bool) -> int:
        repeat_getter = getattr(event, "count", None)
        repeat_count = 1
        if callable(repeat_getter):
            try:
                repeat_count = max(1, int(repeat_getter()))
            except Exception:
                repeat_count = 1
        l_param = repeat_count & 0xFFFF
        scan_code = self._map_virtual_key(vk_code)
        l_param |= (scan_code & 0xFF) << 16
        if vk_code in self._EXTENDED_KEY_CODES:
            l_param |= 1 << 24
        auto_repeat_getter = getattr(event, "isAutoRepeat", None)
        is_auto_repeat = False
        if callable(auto_repeat_getter):
            try:
                is_auto_repeat = bool(auto_repeat_getter())
            except Exception:
                is_auto_repeat = False
        if is_press:
            if is_auto_repeat:
                l_param |= 1 << 30
        else:
            l_param |= 1 << 30
            l_param |= 1 << 31
        return l_param & 0xFFFFFFFF

    def _deliver_key_message(self, hwnd: int, message: int, vk_code: int, l_param: int) -> bool:
        delivered = False
        if win32api is not None:
            try:
                delivered = bool(win32api.PostMessage(hwnd, message, vk_code, l_param))
            except Exception:
                delivered = False
        if delivered:
            return True
        if _USER32 is None:
            return False
        result = ctypes.c_size_t()
        try:
            sent = _USER32.SendMessageTimeoutW(
                hwnd,
                message,
                wintypes.WPARAM(vk_code),
                wintypes.LPARAM(l_param),
                self._SMTO_ABORTIFHUNG,
                30,
                ctypes.byref(result),
            )
        except Exception:
            sent = 0
        return bool(sent)

    def _is_overlay_window(self, hwnd: int) -> bool:
        try:
            overlay_hwnd = int(self.overlay.winId()) if self.overlay.winId() else 0
        except Exception:
            overlay_hwnd = 0
        return hwnd != 0 and hwnd == overlay_hwnd

    def _is_keyboard_target(self, hwnd: int, *, require_visible: bool) -> bool:
        if hwnd == 0 or self._is_overlay_window(hwnd):
            return False
        if win32gui is None:
            return False
        try:
            if not win32gui.IsWindow(hwnd):
                return False
            if require_visible and not win32gui.IsWindowVisible(hwnd):
                return False
        except Exception:
            return False
        return True

    def _gather_thread_focus_handles(self, target: int) -> Iterable[int]:
        if _USER32 is None:
            return ()
        info = self._GuiThreadInfo()
        pid = wintypes.DWORD()
        try:
            thread_id = _USER32.GetWindowThreadProcessId(
                wintypes.HWND(target), ctypes.byref(pid)
            )
        except Exception:
            return ()
        if not thread_id:
            return ()
        info.cbSize = ctypes.sizeof(info)
        try:
            ok = bool(_USER32.GetGUIThreadInfo(thread_id, ctypes.byref(info)))
        except Exception:
            ok = False
        if not ok:
            return ()
        handles = (
            info.hwndFocus,
            info.hwndActive,
            info.hwndCapture,
            info.hwndMenuOwner,
            info.hwndCaret,
        )
        return tuple(int(h) for h in handles if h)

    def _collect_descendant_windows(self, root: int) -> Iterable[int]:
        if win32gui is None:
            return ()
        queue: deque[int] = deque([root])
        discovered: Set[int] = {root}
        results: List[int] = []
        buffer = self._child_buffer
        while queue and len(results) < self._MAX_CHILD_FORWARDS:
            parent = queue.popleft()
            buffer.clear()

            def _collector(child_hwnd: int, acc: List[int]) -> bool:
                if child_hwnd not in discovered:
                    acc.append(child_hwnd)
                return len(acc) < self._MAX_CHILD_FORWARDS

            try:
                win32gui.EnumChildWindows(parent, _collector, buffer)
            except Exception:
                continue
            snapshot = list(buffer)
            buffer.clear()
            for child in snapshot:
                if child in discovered:
                    continue
                discovered.add(child)
                results.append(child)
                if len(results) >= self._MAX_CHILD_FORWARDS:
                    break
                queue.append(child)
        return tuple(results)

    def _overlay_rect_tuple(self) -> Optional[Tuple[int, int, int, int]]:
        rect = self.overlay.geometry()
        if rect.isNull():
            return None
        left = rect.left()
        top = rect.top()
        right = left + rect.width()
        bottom = top + rect.height()
        return left, top, right, bottom

    def _rect_intersects_overlay(self, rect: Tuple[int, int, int, int]) -> bool:
        overlay_rect = self._overlay_rect_tuple()
        if overlay_rect is None:
            return False
        left, top, right, bottom = rect
        o_left, o_top, o_right, o_bottom = overlay_rect
        return not (right <= o_left or left >= o_right or bottom <= o_top or top >= o_bottom)

    def _fallback_is_target_window_valid(self, hwnd: int) -> bool:
        if _USER32 is None or hwnd == 0:
            return False
        if self._belongs_to_current_process(hwnd):
            return False
        overlay_hwnd = int(self.overlay.winId()) if self.overlay.winId() else 0
        if hwnd == overlay_hwnd:
            return False
        if not _user32_is_window(hwnd):
            return False
        if not _user32_is_window_visible(hwnd) or _user32_is_window_iconic(hwnd):
            return False
        rect = _user32_window_rect(hwnd)
        if not rect:
            return False
        return self._rect_intersects_overlay(rect)

    def _fallback_is_candidate_window(self, hwnd: int) -> bool:
        if _USER32 is None or hwnd == 0:
            return False
        if self._belongs_to_current_process(hwnd):
            return False
        class_name = _user32_window_class_name(hwnd)
        rect = _user32_window_rect(hwnd)
        if not class_name or not rect:
            return False
        class_name = class_name.strip().lower()
        role = self._classify_role_from_class_name(class_name)
        if role == "slideshow":
            return True
        if role == "document":
            return self._is_document_window_valid(hwnd, rect=rect, index=0, foreground=None)
        left, top, right, bottom = rect
        width = max(0, right - left)
        height = max(0, bottom - top)
        overlay_rect = self._overlay_rect_tuple()
        if overlay_rect is None:
            return False
        o_width = overlay_rect[2] - overlay_rect[0]
        o_height = overlay_rect[3] - overlay_rect[1]
        if o_width <= 0 or o_height <= 0:
            return False
        width_diff = abs(width - o_width)
        height_diff = abs(height - o_height)
        return width >= 400 and height >= 300 and width_diff <= 64 and height_diff <= 64

    def _get_window_styles(self, hwnd: int) -> Tuple[Optional[int], Optional[int]]:
        style: Optional[int] = None
        ex_style: Optional[int] = None
        try:
            if win32gui is not None:
                style = int(win32gui.GetWindowLong(hwnd, getattr(win32con, "GWL_STYLE", -16)))
                ex_style = int(win32gui.GetWindowLong(hwnd, getattr(win32con, "GWL_EXSTYLE", -20)))
            elif _USER32 is not None:
                gwl_style = getattr(win32con, "GWL_STYLE", -16)
                gwl_exstyle = getattr(win32con, "GWL_EXSTYLE", -20)
                style = int(_USER32.GetWindowLongW(wintypes.HWND(hwnd), gwl_style))
                ex_style = int(_USER32.GetWindowLongW(wintypes.HWND(hwnd), gwl_exstyle))
        except Exception:
            style = style if isinstance(style, int) else None
            ex_style = ex_style if isinstance(ex_style, int) else None
        return style, ex_style

    def _has_window_caption(self, hwnd: int) -> Optional[bool]:
        style, _ = self._get_window_styles(hwnd)
        if style is None:
            return None
        caption_flag = getattr(win32con, "WS_CAPTION", 0x00C00000)
        return bool(style & caption_flag)

    def _is_topmost_window(self, hwnd: int) -> Optional[bool]:
        _, ex_style = self._get_window_styles(hwnd)
        if ex_style is None:
            return None
        topmost_flag = getattr(win32con, "WS_EX_TOPMOST", 0x00000008)
        return bool(ex_style & topmost_flag)

    def _get_window_rect_generic(self, hwnd: int) -> Optional[Tuple[int, int, int, int]]:
        if win32gui is not None:
            try:
                rect = win32gui.GetWindowRect(hwnd)
                if rect:
                    return int(rect[0]), int(rect[1]), int(rect[2]), int(rect[3])
            except Exception:
                pass
        return _user32_window_rect(hwnd)

    def _candidate_score(
        self,
        hwnd: int,
        *,
        class_name: Optional[str] = None,
        rect: Optional[Tuple[int, int, int, int]] = None,
    ) -> int:
        if rect is None:
            rect = self._get_window_rect_generic(hwnd)
        if rect is None:
            return -1
        left, top, right, bottom = rect
        width = max(0, right - left)
        height = max(0, bottom - top)
        if width == 0 or height == 0:
            return -1
        if class_name is None:
            if win32gui is not None:
                try:
                    class_name = win32gui.GetClassName(hwnd)
                except Exception:
                    class_name = ""
            if not class_name:
                class_name = _user32_window_class_name(hwnd)
        class_name = (class_name or "").strip().lower()

        score = 0
        if class_name in self._SLIDESHOW_PRIORITY_CLASSES:
            score += 400
        elif self._is_slideshow_class(class_name):
            if class_name in self._SLIDESHOW_SECONDARY_CLASSES:
                score += 240
            else:
                score += 200
        elif class_name in self._DOCUMENT_WINDOW_CLASSES:
            score += 120

        has_caption = self._has_window_caption(hwnd)
        if has_caption is False:
            score += 160
        elif has_caption is True:
            score -= 40

        is_topmost = self._is_topmost_window(hwnd)
        if is_topmost:
            score += 40

        overlay_rect = self._overlay_rect_tuple()
        if overlay_rect is not None:
            o_width = max(0, overlay_rect[2] - overlay_rect[0])
            o_height = max(0, overlay_rect[3] - overlay_rect[1])
            if o_width > 0 and o_height > 0:
                width_diff = abs(width - o_width)
                height_diff = abs(height - o_height)
                size_penalty = min(width_diff + height_diff, 800)
                score += max(0, 220 - size_penalty // 2)
                area = width * height
                overlay_area = o_width * o_height
                if overlay_area > 0:
                    ratio = min(area, overlay_area) / max(area, overlay_area)
                    score += int(ratio * 120)
                overlap_x = max(0, min(right, overlay_rect[2]) - max(left, overlay_rect[0]))
                overlap_y = max(0, min(bottom, overlay_rect[3]) - max(top, overlay_rect[1]))
                overlap_area = overlap_x * overlap_y
                if overlap_area > 0 and area > 0:
                    score += int((overlap_area / area) * 140)

        return score

    def _is_top_level_candidate(self, hwnd: int) -> bool:
        if hwnd == 0:
            return False
        if win32gui is not None:
            try:
                if not win32gui.IsWindow(hwnd) or not win32gui.IsWindowVisible(hwnd):
                    return False
                if win32gui.IsIconic(hwnd):
                    return False
            except Exception:
                return False
            return True
        if _USER32 is None:
            return False
        if not _user32_is_window(hwnd):
            return False
        if not _user32_is_window_visible(hwnd):
            return False
        if _user32_is_window_iconic(hwnd):
            return False
        return True

    def _gather_target_candidates(self) -> List[Tuple[int, str, str, Tuple[int, int, int, int]]]:
        overlay_hwnd = int(self.overlay.winId()) if self.overlay.winId() else 0
        results: List[Tuple[int, str, str, Tuple[int, int, int, int]]] = []
        seen: Set[int] = set()

        def _append(hwnd: int) -> None:
            if hwnd in seen or hwnd == 0:
                return
            seen.add(hwnd)
            if hwnd == overlay_hwnd:
                return
            if self._belongs_to_current_process(hwnd):
                return
            if not self._is_top_level_candidate(hwnd):
                return
            rect = self._get_window_rect_generic(hwnd)
            if not rect or not self._rect_intersects_overlay(rect):
                return
            class_name = self._window_class_name(hwnd)
            role = self._classify_role_from_class_name(class_name)
            if role is None:
                return
            results.append((hwnd, role, class_name, rect))

        if win32gui is not None:
            def _enum(hwnd: int, _unused) -> bool:
                _append(hwnd)
                return True

            try:
                win32gui.EnumWindows(_enum, None)
            except Exception:
                pass
        elif _USER32 is not None and _WNDENUMPROC is not None:
            def _enum(hwnd: int, _l_param: int) -> int:
                _append(int(hwnd))
                return True

            try:
                _USER32.EnumWindows(_WNDENUMPROC(_enum), 0)
            except Exception:
                pass
        return results

    def _is_document_window_valid(
        self,
        self_hwnd: int,
        *,
        rect: Tuple[int, int, int, int],
        index: int,
        foreground: Optional[int],
    ) -> bool:
        hwnd = self_hwnd
        if not self._is_target_window_valid(hwnd):
            return False
        if foreground:
            foreground_top = self._top_level_hwnd(foreground)
            if foreground_top == 0:
                foreground_top = foreground
            if self._top_level_hwnd(hwnd) != foreground_top:
                return False
        elif index != 0:
            return False
        if not self._is_window_maximized(hwnd):
            return False
        left, top, right, bottom = rect
        width = max(0, right - left)
        height = max(0, bottom - top)
        if width < 400 or height < 300:
            return False
        return True

    def _document_candidate_score(
        self,
        hwnd: int,
        rect: Tuple[int, int, int, int],
        index: int,
    ) -> int:
        if index != 0:
            return -1
        left, top, right, bottom = rect
        width = max(0, right - left)
        height = max(0, bottom - top)
        overlay_rect = self._overlay_rect_tuple()
        score = 320
        if overlay_rect is not None:
            o_width = max(0, overlay_rect[2] - overlay_rect[0])
            o_height = max(0, overlay_rect[3] - overlay_rect[1])
            if o_width > 0 and o_height > 0:
                width_diff = abs(width - o_width)
                height_diff = abs(height - o_height)
                size_penalty = min(width_diff + height_diff, 800)
                score += max(0, 240 - size_penalty // 2)
                area = width * height
                overlay_area = o_width * o_height
                if overlay_area > 0 and area > 0:
                    ratio = min(area, overlay_area) / max(area, overlay_area)
                    score += int(ratio * 80)
        return score

    def _belongs_to_current_process(self, hwnd: int) -> bool:
        if _USER32 is None or hwnd == 0:
            return False
        pid = wintypes.DWORD()
        try:
            _USER32.GetWindowThreadProcessId(wintypes.HWND(hwnd), ctypes.byref(pid))
        except Exception:
            return False
        return int(pid.value) == self._self_pid

    def _next_window_in_z_order(self, hwnd: int) -> int:
        if hwnd == 0:
            return 0
        if win32gui is not None:
            command = getattr(win32con, "GW_HWNDNEXT", 2) if win32con is not None else 2
            try:
                nxt = win32gui.GetWindow(hwnd, command)
                if nxt:
                    return int(nxt)
            except Exception:
                pass
        if _USER32 is not None:
            command = getattr(win32con, "GW_HWNDNEXT", 2) if win32con is not None else 2
            try:
                nxt = _USER32.GetWindow(wintypes.HWND(hwnd), command)
                if nxt:
                    return int(nxt)
            except Exception:
                pass
        return 0

    def _effective_foreground_window(self) -> int:
        overlay_hwnd = int(self.overlay.winId()) if self.overlay.winId() else 0
        foreground = 0
        if win32gui is not None:
            try:
                fg = win32gui.GetForegroundWindow()
                if fg:
                    foreground = int(fg)
            except Exception:
                foreground = 0
        if not foreground:
            foreground = _user32_get_foreground_window()
        visited: Set[int] = set()
        hwnd = foreground
        while hwnd and hwnd not in visited:
            visited.add(hwnd)
            top = self._top_level_hwnd(hwnd)
            if top == overlay_hwnd or self._belongs_to_current_process(top):
                hwnd = self._next_window_in_z_order(top)
                continue
            if self._is_top_level_candidate(top):
                return top
            hwnd = self._next_window_in_z_order(top)
        return 0

    def _is_window_maximized(self, hwnd: int) -> bool:
        if win32gui is not None:
            try:
                if win32gui.IsZoomed(hwnd):
                    return True
            except Exception:
                pass
            try:
                placement = win32gui.GetWindowPlacement(hwnd)
                if placement and placement[1] == getattr(win32con, "SW_SHOWMAXIMIZED", 3):
                    return True
            except Exception:
                pass
        if _USER32 is not None:
            try:
                if _USER32.IsZoomed(wintypes.HWND(hwnd)):
                    return True
            except Exception:
                pass
        return False

    def _document_nav_delta(self, vk_code: int) -> int:
        return self._DOCUMENT_NAVIGATION_KEYS.get(vk_code, 0)

    def _get_default_wheel_point(self, hwnd: int) -> Optional[Tuple[int, int]]:
        rect = self._get_window_rect_generic(hwnd)
        if not rect:
            return None
        left, top, right, bottom = rect
        width = max(0, right - left)
        height = max(0, bottom - top)
        return left + width // 2, top + height // 2

    def _send_wheel_message(
        self,
        hwnd: int,
        delta: int,
        *,
        modifiers: int = 0,
        point: Optional[Tuple[int, int]] = None,
    ) -> bool:
        if hwnd == 0:
            return False
        if point is None:
            point = self._get_default_wheel_point(hwnd)
            if point is None:
                return False
        x, y = point
        delta_word = ctypes.c_short(delta).value & 0xFFFF
        w_param = (ctypes.c_ushort(modifiers).value & 0xFFFF) | (delta_word << 16)
        x_word = ctypes.c_short(int(x)).value & 0xFFFF
        y_word = ctypes.c_short(int(y)).value & 0xFFFF
        l_param = x_word | (y_word << 16)
        delivered = False
        if win32api is not None and win32con is not None:
            try:
                delivered = bool(win32api.PostMessage(hwnd, win32con.WM_MOUSEWHEEL, w_param, l_param))
            except Exception:
                delivered = False
        if not delivered and _USER32 is not None and win32con is not None:
            result = ctypes.c_size_t()
            try:
                sent = _USER32.SendMessageTimeoutW(
                    hwnd,
                    win32con.WM_MOUSEWHEEL,
                    wintypes.WPARAM(w_param),
                    wintypes.LPARAM(l_param),
                    self._SMTO_ABORTIFHUNG,
                    30,
                    ctypes.byref(result),
                )
                delivered = bool(sent)
            except Exception:
                delivered = False
        return delivered

    def _should_force_wheel_injection(self, hwnd: int, role_hint: Optional[str]) -> bool:
        if hwnd == 0:
            return False
        if role_hint == "document":
            return False
        class_name = self._window_class_name(hwnd)
        if not class_name:
            return False
        return class_name in self._FORCE_WHEEL_INJECTION_CLASSES

    def _deliver_wheel_messages(
        self,
        hwnd: int,
        delta: int,
        *,
        modifiers: int = 0,
        point: Optional[Tuple[int, int]] = None,
    ) -> bool:
        if hwnd == 0:
            return False
        effective_point = point or self._get_default_wheel_point(hwnd)
        if effective_point is None:
            return False
        for candidate in self._iter_wheel_targets(hwnd):
            if self._send_wheel_message(
                candidate,
                delta,
                modifiers=modifiers,
                point=effective_point,
            ):
                return True
        return self._send_wheel_message(
            hwnd,
            delta,
            modifiers=modifiers,
            point=effective_point,
        )

    def _send_wheel_via_injection(
        self,
        hwnd: int,
        delta: int,
        *,
        modifiers: int = 0,
        point: Optional[Tuple[int, int]] = None,
        allow_message_fallback: bool = False,
    ) -> bool:
        if hwnd == 0:
            return False
        delivered = False
        passthrough_applied = False
        original_cursor: Optional[Tuple[int, int]] = None
        with self._keyboard_capture_guard():
            focus_ok = self.bring_target_to_foreground(hwnd)
            if not focus_ok:
                focus_ok = self._activate_window_for_input(hwnd)
            attach_pair = self._attach_to_target_thread(hwnd)
            try:
                injected = False
                if focus_ok:
                    original_cursor = self._get_cursor_pos()
                    passthrough_applied = self._set_overlay_passthrough(True)
                    if passthrough_applied:
                        QApplication.processEvents()
                    if point is not None:
                        self._set_cursor_pos(point)
                    self._injecting_wheel = True
                    try:
                        injected = self._send_wheel_input(delta)
                    finally:
                        if not injected:
                            self._injecting_wheel = False
                    if injected:
                        delivered = True
                if not delivered and allow_message_fallback:
                    delivered = self._deliver_wheel_messages(
                        hwnd,
                        delta,
                        modifiers=modifiers,
                        point=point,
                    )
            finally:
                self._detach_from_target_thread(attach_pair)
        if passthrough_applied:
            self._restore_overlay_interaction()
        if original_cursor is not None:
            self._set_cursor_pos(original_cursor)
        self._injecting_wheel = False
        return delivered

    def _send_navigation_wheel_impl(
        self,
        hwnd: int,
        delta: int,
        *,
        role_hint: str,
        modifiers: int = 0,
        source_event: Optional[QWheelEvent] = None,
        injection_only: bool = False,
    ) -> bool:
        if hwnd == 0:
            return False
        point: Optional[Tuple[int, int]] = None
        if source_event is not None:
            try:
                global_pos = source_event.globalPosition().toPoint()
                point = (int(global_pos.x()), int(global_pos.y()))
            except Exception:
                point = None
        default_point = point if point is not None else self._get_default_wheel_point(hwnd)
        force_injection = self._should_force_wheel_injection(hwnd, role_hint)
        delivered = False
        message_attempted = injection_only
        if not force_injection and not injection_only:
            delivered = self._deliver_wheel_messages(
                hwnd,
                delta,
                modifiers=modifiers,
                point=default_point,
            )
            message_attempted = True
        if not delivered:
            delivered = self._send_wheel_via_injection(
                hwnd,
                delta,
                modifiers=modifiers,
                point=default_point,
                allow_message_fallback=not message_attempted,
            )
        if delivered:
            self._set_cached_target(hwnd, role_hint)
        else:
            self.clear_cached_target()
        return delivered

    def _send_document_wheel(
        self,
        hwnd: int,
        delta: int,
        *,
        modifiers: int = 0,
        source_event: Optional[QWheelEvent] = None,
    ) -> bool:
        return self._send_navigation_wheel_impl(
            hwnd,
            delta,
            role_hint="document",
            modifiers=modifiers,
            source_event=source_event,
        )

    def _fallback_detect_presentation_window_user32(self) -> Optional[int]:
        return self._detect_presentation_window()

    def _is_target_window_valid(self, hwnd: int) -> bool:
        if hwnd == 0:
            return False
        if self._belongs_to_current_process(hwnd):
            return False
        if win32gui is None:
            return self._fallback_is_target_window_valid(hwnd)
        try:
            if hwnd == int(self.overlay.winId()):
                return False
            if not win32gui.IsWindow(hwnd) or not win32gui.IsWindowVisible(hwnd):
                return False
            if win32gui.IsIconic(hwnd):
                return False
            rect = win32gui.GetWindowRect(hwnd)
        except Exception:
            return False
        if not rect:
            return False
        return self._rect_intersects_overlay(rect)

    def _is_candidate_window(self, hwnd: int) -> bool:
        if win32gui is None:
            return self._fallback_is_candidate_window(hwnd)
        if self._belongs_to_current_process(hwnd):
            return False
        try:
            class_name = win32gui.GetClassName(hwnd)
        except Exception:
            class_name = ""
        class_name = class_name.strip().lower()
        if not class_name:
            return False
        role = self._classify_role_from_class_name(class_name)
        if role == "slideshow":
            return True
        if role == "document":
            return True
        try:
            rect = win32gui.GetWindowRect(hwnd)
        except Exception:
            return False
        if not rect:
            return False
        left, top, right, bottom = rect
        width = max(0, right - left)
        height = max(0, bottom - top)
        overlay_rect = self._overlay_rect_tuple()
        if overlay_rect is None:
            return False
        o_width = overlay_rect[2] - overlay_rect[0]
        o_height = overlay_rect[3] - overlay_rect[1]
        if o_width <= 0 or o_height <= 0:
            return False
        width_diff = abs(width - o_width)
        height_diff = abs(height - o_height)
        return width >= 400 and height >= 300 and width_diff <= 64 and height_diff <= 64

    def _detect_presentation_window(self) -> Optional[int]:
        candidates = self._gather_target_candidates()
        foreground_hwnd = self._effective_foreground_window()
        self._last_detected_role = None
        if not candidates:
            return None
        best_slideshow: Tuple[Optional[int], int] = (None, -1)
        best_document: Tuple[Optional[int], int] = (None, -1)
        for index, (hwnd, role, class_name, rect) in enumerate(candidates):
            if role == "slideshow":
                score = self._candidate_score(hwnd, class_name=class_name, rect=rect)
                if score > best_slideshow[1]:
                    best_slideshow = (hwnd, score)
            elif role == "document":
                if not self._is_document_window_valid(
                    hwnd,
                    rect=rect,
                    index=index,
                    foreground=foreground_hwnd,
                ):
                    continue
                score = self._document_candidate_score(hwnd, rect, index)
                if score > best_document[1]:
                    best_document = (hwnd, score)
        if best_slideshow[0] is not None:
            self._last_detected_role = "slideshow"
            return best_slideshow[0]
        if best_document[0] is not None:
            self._last_detected_role = "document"
            return best_document[0]
        return None

    def _resolve_presentation_target(self) -> Optional[int]:
        hwnd = self._last_target_hwnd
        if hwnd and self._is_target_window_valid(hwnd):
            if self._should_refresh_cached_target(hwnd):
                refreshed = self._detect_presentation_window()
                if refreshed and refreshed != hwnd and self._is_target_window_valid(refreshed):
                    self._set_cached_target(refreshed, self._last_detected_role)
                    return refreshed
            return hwnd
        refreshed = self._detect_presentation_window()
        if refreshed and self._is_target_window_valid(refreshed):
            self._set_cached_target(refreshed, self._last_detected_role)
            return refreshed
        self.clear_cached_target()
        return None


@dataclass
class _PendingToolRestoreState:
    mode: str
    shape: Optional[str]
    restore_on_move: bool


class OverlayWindow(QWidget):
    _KNOWN_PRESENTATION_CLASSES = _PresentationForwarder._KNOWN_PRESENTATION_CLASSES
    _KNOWN_PRESENTATION_PREFIXES = _PresentationForwarder._KNOWN_PRESENTATION_PREFIXES
    _SLIDESHOW_PRIORITY_CLASSES = _PresentationForwarder._SLIDESHOW_PRIORITY_CLASSES
    _SLIDESHOW_SECONDARY_CLASSES = _PresentationForwarder._SLIDESHOW_SECONDARY_CLASSES
    _DOCUMENT_WINDOW_CLASSES = _PresentationForwarder._DOCUMENT_WINDOW_CLASSES
    _DOCUMENT_WINDOW_PREFIXES = _PresentationForwarder._DOCUMENT_WINDOW_PREFIXES

    def __init__(self, settings_manager: SettingsManager) -> None:
        super().__init__(None, Qt.WindowType.FramelessWindowHint | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.settings_manager = settings_manager
        s = self.settings_manager.load_settings().get("Paint", {})
        self.pen_size = int(s.get("brush_size", "12"))
        self.pen_color = QColor(s.get("brush_color", "#ff0000"))
        self._mode_state = ModeState(
            tool=ToolMode.BRUSH,
            interaction=InteractionMode.DRAWING,
            passthrough=False,
            canvas_hidden=False,
            keyboard_grabbed=False,
        )
        self.mode = self._mode_state.tool.value
        self.current_shape: Optional[str] = None
        self.shape_start_point: Optional[QPoint] = None
        self.drawing = False
        self.last_point = QPointF(); self.prev_point = QPointF()
        self.last_width = float(self.pen_size); self.last_time = time.time()
        self._last_brush_color = QColor(self.pen_color)
        self._last_brush_size = max(1, self.pen_size)
        self._last_draw_mode = "brush"
        self._last_shape_type: Optional[str] = None
        self._restoring_tool = False
        self._eraser_last_point: Optional[QPoint] = None
        self._stroke_points: deque[QPointF] = deque(maxlen=8)
        self._stroke_timestamps: deque[float] = deque(maxlen=8)
        self._stroke_speed: float = 0.0
        self._stroke_last_midpoint: Optional[QPointF] = None
        self._stroke_filter_point: Optional[QPointF] = None
        self._brush_painter: Optional[QPainter] = None
        self._eraser_painter: Optional[QPainter] = None
        self._last_target_hwnd: Optional[int] = None
        self._pending_tool_restore: Optional[_PendingToolRestoreState] = None
        self._scroll_restore_pending = False
        self._scroll_restore_policy: Optional[str] = None
        self._scroll_restore_info: Optional[Tuple[str, Optional[str]]] = None
        self._last_scroll_cursor_inside_toolbar = False
        self._toolbar_hover_restore: Optional[Tuple[str, Optional[str]]] = None
        base_width = float(max(1, self.pen_size))
        self._brush_pen = QPen(
            self.pen_color,
            base_width,
            Qt.PenStyle.SolidLine,
            Qt.PenCapStyle.RoundCap,
            Qt.PenJoinStyle.RoundJoin,
        )
        fade_color = QColor(self.pen_color)
        fade_color.setAlpha(160)
        self._brush_shadow_pen = QPen(
            fade_color,
            base_width * 1.2,
            Qt.PenStyle.SolidLine,
            Qt.PenCapStyle.RoundCap,
            Qt.PenJoinStyle.RoundJoin,
        )
        self._last_preview_bounds: Optional[QRect] = None
        self.whiteboard_active = False
        self.whiteboard_color = QColor(0, 0, 0, 0); self.last_board_color = QColor("#ffffff")
        self.cursor_pixmap = QPixmap()
        self._canvas_hidden = False
        self._mode_canvas_visible = True
        self._eraser_stroker = QPainterPathStroker()
        self._eraser_stroker.setCapStyle(Qt.PenCapStyle.RoundCap)
        self._eraser_stroker.setJoinStyle(Qt.PenJoinStyle.RoundJoin)
        self._eraser_stroker_width = 0.0
        self._forwarder: Optional[_PresentationForwarder] = (
            _PresentationForwarder(self) if _PresentationForwarder.is_supported() else None
        )
        self.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.setMouseTracking(True)
        self._keyboard_grabbed = False

        self._build_scene()
        self.history: List[QPixmap] = []
        self._history_limit = 30
        self.toolbar = FloatingToolbar(self, self.settings_manager)
        self._nav_repeat_timer = QTimer(self)
        self._nav_repeat_timer.setInterval(120)
        self._nav_repeat_timer.timeout.connect(self._on_navigation_repeat_timeout)
        self._nav_repeat_delay_timer = QTimer(self)
        self._nav_repeat_delay_timer.setSingleShot(True)
        self._nav_repeat_delay_timer.setInterval(280)
        self._nav_repeat_delay_timer.timeout.connect(self._start_navigation_repeat)
        self._nav_active_vk: Optional[int] = None
        self._nav_passthrough_active = False
        self._nav_restore_cursor_pos: Optional[QPoint] = None
        self._nav_restore_timer = QTimer(self)
        self._nav_restore_timer.setInterval(80)
        self._nav_restore_timer.timeout.connect(self._check_navigation_restore)
        self._nav_forced_passthrough = False
        self._nav_restore_deadline: Optional[float] = None
        self._nav_direct_keys: Set[int] = set()
        self._nav_cursor_canvas_override = False
        self._interaction_mode = "drawing"
        self._interaction_before_navigation: Optional[str] = None
        self._navigation_origin: Optional[NavigationOriginState] = None
        self._navigation_restore_to_cursor = False
        self.set_mode(ToolMode.BRUSH.value, initial=True)
        self.toolbar.update_undo_state(False)

    def raise_toolbar(self) -> None:
        if getattr(self, "toolbar", None) is not None:
            self.toolbar.show()
            self.toolbar.raise_()

    def _is_point_inside_toolbar(self, global_point: QPoint) -> bool:
        toolbar = getattr(self, "toolbar", None)
        if toolbar is None or not toolbar.isVisible():
            return False
        try:
            top_left = toolbar.mapToGlobal(QPoint(0, 0))
        except Exception:
            return False
        rect = QRect(top_left, toolbar.size())
        return rect.contains(global_point)

    def _schedule_tool_restore(
        self,
        mode: Optional[str],
        shape: Optional[str],
        *,
        restore_on_move: bool,
    ) -> None:
        if mode not in {"brush", "shape", "eraser"}:
            self._pending_tool_restore = None
            return
        tracked_shape = shape if mode == "shape" else None
        self._pending_tool_restore = _PendingToolRestoreState(
            mode=mode,
            shape=tracked_shape,
            restore_on_move=restore_on_move,
        )

    def _set_interaction_mode(self, mode: InteractionMode) -> None:
        if self._interaction_mode != mode.value:
            self._interaction_mode = mode.value

    def _coerce_tool_mode(self, mode: Optional[Union[str, ToolMode]]) -> Optional[ToolMode]:
        if not mode:
            return None
        if isinstance(mode, ToolMode):
            return mode
        try:
            return ToolMode(str(mode))
        except ValueError:
            return None

    def _build_mode_state(self, tool: ToolMode) -> ModeState:
        passthrough = tool == ToolMode.CURSOR
        interaction = InteractionMode.CURSOR if tool == ToolMode.CURSOR else InteractionMode.DRAWING
        canvas_hidden = tool == ToolMode.CURSOR
        keyboard_grabbed = not passthrough
        return ModeState(
            tool=tool,
            interaction=interaction,
            passthrough=passthrough,
            canvas_hidden=canvas_hidden,
            keyboard_grabbed=keyboard_grabbed,
        )

    def _snapshot_mode_state(self, template: ModeState) -> ModeState:
        return ModeState(
            tool=template.tool,
            interaction=template.interaction,
            passthrough=template.passthrough,
            canvas_hidden=bool(self._canvas_hidden),
            keyboard_grabbed=bool(self._keyboard_grabbed),
        )

    def _apply_mode_visibility(
        self,
        previous: ModeState,
        target: ModeState,
        *,
        initial: bool = False,
    ) -> None:
        self._apply_input_passthrough(target.passthrough)
        self._set_canvas_hidden(target.canvas_hidden)
        if target.passthrough:
            if not self.isVisible():
                self.show()
            if not initial:
                self._release_keyboard_capture()
            return
        if not self.isVisible():
            self.show()
        self._ensure_keyboard_capture()
        if initial:
            return

    def _refresh_mode_visibility(self, *, initial: bool = False) -> None:
        current_state = getattr(self, "_mode_state", None)
        if not isinstance(current_state, ModeState):
            return
        target_state = self._build_mode_state(current_state.tool)
        self._apply_mode_visibility(current_state, target_state, initial=initial)
        self._mode_canvas_visible = not target_state.canvas_hidden
        self._mode_state = self._snapshot_mode_state(target_state)

    def _event_global_point(self, event: Optional[Any]) -> Optional[QPoint]:
        if event is None:
            return self._safe_global_cursor_pos()
        for attr in ("globalPosition", "globalPos", "globalPosF"):
            candidate = getattr(event, attr, None)
            if candidate is None:
                continue
            try:
                value = candidate() if callable(candidate) else candidate
            except Exception:
                continue
            if isinstance(value, QPoint):
                return QPoint(value)
            if isinstance(value, QPointF):
                return value.toPoint()
        return self._safe_global_cursor_pos()

    def _build_input_context(
        self,
        event: Optional[Any],
        *,
        force_toolbar: bool = False,
    ) -> InputContext:
        global_pos = self._event_global_point(event)
        inside_toolbar = True if force_toolbar else bool(
            global_pos and self._is_point_inside_toolbar(global_pos)
        )
        return InputContext(
            global_pos=global_pos,
            inside_toolbar=inside_toolbar,
            tool=self._mode_state.tool,
        )

    def _maybe_restore_tool_for_press(self, context: InputContext) -> None:
        pending = self._pending_tool_restore
        if not pending:
            return
        if context.tool == ToolMode.CURSOR or self._nav_forced_passthrough:
            if not context.inside_toolbar:
                self._restore_pending_tool()

    def _handle_move_restore(self, context: InputContext) -> None:
        pending = self._pending_tool_restore
        if not (pending and pending.restore_on_move):
            return
        if context.global_pos is None:
            return
        self._handle_pointer_navigation_restore(context.global_pos)

    def _prepare_navigation_for_pointer(self, context: InputContext, *, reason: str) -> None:
        if self._coerce_tool_mode(self.mode) is None:
            return
        if reason == "wheel":
            self.drawing = False
        self._engage_navigation_for_context(context)

    def _engage_navigation_for_context(self, context: InputContext) -> None:
        pointer = context.global_pos or self._safe_global_cursor_pos() or QPoint()
        self._engage_navigation_transition(pointer, inside_toolbar=context.inside_toolbar)

    def _set_navigation_origin(
        self,
        mode: ToolMode,
        shape: Optional[str],
        *,
        inside_toolbar: bool,
    ) -> None:
        tracked_shape = shape if mode == ToolMode.SHAPE else None
        self._navigation_origin = NavigationOriginState(
            mode=mode,
            shape=tracked_shape,
            inside_toolbar=inside_toolbar,
        )

    def _update_navigation_origin_location(self, *, inside_toolbar: bool) -> None:
        origin = self._navigation_origin
        if not origin:
            return
        self._navigation_origin = NavigationOriginState(
            mode=origin.mode,
            shape=origin.shape,
            inside_toolbar=inside_toolbar,
        )

    def _clear_navigation_origin(self) -> None:
        self._navigation_origin = None

    def _navigation_origin_from_toolbar(self) -> bool:
        origin = self._navigation_origin
        return bool(origin and origin.inside_toolbar)

    def _engage_navigation_transition(
        self,
        pointer: Optional[QPoint],
        *,
        inside_toolbar: bool,
    ) -> None:
        prev_mode = self._coerce_tool_mode(self.mode)
        if prev_mode is None:
            self._clear_navigation_passthrough()
            return
        prev_shape = self.current_shape if prev_mode == ToolMode.SHAPE else None
        if prev_mode in {ToolMode.BRUSH, ToolMode.SHAPE}:
            self._update_last_tool_snapshot()
        reference = pointer
        if reference is None:
            reference = self._safe_global_cursor_pos()
        if reference is None:
            reference = QPoint()
        restore_on_move = prev_mode != ToolMode.CURSOR and not inside_toolbar
        self._begin_navigation_passthrough(
            prev_mode.value,
            prev_shape,
            restore_on_move=restore_on_move,
            cursor_reference=reference,
            origin_inside_toolbar=inside_toolbar,
        )

    def _handle_toolbar_navigation_enter(self) -> None:
        if self.mode == ToolMode.CURSOR.value:
            return
        pointer = self._safe_global_cursor_pos() or QPoint()
        self._engage_navigation_transition(pointer, inside_toolbar=True)

    def _handle_toolbar_navigation_leave(self) -> None:
        origin = self._navigation_origin
        if not origin or not origin.inside_toolbar:
            return
        if self._pending_tool_restore:
            self._restore_pending_tool()
            return
        if origin.mode == ToolMode.CURSOR:
            self._clear_navigation_passthrough()
        elif self._interaction_mode == InteractionMode.NAVIGATION.value:
            self._clear_navigation_passthrough()

    def handle_toolbar_navigation_wheel(self, event: QWheelEvent) -> bool:
        if self.mode not in {"brush", "shape", "eraser", "cursor"}:
            return False
        self.wheelEvent(event)
        return event.isAccepted()

    def handle_toolbar_navigation_key_event(
        self, event: QKeyEvent, *, pressed: bool
    ) -> bool:
        if self._interaction_mode != InteractionMode.NAVIGATION.value:
            return False
        key_code = int(event.key())
        if key_code not in self._NAVIGATION_KEY_CODES:
            return False
        if pressed:
            self.keyPressEvent(event)
        else:
            self.keyReleaseEvent(event)
        return event.isAccepted()

    def _schedule_toolbar_navigation_cleanup(self) -> None:
        if not self._navigation_origin_from_toolbar():
            return
        if self._pending_tool_restore:
            QTimer.singleShot(0, self._restore_pending_tool)
            return
        if (
            self.mode == ToolMode.CURSOR.value
            and self._interaction_mode == InteractionMode.NAVIGATION.value
        ):
            QTimer.singleShot(0, self._clear_navigation_passthrough)

    def _handle_pointer_navigation_restore(self, global_pos: QPoint) -> None:
        pending = self._pending_tool_restore
        if not pending or not pending.restore_on_move:
            return
        origin = self._navigation_origin
        if origin and origin.inside_toolbar:
            pending.restore_on_move = False
            if self._nav_restore_timer.isActive():
                self._nav_restore_timer.stop()
            self._nav_restore_deadline = None
            return
        if not (self.mode == ToolMode.CURSOR.value or self._nav_forced_passthrough):
            return
        if self._is_point_inside_toolbar(global_pos):
            pending.restore_on_move = False
            if self._nav_restore_timer.isActive():
                self._nav_restore_timer.stop()
            self._nav_restore_deadline = None
        else:
            self._restore_pending_tool()

    def _handle_navigation_key_event(self, event: QKeyEvent) -> None:
        context = self._build_input_context(event)
        self._engage_navigation_for_context(context)

    def _handle_navigation_button_trigger(self) -> None:
        context = self._build_input_context(None, force_toolbar=True)
        self._engage_navigation_for_context(context)

    @property
    def navigation_passthrough_active(self) -> bool:
        return bool(self._nav_passthrough_active)

    def _safe_global_cursor_pos(self) -> Optional[QPoint]:
        try:
            return QCursor.pos()
        except Exception:
            return None

    _NAVIGATION_KEY_CODES: Set[int] = {
        int(Qt.Key.Key_Up),
        int(Qt.Key.Key_Down),
        int(Qt.Key.Key_Left),
        int(Qt.Key.Key_Right),
        int(Qt.Key.Key_PageUp),
        int(Qt.Key.Key_PageDown),
        int(Qt.Key.Key_Space),
        int(Qt.Key.Key_Return),
        int(Qt.Key.Key_Enter),
    }

    _NAVIGATION_QT_TO_VK: Dict[int, int] = {
        int(Qt.Key.Key_Up): VK_UP,
        int(Qt.Key.Key_Down): VK_DOWN,
        int(Qt.Key.Key_Left): VK_LEFT,
        int(Qt.Key.Key_Right): VK_RIGHT,
        int(Qt.Key.Key_PageUp): VK_PRIOR,
        int(Qt.Key.Key_PageDown): VK_NEXT,
        int(Qt.Key.Key_Space): VK_SPACE,
        int(Qt.Key.Key_Return): VK_RETURN,
        int(Qt.Key.Key_Enter): VK_RETURN,
    }

    _NAVIGATION_WHEEL_MAP: Dict[int, int] = {
        VK_UP: WHEEL_DELTA,
        VK_LEFT: WHEEL_DELTA,
        VK_PRIOR: WHEEL_DELTA,
        VK_DOWN: -WHEEL_DELTA,
        VK_RIGHT: -WHEEL_DELTA,
        VK_NEXT: -WHEEL_DELTA,
        VK_SPACE: -WHEEL_DELTA,
        VK_RETURN: -WHEEL_DELTA,
    }

    def _begin_navigation_passthrough(
        self,
        prev_mode: str,
        prev_shape: Optional[str],
        *,
        restore_on_move: bool,
        cursor_reference: Optional[QPoint] = None,
        origin_inside_toolbar: Optional[bool] = None,
    ) -> None:
        tool_mode = self._coerce_tool_mode(prev_mode)
        if tool_mode is None:
            self._clear_navigation_passthrough()
            return
        inside_toolbar = (
            bool(origin_inside_toolbar)
            if origin_inside_toolbar is not None
            else not restore_on_move
        )
        if (
            self._nav_passthrough_active
            and self._navigation_origin_mode == prev_mode
            and self._navigation_origin_inside_toolbar == inside_toolbar
            and (
                prev_mode != "shape"
                or self._navigation_origin_shape == (prev_shape if prev_mode == "shape" else None)
            )
        ):
            if prev_mode == "cursor" and self._canvas_hidden:
                self._nav_cursor_canvas_override = True
                self._set_canvas_hidden(False)
            if cursor_reference is not None:
                self._nav_restore_cursor_pos = QPoint(cursor_reference)
            return

        self._set_navigation_origin(tool_mode, prev_shape, inside_toolbar=inside_toolbar)
        if not self._nav_passthrough_active:
            self._interaction_before_navigation = self._interaction_mode
        self._interaction_mode = "navigation"
        self._navigation_origin_mode = prev_mode
        self._navigation_origin_shape = prev_shape if prev_mode == "shape" else None
        self._navigation_origin_inside_toolbar = inside_toolbar
        restore_cursor_passthrough = prev_mode == "cursor" and not self.whiteboard_active
        self._navigation_restore_to_cursor = restore_cursor_passthrough
        self._nav_passthrough_active = True
        if cursor_reference is not None:
            self._nav_restore_cursor_pos = QPoint(cursor_reference)
        else:
            pos = self._safe_global_cursor_pos()
            self._nav_restore_cursor_pos = QPoint(pos) if isinstance(pos, QPoint) else pos
        nav_from_drawing = prev_mode != "cursor"
        needs_transparent_state = nav_from_drawing or restore_cursor_passthrough
        was_transparent = False
        if needs_transparent_state:
            was_transparent = self.testAttribute(
                Qt.WidgetAttribute.WA_TransparentForMouseEvents
            ) or bool(
                self.windowFlags()
                & Qt.WindowType.WindowTransparentForInput
            )
        if nav_from_drawing:
            if was_transparent:
                self._apply_input_passthrough(False)
            self._nav_forced_passthrough = True
            if not self._keyboard_grabbed:
                self._ensure_keyboard_capture()
        else:
            self._nav_forced_passthrough = False
            if restore_cursor_passthrough:
                if was_transparent:
                    self._apply_input_passthrough(False)
                if not self._keyboard_grabbed:
                    self._ensure_keyboard_capture()
        if prev_mode == "cursor":
            should_restore_canvas = bool(self._canvas_hidden) or self._nav_cursor_canvas_override
            self._nav_cursor_canvas_override = should_restore_canvas
            if should_restore_canvas and self._canvas_hidden:
                self._set_canvas_hidden(False)
            restore_mode, restore_shape = self._cursor_navigation_restore_target()
            if restore_mode:
                self._schedule_tool_restore(
                    restore_mode,
                    restore_shape,
                    restore_on_move=False,
                )
            else:
                self._pending_tool_restore = None
            self._nav_restore_timer.stop()
        else:
            self._nav_cursor_canvas_override = False
            self._schedule_tool_restore(prev_mode, prev_shape, restore_on_move=restore_on_move)
            pending = self._pending_tool_restore
            if restore_on_move and pending:
                if self._nav_restore_cursor_pos is None:
                    pos = self._safe_global_cursor_pos()
                    self._nav_restore_cursor_pos = QPoint(pos) if isinstance(pos, QPoint) else pos
                self._nav_restore_deadline = time.monotonic() + NAVIGATION_RESTORE_TIMEOUT
                if not self._nav_restore_timer.isActive():
                    self._nav_restore_timer.start()
            else:
                self._nav_restore_deadline = None
                self._nav_restore_timer.stop()
        self.update_toolbar_state()

    def _is_passthrough_enabled(self) -> bool:
        try:
            attr_passthrough = self.testAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        except Exception:
            attr_passthrough = False
        window_passthrough = False
        if hasattr(Qt.WindowType, "WindowTransparentForInput"):
            try:
                window_passthrough = self._has_window_input_passthrough()
            except Exception:
                window_passthrough = False
        return bool(attr_passthrough or window_passthrough)

    @contextlib.contextmanager
    def _navigation_input_transfer(self):
        original_passthrough = self._is_passthrough_enabled()
        had_keyboard = self._keyboard_grabbed
        if original_passthrough:
            if had_keyboard:
                self._release_keyboard_capture()
        else:
            self._apply_input_passthrough(True)
        try:
            yield
        finally:
            if original_passthrough:
                if had_keyboard and not self._keyboard_grabbed and self.mode != "cursor":
                    self._ensure_keyboard_capture()
            else:
                self._apply_input_passthrough(False)
                if had_keyboard and self.mode != "cursor":
                    self._ensure_keyboard_capture()
                if self.mode != "cursor":
                    self.update()

    def _clear_navigation_passthrough(self) -> None:
        self._nav_passthrough_active = False
        self._nav_restore_cursor_pos = None
        if self._nav_restore_timer.isActive():
            self._nav_restore_timer.stop()
        self._nav_restore_deadline = None
        if self._nav_forced_passthrough and self.mode != ToolMode.CURSOR.value:
            self._apply_input_passthrough(False)
            self._ensure_keyboard_capture()
        restore_cursor_passthrough = (
            self._navigation_restore_to_cursor
            and self.mode == "cursor"
            and not self.whiteboard_active
        )
        self._nav_forced_passthrough = False
        previous = self._interaction_before_navigation
        if previous in {InteractionMode.DRAWING.value, InteractionMode.CURSOR.value}:
            self._interaction_mode = previous
        elif self.mode == ToolMode.CURSOR.value:
            self._set_interaction_mode(InteractionMode.CURSOR)
        else:
            self._set_interaction_mode(InteractionMode.DRAWING)
        self._interaction_before_navigation = None
        self._navigation_origin_mode = None
        self._navigation_origin_shape = None
        self._navigation_origin_inside_toolbar = False
        self._clear_navigation_origin()
        if restore_cursor_passthrough:
            self._apply_input_passthrough(True)
        self._navigation_restore_to_cursor = False
        if self._nav_cursor_canvas_override and self.mode == ToolMode.CURSOR.value:
            self._set_canvas_hidden(True)
        self._nav_cursor_canvas_override = False
        self.update_toolbar_state()

    def _check_navigation_restore(self) -> None:
        pending = self._pending_tool_restore
        if pending and pending.restore_on_move and self._navigation_origin_from_toolbar():
            pending.restore_on_move = False
            self._nav_restore_timer.stop()
            self._nav_restore_deadline = None
            return
        if not (
            self._nav_passthrough_active
            and pending
            and pending.restore_on_move
            and not self._navigation_origin_from_toolbar()
            and (self.mode == ToolMode.CURSOR.value or self._nav_forced_passthrough)
        ):
            if not (pending and pending.restore_on_move):
                self._nav_restore_timer.stop()
                self._nav_restore_deadline = None
            return
        current_pos = self._safe_global_cursor_pos()
        if current_pos is None:
            self._nav_restore_timer.stop()
            self._nav_restore_deadline = None
            self._restore_pending_tool()
            return
        if self._nav_restore_cursor_pos is None:
            self._nav_restore_cursor_pos = QPoint(current_pos)
            return
        if current_pos == self._nav_restore_cursor_pos:
            deadline = self._nav_restore_deadline
            if deadline is not None and time.monotonic() >= deadline:
                self._nav_restore_timer.stop()
                self._nav_restore_deadline = None
                self._restore_pending_tool()
            return
        if self._is_point_inside_toolbar(current_pos):
            pending.restore_on_move = False
            self._nav_restore_timer.stop()
            self._nav_restore_deadline = None
            return
        self._nav_restore_deadline = None
        self._restore_pending_tool()

    def _restore_pending_tool(self) -> None:
        pending = self._pending_tool_restore
        if not pending:
            return
        self._pending_tool_restore = None
        self._clear_navigation_passthrough()
        self._restore_last_tool(pending.mode, shape_type=pending.shape)

    def _after_navigation_wheel(self, event: QWheelEvent) -> None:
        if self.whiteboard_active:
            return
        prev_mode = self.mode
        if prev_mode not in {"brush", "shape", "eraser", "cursor"}:
            return
        prev_shape = self.current_shape if prev_mode == "shape" else None
        if prev_mode in {"brush", "shape"}:
            self._update_last_tool_snapshot()
        try:
            global_pos = event.globalPosition().toPoint()
        except Exception:
            fallback = self._safe_global_cursor_pos()
            global_pos = fallback if fallback is not None else QPoint()
        inside_toolbar = self._is_point_inside_toolbar(global_pos)
        restore_on_move = prev_mode != "cursor" and not inside_toolbar
        self.drawing = False
        self._begin_navigation_passthrough(
            prev_mode,
            prev_shape,
            restore_on_move=restore_on_move,
            cursor_reference=global_pos,
            origin_inside_toolbar=inside_toolbar,
        )

    def _after_navigation_key(self, event: QKeyEvent) -> None:
        if self.whiteboard_active:
            return
        key_code = int(event.key())
        if key_code not in self._NAVIGATION_KEY_CODES:
            return
        modifiers = event.modifiers()
        if modifiers & (
            Qt.KeyboardModifier.ControlModifier
            | Qt.KeyboardModifier.AltModifier
            | Qt.KeyboardModifier.MetaModifier
        ):
            return
        prev_mode = self.mode
        if prev_mode not in {"brush", "shape", "eraser", "cursor"}:
            return
        context = self._build_input_context(event)
        self._engage_navigation_for_context(context)

    def _build_scene(self) -> None:
        virtual = QRect()
        for screen in QApplication.screens():
            virtual = virtual.united(screen.geometry())
        self.setGeometry(virtual)
        self.canvas = QPixmap(self.size()); self.canvas.fill(Qt.GlobalColor.transparent)
        self.temp_canvas = QPixmap(self.size()); self.temp_canvas.fill(Qt.GlobalColor.transparent)
        self._canvas_hidden = False

    # ---- ͼ��ͼ����� ----
    def _ensure_brush_painter(self) -> QPainter:
        painter = self._brush_painter
        if painter is None:
            painter = QPainter(self.canvas)
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            self._brush_painter = painter
        return painter

    def _release_brush_painter(self) -> None:
        if self._brush_painter is not None:
            self._brush_painter.end()
            self._brush_painter = None

    def _ensure_eraser_painter(self) -> QPainter:
        painter = self._eraser_painter
        if painter is None:
            painter = QPainter(self.canvas)
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            painter.setCompositionMode(QPainter.CompositionMode.CompositionMode_Clear)
            self._eraser_painter = painter
        return painter

    def _release_eraser_painter(self) -> None:
        if self._eraser_painter is not None:
            self._eraser_painter.end()
            self._eraser_painter = None

    def _release_canvas_painters(self) -> None:
        self._release_brush_painter()
        self._release_eraser_painter()

    def _set_canvas_hidden(self, hidden: bool) -> None:
        """Switch rendering visibility for the main canvas without losing content."""
        hidden = bool(hidden)
        if hidden == self._canvas_hidden and self._mode_canvas_visible == (not hidden):
            return
        self._canvas_hidden = hidden
        self._mode_canvas_visible = not hidden
        self.update()

    def _update_brush_pen_appearance(self, width: float, fade_alpha: int) -> None:
        width = max(0.6, float(width))
        fade_alpha = max(0, min(255, int(fade_alpha)))
        self._brush_pen.setColor(self.pen_color)
        self._brush_pen.setWidthF(width)
        fade_color = QColor(self.pen_color)
        fade_color.setAlpha(fade_alpha)
        self._brush_shadow_pen.setColor(fade_color)
        self._brush_shadow_pen.setWidthF(width * 1.25)

    def show_overlay(self) -> None:
        self.show(); self.toolbar.show(); self.raise_toolbar()
        self.set_mode(self.mode, self.current_shape)

    def hide_overlay(self) -> None:
        self._release_keyboard_capture()
        self.hide(); self.toolbar.hide()
        self.save_settings(); self.save_window_position()

    def open_pen_settings(self) -> None:
        pm, ps = self.mode, self.current_shape
        d = PenSettingsDialog(self.toolbar, self.pen_size, self.pen_color.name())
        if d.exec():
            self.pen_size, self.pen_color = d.get_settings()
            self._update_brush_pen_appearance(max(1.0, float(self.pen_size)), 180)
            self.update_cursor()
        self.set_mode(pm, ps)
        self.raise_toolbar()

    def open_board_color_dialog(self) -> None:
        d = BoardColorDialog(self.toolbar)
        if d.exec():
            c = d.get_color()
            if c:
                self.last_board_color = c
                self.whiteboard_color = c
                self.whiteboard_active = True
                if self._forwarder:
                    self._forwarder.clear_cached_target()
                self.toolbar.update_whiteboard_button_state(True)
                self._refresh_mode_visibility(initial=False)
                self.raise_toolbar()
                self.update()

    def toggle_whiteboard(self) -> None:
        self.whiteboard_active = not self.whiteboard_active
        self.whiteboard_color = self.last_board_color if self.whiteboard_active else QColor(0, 0, 0, 0)
        if self._forwarder and self.whiteboard_active:
            self._forwarder.clear_cached_target()
        self._refresh_mode_visibility(initial=False)
        self.raise_toolbar()
        self.toolbar.update_whiteboard_button_state(self.whiteboard_active)
        self.update()

    def set_mode(
        self,
        mode: Union[str, ToolMode],
        shape_type: Optional[str] = None,
        *,
        initial: bool = False,
    ) -> None:
        target_mode = self._coerce_tool_mode(mode)
        if target_mode is None:
            return
        prev_state = getattr(self, "_mode_state", None)
        if not isinstance(prev_state, ModeState):
            prev_state = self._build_mode_state(target_mode)
        if target_mode != ToolMode.CURSOR or self._nav_passthrough_active:
            self._clear_navigation_passthrough()
        if prev_state.tool != target_mode:
            self._release_canvas_painters()
        focus_on_cursor = bool(self._forwarder) and target_mode == ToolMode.CURSOR and not initial
        state_template = self._build_mode_state(target_mode)
        self.mode = state_template.tool.value
        if state_template.interaction == InteractionMode.CURSOR:
            self._set_interaction_mode(InteractionMode.CURSOR)
            self._clear_navigation_origin()
            self._navigation_restore_to_cursor = False
        else:
            self._set_interaction_mode(InteractionMode.DRAWING)
        if not self._restoring_tool:
            self._pending_tool_restore = None
        if shape_type is not None or state_template.tool != ToolMode.SHAPE:
            self.current_shape = shape_type
        if state_template.tool != ToolMode.SHAPE:
            self.shape_start_point = None
            self._last_preview_bounds = None
        if state_template.tool != ToolMode.ERASER:
            self._eraser_last_point = None
        if self._forwarder and state_template.tool == ToolMode.CURSOR:
            self._forwarder.clear_cached_target()
        if state_template.tool in {ToolMode.BRUSH, ToolMode.SHAPE} and not self._restoring_tool:
            self._update_last_tool_snapshot()
        self._apply_mode_visibility(prev_state, state_template, initial=initial)
        self._mode_canvas_visible = not state_template.canvas_hidden
        self._mode_state = self._snapshot_mode_state(state_template)
        if focus_on_cursor and self._forwarder:
            self._forwarder.focus_presentation_window()
        if not initial:
            self.raise_toolbar()
        self.update_toolbar_state()
        self.update_cursor()

    def update_toolbar_state(self) -> None:
        if not getattr(self, 'toolbar', None):
            return
        interaction = getattr(self, "_interaction_mode", InteractionMode.DRAWING.value)
        navigation_from_cursor = (
            interaction == InteractionMode.NAVIGATION.value
            and self._navigation_origin is not None
            and self._navigation_origin.mode == ToolMode.CURSOR
        )
        highlight_mode: Optional[str] = None
        if (
            interaction == InteractionMode.NAVIGATION.value
            and navigation_from_cursor
        ):
            restore_mode, _ = self._cursor_navigation_restore_target()
            highlight_mode = restore_mode if restore_mode is not None else ""
        self.toolbar.update_tool_states(
            self.mode,
            interaction,
            self.pen_color,
            navigation_from_cursor=navigation_from_cursor,
            highlight_mode=highlight_mode,
        )

    def _update_last_tool_snapshot(self) -> None:
        if self.pen_color.isValid():
            self._last_brush_color = QColor(self.pen_color)
        if self.pen_size > 0:
            self._last_brush_size = max(1, int(self.pen_size))
        if self.mode in {ToolMode.BRUSH.value, ToolMode.SHAPE.value}:
            self._last_draw_mode = self.mode
            if self.mode == ToolMode.SHAPE.value:
                self._last_shape_type = self.current_shape

    def _cursor_navigation_restore_target(self) -> Tuple[Optional[str], Optional[str]]:
        last_mode = getattr(self, "_last_draw_mode", None)
        restore_mode: Optional[str]
        restore_shape: Optional[str] = None
        if last_mode in {"brush", "shape", "eraser"}:
            restore_mode = last_mode
            if restore_mode == "shape":
                restore_shape = self._last_shape_type or self.current_shape
        else:
            restore_mode = "brush"
        if restore_mode == "shape" and not restore_shape:
            restore_shape = self.current_shape or self._last_shape_type
        return restore_mode, restore_shape

    def _restore_last_tool(self, preferred_mode: Optional[str] = None, *, shape_type: Optional[str] = None) -> None:
        self._pending_tool_restore = None
        if isinstance(self._last_brush_color, QColor) and self._last_brush_color.isValid():
            self.pen_color = QColor(self._last_brush_color)
        if isinstance(self._last_brush_size, int) and self._last_brush_size > 0:
            self.pen_size = max(1, int(self._last_brush_size))
        target_mode = preferred_mode
        target_shape: Optional[str] = None
        if target_mode == "shape":
            target_shape = shape_type or self._last_shape_type or self.current_shape
        if target_mode not in {"brush", "shape", "eraser"}:
            if self._last_draw_mode == "shape":
                target_mode = "shape"
                target_shape = shape_type or self._last_shape_type or self.current_shape
            else:
                target_mode = "brush"
        self._restoring_tool = True
        try:
            self.set_mode(target_mode, shape_type=target_shape)
        finally:
            self._restoring_tool = False
        if target_mode in {"brush", "shape"}:
            self._update_last_tool_snapshot()

    def toggle_eraser_mode(self) -> None:
        """切换橡皮模式；再次点击会恢复上一次的画笔配置。"""
        if self.mode == ToolMode.ERASER.value:
            self._restore_last_tool()
        else:
            self._update_last_tool_snapshot()
            self.set_mode(ToolMode.ERASER.value)

    def toggle_cursor_mode(self) -> None:
        """切换光标模式；再次点击在光标与导航之间切换。"""
        if self.mode == "cursor":
            pointer = self._safe_global_cursor_pos() or QPoint()
            if (
                self._interaction_mode == "navigation"
                and self._navigation_origin_mode == "cursor"
            ):
                self._clear_navigation_passthrough()
            else:
                self._begin_navigation_passthrough(
                    "cursor",
                    None,
                    restore_on_move=False,
                    cursor_reference=pointer,
                    origin_inside_toolbar=self._is_point_inside_toolbar(pointer),
                )
            return
        self._update_last_tool_snapshot()
        self.set_mode(ToolMode.CURSOR.value)

    def go_to_next_slide(self) -> None:
        self._send_slide_virtual_key(VK_DOWN)

    def go_to_previous_slide(self) -> None:
        self._send_slide_virtual_key(VK_UP)

    def _send_slide_virtual_key(
        self, vk_code: int, *, repeated: bool = False, engaged: bool = False
    ) -> bool:
        if self.whiteboard_active:
            return False
        if vk_code == 0:
            return False
        if repeated:
            success, had_grab = self._dispatch_navigation_session(vk_code=vk_code)
            if success and self.mode != ToolMode.CURSOR.value and not had_grab:
                QTimer.singleShot(100, self._ensure_keyboard_capture)
            if not success:
                self._pending_tool_restore = None
                self._clear_navigation_passthrough()
            return success
        if not engaged:
            if self._coerce_tool_mode(self.mode) is None:
                return False
            context = self._build_input_context(None)
            self._engage_navigation_for_context(context)
        success, had_grab = self._dispatch_navigation_session(vk_code=vk_code)
        if not success:
            self._pending_tool_restore = None
            self._clear_navigation_passthrough()
            return False
        if self.mode != ToolMode.CURSOR.value and not had_grab:
            QTimer.singleShot(100, self._ensure_keyboard_capture)
        self.raise_toolbar()
        return True

    def handle_navigation_button_press(self, direction: str) -> None:
        self._nav_repeat_delay_timer.stop()
        self._nav_repeat_timer.stop()
        if self.whiteboard_active:
            self._nav_active_vk = None
            return
        vk_code = VK_DOWN if direction.lower() == "down" else VK_UP
        self._handle_navigation_button_trigger()
        if self._send_slide_virtual_key(vk_code, engaged=True):
            self._nav_active_vk = vk_code
            self._nav_repeat_delay_timer.start()
        else:
            self._nav_active_vk = None
            self._nav_repeat_delay_timer.stop()
            self._nav_repeat_timer.stop()

    def handle_navigation_button_release(self) -> None:
        self._nav_repeat_delay_timer.stop()
        self._nav_repeat_timer.stop()
        self._nav_active_vk = None
        pending = self._pending_tool_restore
        if pending:
            if getattr(self, "toolbar", None) is not None and self.toolbar.underMouse():
                return
            self._restore_pending_tool()
        elif (
            self._interaction_mode == InteractionMode.NAVIGATION.value
            and self.mode == ToolMode.CURSOR.value
        ):
            self._clear_navigation_passthrough()

    def _start_navigation_repeat(self) -> None:
        if self._nav_active_vk is None:
            return
        if not self._nav_repeat_timer.isActive():
            self._nav_repeat_timer.start()

    def _on_navigation_repeat_timeout(self) -> None:
        vk_code = self._nav_active_vk
        if vk_code is None:
            self._nav_repeat_timer.stop()
            return
        if not self._send_slide_virtual_key(vk_code, repeated=True):
            self._nav_repeat_timer.stop()
            self._nav_active_vk = None

    def _navigation_wheel_delta(self, vk_code: int) -> int:
        return self._NAVIGATION_WHEEL_MAP.get(vk_code, 0)

    def _send_navigation_wheel_message_only(self, delta: int) -> bool:
        if delta == 0 or self._forwarder is None:
            return False
        try:
            return self._forwarder.send_navigation_wheel_message(delta)
        except Exception:
            return False

    def _send_navigation_wheel(
        self,
        vk_code: int,
        delta: int,
        *,
        injection_only: bool = False,
    ) -> bool:
        if delta == 0:
            return False
        delivered = False
        if self._forwarder is not None:
            try:
                delivered = self._forwarder.send_navigation_wheel(
                    delta,
                    injection_only=injection_only,
                )
            except Exception:
                delivered = False
            if delivered:
                return True
            if not injection_only:
                try:
                    if self._forwarder.get_last_target_role() == "document":
                        delivered = self._forwarder.send_document_scroll_for_vk(vk_code)
                except Exception:
                    delivered = False
                if delivered:
                    return True
        focused = self._focus_presentation_window_fallback()
        if not focused:
            self._focus_presentation_window_fallback()
        return self._fallback_scroll_with_wheel(delta)

    def _send_virtual_key_direct(self, vk_code: int) -> bool:
        if vk_code == 0:
            return False
        if self._forwarder is not None:
            try:
                delivered = self._forwarder.send_virtual_key(vk_code)
            except Exception:
                delivered = False
            else:
                if delivered:
                    return True
            try:
                self._forwarder.clear_cached_target()
            except Exception:
                pass
        self._focus_presentation_window_fallback()
        return self._fallback_send_virtual_key(vk_code)

    def _dispatch_virtual_key(self, vk_code: int) -> bool:
        if vk_code == 0:
            return False
        wheel_delta = self._navigation_wheel_delta(vk_code)
        with self._navigation_input_transfer():
            success = self._perform_navigation_dispatch(
                vk_code,
                wheel_delta=wheel_delta,
                prefer_scroll=bool(wheel_delta),
            )
        if success and self.mode != "cursor":
            QTimer.singleShot(100, self._ensure_keyboard_capture)
        return success

    def cancel_pending_tool_restore(self) -> None:
        self._pending_tool_restore = None
        self._clear_navigation_passthrough()

    def on_toolbar_mouse_enter(self) -> None:
        if self.mode == ToolMode.CURSOR.value:
            self.raise_toolbar()
            return
        if self._interaction_mode != InteractionMode.NAVIGATION.value:
            self._handle_toolbar_navigation_enter()
        else:
            if self._pending_tool_restore:
                self._pending_tool_restore.restore_on_move = False
                if self._nav_restore_timer.isActive():
                    self._nav_restore_timer.stop()
                self._nav_restore_deadline = None
            self._update_navigation_origin_location(inside_toolbar=True)
        self.raise_toolbar()

    def on_toolbar_mouse_leave(self) -> None:
        self._nav_repeat_delay_timer.stop()
        self._nav_repeat_timer.stop()
        self._nav_active_vk = None
        if getattr(self, "toolbar", None) is not None and self.toolbar.underMouse():
            return
        if self._interaction_mode != InteractionMode.NAVIGATION.value:
            return
        self._handle_toolbar_navigation_leave()

    def _fallback_scroll_with_wheel(self, delta: int) -> bool:
        if delta == 0:
            return False
        if win32api is not None and win32con is not None:
            try:
                win32api.mouse_event(win32con.MOUSEEVENTF_WHEEL, 0, 0, delta, 0)
                return True
            except Exception:
                pass
        if _USER32 is None:
            return False
        try:
            _USER32.mouse_event(0x0800, 0, 0, delta, 0)
            return True
        except Exception:
            return False

    def _fallback_send_virtual_key(self, vk_code: int) -> bool:
        if vk_code == 0 or _USER32 is None:
            return False
        if self._forwarder and self._forwarder.get_last_target_role() == "document":
            if self._forwarder.send_document_scroll_for_vk(vk_code):
                return True
        wheel_delta = 0
        if vk_code in {VK_UP, VK_PRIOR}:
            wheel_delta = WHEEL_DELTA
        elif vk_code in {VK_DOWN, VK_NEXT}:
            wheel_delta = -WHEEL_DELTA
        if wheel_delta and self._fallback_scroll_with_wheel(wheel_delta):
            return True
        try:
            scan_code = _USER32.MapVirtualKeyW(vk_code, 0) if hasattr(_USER32, "MapVirtualKeyW") else 0
        except Exception:
            scan_code = 0
        flags = KEYEVENTF_EXTENDEDKEY if vk_code in _NAVIGATION_EXTENDED_KEYS else 0
        try:
            _USER32.keybd_event(vk_code, scan_code, flags, 0)
            _USER32.keybd_event(vk_code, scan_code, flags | KEYEVENTF_KEYUP, 0)
            return True
        except Exception:
            return False

    def update_cursor(self) -> None:
        if self.mode == "cursor":
            self.setCursor(Qt.CursorShape.ArrowCursor); return
        if self.mode == "shape":
            self.setCursor(Qt.CursorShape.CrossCursor); return
        d = max(10, int(self.pen_size * (3.2 if self.mode == "eraser" else 2.2)))
        self.cursor_pixmap = QPixmap(d, d); self.cursor_pixmap.fill(Qt.GlobalColor.transparent)
        p = QPainter(self.cursor_pixmap); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        if self.mode == "eraser":
            p.setBrush(QBrush(Qt.GlobalColor.white)); p.setPen(QPen(QColor("#555"), 2))
        else:
            p.setBrush(QBrush(self.pen_color)); p.setPen(QPen(Qt.GlobalColor.black, 2))
        p.drawEllipse(1, 1, d - 2, d - 2); p.end()
        self.setCursor(QCursor(self.cursor_pixmap, d // 2, d // 2))

    def _apply_dirty_region(self, region: Optional[Union[QRect, QRectF]]) -> None:
        if not region:
            return
        if isinstance(region, QRectF):
            rect = region.toAlignedRect()
        else:
            rect = QRect(region)
        if rect.isNull() or rect.width() <= 0 or rect.height() <= 0:
            return
        inflated = rect.adjusted(-4, -4, 4, 4)
        target = inflated.intersected(self.rect())
        if target.isValid() and not target.isNull():
            self.update(target)
        else:
            self.update()

    def _shape_dirty_bounds(
        self,
        start_point: Optional[QPoint],
        end_point: Optional[Union[QPoint, QPointF]],
        pen_width: int,
    ) -> Optional[QRect]:
        if start_point is None or end_point is None:
            return None
        if isinstance(end_point, QPointF):
            end = end_point.toPoint()
        else:
            end = end_point
        rect = QRect(start_point, end).normalized()
        if rect.isNull():
            rect = QRect(end, end)
        margin = max(4, int(max(1, pen_width) * 2))
        return rect.adjusted(-margin, -margin, margin, margin)

    # ---- 系统级穿透 ----
    def _has_window_input_passthrough(self) -> bool:
        """Return whether the overlay currently ignores input at the window level."""
        try:
            flag_value = Qt.WindowType.WindowTransparentForInput
        except AttributeError:
            return False
        try:
            flags = self.windowFlags()
        except Exception:
            return False
        try:
            flags_int = int(flags)
        except (TypeError, ValueError):
            try:
                flags_int = int(Qt.WindowType(flags))
            except Exception:
                return False
        return bool(flags_int & int(flag_value))

    def _apply_input_passthrough(self, enabled: bool) -> None:
        """Toggle system-level input passthrough for cursor/navigation mode."""
        attr_changed = self.testAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents) != enabled
        if attr_changed:
            self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, enabled)
        flag_supported = hasattr(Qt.WindowType, "WindowTransparentForInput")
        flag_changed = False
        if flag_supported:
            has_flag = self._has_window_input_passthrough()
            if has_flag != enabled:
                self.setWindowFlag(Qt.WindowType.WindowTransparentForInput, enabled)
                flag_changed = True
        if enabled and self._keyboard_grabbed:
            self._release_keyboard_capture()
        if (attr_changed or flag_changed) and self.isVisible():
            super().show()  # Force Qt to apply the new flags

    def _ensure_keyboard_capture(self) -> None:
        if not self._keyboard_grabbed:
            try:
                self.grabKeyboard()
                self._keyboard_grabbed = True
            except Exception:
                self._keyboard_grabbed = False
        try:
            self.raise_()
            self.activateWindow()
        except Exception:
            pass
        self.setFocus(Qt.FocusReason.ActiveWindowFocusReason)

    def _release_keyboard_capture(self) -> None:
        if not self._keyboard_grabbed:
            return
        try:
            self.releaseKeyboard()
        except Exception:
            pass
        self._keyboard_grabbed = False

    @contextlib.contextmanager
    def _temporarily_release_keyboard(self):
        had_keyboard_grab = self._keyboard_grabbed
        if had_keyboard_grab:
            self._release_keyboard_capture()
        try:
            yield had_keyboard_grab
        finally:
            if had_keyboard_grab:
                self._ensure_keyboard_capture()

    @contextlib.contextmanager
    def _navigation_injection_guard(self):
        """Maintain keyboard grab state while navigation events are dispatched."""
        had_keyboard_grab = self._keyboard_grabbed
        try:
            yield had_keyboard_grab
        finally:
            if had_keyboard_grab and not self._keyboard_grabbed:
                self._ensure_keyboard_capture()

    def _extract_wheel_delta(self, event: Optional[QWheelEvent]) -> int:
        if event is None:
            return 0
        delta = 0
        try:
            angle_delta = event.angleDelta()
        except Exception:
            angle_delta = QPoint()
        if isinstance(angle_delta, QPoint):
            delta = int(angle_delta.y() or angle_delta.x() or 0)
        if delta == 0:
            try:
                pixel_delta = event.pixelDelta()
            except Exception:
                pixel_delta = QPoint()
            if isinstance(pixel_delta, QPoint):
                delta = int(pixel_delta.y() or pixel_delta.x() or 0)
        return delta

    def _forward_wheel_event(self, event: QWheelEvent) -> bool:
        delivered = False
        if self._forwarder is not None:
            try:
                delivered = self._forwarder.forward_wheel(event)
            except Exception:
                delivered = False
            if delivered:
                return True
            try:
                self._forwarder.clear_cached_target()
            except Exception:
                pass
        delta = self._extract_wheel_delta(event)
        if delta == 0:
            return False
        focused = self._focus_presentation_window_fallback()
        if not focused:
            self._focus_presentation_window_fallback()
        return self._fallback_scroll_with_wheel(delta)

    def _perform_navigation_dispatch(
        self,
        vk_code: Optional[int],
        *,
        wheel_delta: int,
        prefer_scroll: bool,
    ) -> bool:
        if vk_code is None or vk_code == 0:
            return False
        attempts: List[Callable[[], bool]] = []
        added_attempts: Set[Tuple[Any, ...]] = set()

        def add_attempt(key: Tuple[Any, ...], func: Callable[[], bool]) -> None:
            if key in added_attempts:
                return
            added_attempts.add(key)
            attempts.append(func)

        if wheel_delta and prefer_scroll:
            add_attempt(
                ("wheel_message", wheel_delta),
                lambda delta=wheel_delta: self._send_navigation_wheel_message_only(delta),
            )
            add_attempt(
                ("wheel_injection", vk_code, wheel_delta, True),
                lambda code=vk_code, delta=wheel_delta: self._send_navigation_wheel(
                    code,
                    delta,
                    injection_only=True,
                ),
            )

        if wheel_delta:
            add_attempt(
                ("wheel_injection", vk_code, wheel_delta, False),
                lambda code=vk_code, delta=wheel_delta: self._send_navigation_wheel(
                    code,
                    delta,
                ),
            )

        add_attempt(
            ("virtual_key", vk_code),
            lambda code=vk_code: self._send_virtual_key_direct(code),
        )

        for attempt in attempts:
            try:
                delivered = attempt()
            except Exception:
                delivered = False
            if delivered:
                return True
        return False

    @dataclass
    class _NavigationActionSession:
        overlay: "OverlayWindow"
        vk_code: Optional[int]
        wheel_delta: int
        wheel_event: Optional[QWheelEvent]
        prefer_scroll: bool
        had_keyboard_grab: bool
        success: bool = False

        def execute(self) -> bool:
            if self.wheel_event is not None:
                self.success = self.overlay._forward_wheel_event(self.wheel_event)
                return self.success
            if self.vk_code is None:
                self.success = False
                return False
            self.success = self.overlay._perform_navigation_dispatch(
                self.vk_code,
                wheel_delta=self.wheel_delta,
                prefer_scroll=self.prefer_scroll,
            )
            return self.success

    @contextlib.contextmanager
    def _navigation_action_context(
        self,
        *,
        vk_code: Optional[int] = None,
        wheel_event: Optional[QWheelEvent] = None,
        prefer_scroll: Optional[bool] = None,
    ) -> Iterable["OverlayWindow._NavigationActionSession"]:
        wheel_delta = self._navigation_wheel_delta(vk_code) if vk_code else 0
        resolved_prefer_scroll = prefer_scroll
        if vk_code is not None and resolved_prefer_scroll is None:
            resolved_prefer_scroll = bool(wheel_delta) and (
                self.mode == ToolMode.CURSOR.value
                or self._interaction_mode == InteractionMode.NAVIGATION.value
            )
        with self._navigation_injection_guard() as had_keyboard_grab:
            session = self._NavigationActionSession(
                overlay=self,
                vk_code=vk_code,
                wheel_delta=wheel_delta,
                wheel_event=wheel_event,
                prefer_scroll=bool(resolved_prefer_scroll) if resolved_prefer_scroll is not None else False,
                had_keyboard_grab=had_keyboard_grab,
            )
            yield session

    def _dispatch_navigation_session(
        self,
        *,
        vk_code: Optional[int] = None,
        wheel_event: Optional[QWheelEvent] = None,
        prefer_scroll: Optional[bool] = None,
    ) -> Tuple[bool, bool]:
        if self.whiteboard_active:
            return False, self._keyboard_grabbed
        with self._navigation_action_context(
            vk_code=vk_code,
            wheel_event=wheel_event,
            prefer_scroll=prefer_scroll,
        ) as session:
            handled = session.execute()
        return handled, session.had_keyboard_grab

    def _overlay_rect_tuple(self) -> Optional[Tuple[int, int, int, int]]:
        rect = self.geometry()
        if rect.isNull():
            return None
        left = rect.left()
        top = rect.top()
        right = left + rect.width()
        bottom = top + rect.height()
        return left, top, right, bottom

    def _rect_intersects_overlay(self, rect: Tuple[int, int, int, int]) -> bool:
        overlay = self._overlay_rect_tuple()
        if overlay is None:
            return False
        left, top, right, bottom = rect
        o_left, o_top, o_right, o_bottom = overlay
        return not (right <= o_left or left >= o_right or bottom <= o_top or top >= o_bottom)

    def _fallback_belongs_to_current_process(self, hwnd: int) -> bool:
        if _USER32 is None or hwnd == 0:
            return False
        pid = wintypes.DWORD()
        try:
            _USER32.GetWindowThreadProcessId(wintypes.HWND(hwnd), ctypes.byref(pid))
        except Exception:
            return False
        return pid.value == os.getpid()

    def _fallback_next_window_in_z_order(self, hwnd: int) -> int:
        if _USER32 is None or hwnd == 0:
            return 0
        try:
            command = getattr(win32con, "GW_HWNDNEXT", 2) if win32con is not None else 2
        except Exception:
            command = 2
        try:
            nxt = _USER32.GetWindow(wintypes.HWND(hwnd), command)
        except Exception:
            nxt = 0
        return int(nxt) if nxt else 0

    def _fallback_effective_foreground_window(self) -> int:
        if _USER32 is None:
            return 0
        overlay_hwnd = int(self.winId()) if self.winId() else 0
        foreground = _user32_get_foreground_window()
        visited: Set[int] = set()
        hwnd = foreground
        while hwnd and hwnd not in visited:
            visited.add(hwnd)
            top = _user32_top_level_hwnd(hwnd)
            if top == overlay_hwnd or self._fallback_belongs_to_current_process(top):
                hwnd = self._fallback_next_window_in_z_order(top)
                continue
            if self._fallback_is_target_window_valid(top):
                return top
            hwnd = self._fallback_next_window_in_z_order(top)
        return 0

    def _fallback_is_window_maximized(self, hwnd: int) -> bool:
        if _USER32 is None or hwnd == 0:
            return False
        try:
            return bool(_USER32.IsZoomed(wintypes.HWND(hwnd)))
        except Exception:
            return False

    def _fallback_candidate_role(self, hwnd: int) -> Optional[str]:
        if _USER32 is None or hwnd == 0:
            return None
        if not self._fallback_is_target_window_valid(hwnd):
            return None
        class_name = _user32_window_class_name(hwnd)
        if not class_name:
            return None
        role: Optional[str] = None
        if class_name in self._KNOWN_PRESENTATION_CLASSES or any(
            class_name.startswith(prefix) for prefix in self._KNOWN_PRESENTATION_PREFIXES
        ):
            role = "slideshow"
        elif class_name in self._DOCUMENT_WINDOW_CLASSES or any(
            class_name.startswith(prefix) for prefix in self._DOCUMENT_WINDOW_PREFIXES
        ):
            role = "document"
        if role is None:
            return None
        rect = _user32_window_rect(hwnd)
        if not rect:
            return None
        left, top, right, bottom = rect
        width = max(0, right - left)
        height = max(0, bottom - top)
        if width < 400 or height < 300:
            return None
        overlay = self._overlay_rect_tuple()
        if overlay is None:
            return None
        o_width = overlay[2] - overlay[0]
        o_height = overlay[3] - overlay[1]
        if o_width <= 0 or o_height <= 0:
            return None
        width_diff = abs(width - o_width)
        height_diff = abs(height - o_height)
        if width_diff > 64 or height_diff > 64:
            return None
        if role == "document":
            if not self._fallback_is_window_maximized(hwnd):
                return None
            foreground = self._fallback_effective_foreground_window()
            if foreground:
                target_top = _user32_top_level_hwnd(hwnd)
                foreground_top = _user32_top_level_hwnd(foreground)
                if target_top and foreground_top and target_top != foreground_top:
                    return None
        return role

    def _fallback_is_target_window_valid(self, hwnd: int) -> bool:
        if _USER32 is None or hwnd == 0:
            return False
        overlay_hwnd = int(self.winId()) if self.winId() else 0
        if hwnd == overlay_hwnd:
            return False
        if self._fallback_belongs_to_current_process(hwnd):
            return False
        if not _user32_is_window(hwnd):
            return False
        if not _user32_is_window_visible(hwnd) or _user32_is_window_iconic(hwnd):
            return False
        rect = _user32_window_rect(hwnd)
        if not rect:
            return False
        return self._rect_intersects_overlay(rect)

    def _fallback_is_candidate_window(self, hwnd: int) -> bool:
        return self._fallback_candidate_role(hwnd) is not None

    def _fallback_detect_presentation_window_user32(self) -> Optional[int]:
        if _USER32 is None:
            return None
        overlay_hwnd = int(self.winId()) if self.winId() else 0
        foreground = self._fallback_effective_foreground_window()
        if foreground and foreground != overlay_hwnd:
            role = self._fallback_candidate_role(foreground)
            if role is not None:
                return foreground
        if _WNDENUMPROC is None:
            return None
        candidates: List[Tuple[int, str]] = []

        def _enum_callback(hwnd: int, _l_param: int) -> int:
            if hwnd == overlay_hwnd:
                return True
            if not _user32_is_window_visible(hwnd) or _user32_is_window_iconic(hwnd):
                return True
            rect = _user32_window_rect(hwnd)
            if not rect or not self._rect_intersects_overlay(rect):
                return True
            role = self._fallback_candidate_role(int(hwnd))
            if role is not None:
                candidates.append((int(hwnd), role))
            return True

        enum_proc = _WNDENUMPROC(_enum_callback)
        try:
            _USER32.EnumWindows(enum_proc, 0)
        except Exception:
            return None
        for hwnd, role in candidates:
            if role == "slideshow":
                return hwnd
        for hwnd, role in candidates:
            if role == "document":
                return hwnd
        return None

    def _presentation_window_class(self, hwnd: int) -> str:
        if hwnd == 0:
            return ""
        if win32gui is not None:
            try:
                return win32gui.GetClassName(hwnd).strip().lower()
            except Exception:
                return ""
        return _user32_window_class_name(hwnd)

    def _is_preferred_presentation_class(self, class_name: str) -> bool:
        if not class_name:
            return False
        if class_name in self._SLIDESHOW_PRIORITY_CLASSES:
            return True
        if class_name in self._SLIDESHOW_SECONDARY_CLASSES:
            return True
        return False

    def _should_refresh_cached_presentation_target(self, hwnd: int) -> bool:
        class_name = self._presentation_window_class(hwnd)
        return not self._is_preferred_presentation_class(class_name)

    def _focus_presentation_window_fallback(self) -> bool:
        if _USER32 is None:
            return False
        hwnd = self._resolve_presentation_target()
        if not hwnd:
            hwnd = self._fallback_detect_presentation_window_user32()
            if hwnd and self._fallback_is_target_window_valid(hwnd):
                self._last_target_hwnd = hwnd
        if not hwnd or not self._fallback_is_target_window_valid(hwnd):
            return False
        top_level = _user32_top_level_hwnd(hwnd)
        focused = _user32_focus_window(top_level)
        if not focused:
            focused = _user32_focus_window(hwnd)
        elif hwnd != top_level:
            _user32_focus_window(hwnd)
        return focused

    def _is_target_window_valid(self, hwnd: int) -> bool:
        if win32gui is None:
            return self._fallback_is_target_window_valid(hwnd)
        try:
            if hwnd == 0 or hwnd == int(self.winId()):
                return False
            if not win32gui.IsWindow(hwnd) or not win32gui.IsWindowVisible(hwnd):
                return False
            if win32gui.IsIconic(hwnd):
                return False
            rect = win32gui.GetWindowRect(hwnd)
        except Exception:
            return False
        if not rect:
            return False
        return self._rect_intersects_overlay(rect)

    def _detect_presentation_window(self) -> Optional[int]:
        if win32gui is None:
            return self._fallback_detect_presentation_window_user32()
        overlay_hwnd = int(self.winId()) if self.winId() else 0
        try:
            foreground = win32gui.GetForegroundWindow()
        except Exception:
            foreground = 0
        if foreground and foreground != overlay_hwnd:
            if self._is_candidate_presentation_window(foreground):
                return foreground
        candidates: List[int] = []

        def _enum_callback(hwnd: int, result: List[int]) -> bool:
            if hwnd == overlay_hwnd:
                return True
            try:
                if not win32gui.IsWindowVisible(hwnd) or win32gui.IsIconic(hwnd):
                    return True
                rect = win32gui.GetWindowRect(hwnd)
            except Exception:
                return True
            if not rect or not self._rect_intersects_overlay(rect):
                return True
            result.append(hwnd)
            return True

        try:
            win32gui.EnumWindows(_enum_callback, candidates)
        except Exception:
            return None
        for hwnd in candidates:
            if self._is_candidate_presentation_window(hwnd):
                return hwnd
        return None

    def _resolve_presentation_target(self) -> Optional[int]:
        if win32gui is None:
            hwnd = self._last_target_hwnd
            if hwnd and self._fallback_is_target_window_valid(hwnd):
                if self._should_refresh_cached_presentation_target(hwnd):
                    refreshed = self._fallback_detect_presentation_window_user32()
                    if (
                        refreshed
                        and refreshed != hwnd
                        and self._fallback_is_target_window_valid(refreshed)
                    ):
                        self._last_target_hwnd = refreshed
                        return refreshed
                return hwnd
            hwnd = self._fallback_detect_presentation_window_user32()
            if hwnd and self._fallback_is_target_window_valid(hwnd):
                self._last_target_hwnd = hwnd
                return hwnd
            self._last_target_hwnd = None
            return None
        hwnd = self._last_target_hwnd
        if hwnd and self._is_target_window_valid(hwnd):
            if self._should_refresh_cached_presentation_target(hwnd):
                refreshed = self._detect_presentation_window()
                if refreshed and refreshed != hwnd and self._is_target_window_valid(refreshed):
                    self._last_target_hwnd = refreshed
                    return refreshed
            return hwnd
        hwnd = self._detect_presentation_window()
        if hwnd and self._is_target_window_valid(hwnd):
            self._last_target_hwnd = hwnd
            return hwnd
        self._last_target_hwnd = None
        return None

    def _is_candidate_presentation_window(self, hwnd: int) -> bool:
        if win32gui is None:
            return self._fallback_is_candidate_window(hwnd)
        try:
            class_name = win32gui.GetClassName(hwnd).lower()
        except Exception:
            class_name = ""
        if class_name in self._KNOWN_PRESENTATION_CLASSES:
            return True
        try:
            rect = win32gui.GetWindowRect(hwnd)
        except Exception:
            return False
        if not rect:
            return False
        left, top, right, bottom = rect
        width = max(0, right - left)
        height = max(0, bottom - top)
        overlay = self._overlay_rect_tuple()
        if overlay is None:
            return False
        o_width = overlay[2] - overlay[0]
        o_height = overlay[3] - overlay[1]
        if o_width <= 0 or o_height <= 0:
            return False
        width_diff = abs(width - o_width)
        height_diff = abs(height - o_height)
        return width >= 400 and height >= 300 and width_diff <= 64 and height_diff <= 64

    def _push_history(self) -> None:
        if not isinstance(self.canvas, QPixmap):
            return
        self.history.append(self.canvas.copy())
        if len(self.history) > self._history_limit:
            self.history.pop(0)
        self._update_undo_button()

    def _update_undo_button(self) -> None:
        if getattr(self, "toolbar", None):
            self.toolbar.update_undo_state(bool(self.history))

    def clear_all(self) -> None:
        """清除整块画布，同时根据需要恢复画笔模式。"""
        restore_needed = self.mode not in {"brush", "shape"}
        self._push_history()
        self._release_canvas_painters()
        self.canvas.fill(Qt.GlobalColor.transparent)
        self.temp_canvas.fill(Qt.GlobalColor.transparent)
        self._last_preview_bounds = None
        self.update()
        self._eraser_last_point = None
        if restore_needed:
            self._restore_last_tool()
        else:
            if self.mode in {"brush", "shape"}:
                self._update_last_tool_snapshot()
            self.raise_toolbar()
        self._update_undo_button()

    def use_brush_color(self, color_hex: str) -> None:
        """根据传入的十六进制颜色值启用画笔模式。"""
        color = QColor(color_hex)
        if not color.isValid():
            return
        self.pen_color = color
        base_width = max(1.0, float(self.pen_size))
        self._update_brush_pen_appearance(base_width, 180)
        self.set_mode("brush")

    def undo_last_action(self) -> None:
        if not self.history:
            return
        last = self.history.pop()
        if isinstance(last, QPixmap):
            self._release_canvas_painters()
            self.canvas = last
        else:
            self._update_undo_button()
            return
        self.temp_canvas.fill(Qt.GlobalColor.transparent)
        self.drawing = False
        self._last_preview_bounds = None
        self.update()
        self.raise_toolbar()
        self._update_undo_button()

    def save_settings(self) -> None:
        settings = self.settings_manager.load_settings()
        paint = settings.get("Paint", {})
        paint["brush_size"] = str(self.pen_size)
        paint["brush_color"] = self.pen_color.name()
        settings["Paint"] = paint
        self.settings_manager.save_settings(settings)

    def save_window_position(self) -> None:
        settings = self.settings_manager.load_settings()
        paint = settings.get("Paint", {})
        pos = self.toolbar.pos()
        paint["x"] = str(pos.x()); paint["y"] = str(pos.y())
        settings["Paint"] = paint
        self.settings_manager.save_settings(settings)

    # ---- 画图事件 ----
    def wheelEvent(self, e: QWheelEvent) -> None:
        handled = False
        had_grab = False
        if self.mode in {"brush", "shape", "eraser", "cursor"} and not self.whiteboard_active:
            self._after_navigation_wheel(e)
            handled, had_grab = self._dispatch_navigation_session(
                wheel_event=e,
                prefer_scroll=True,
            )
            if handled and self.mode != ToolMode.CURSOR.value and not had_grab:
                QTimer.singleShot(100, self._ensure_keyboard_capture)
        if not handled and self._forwarder is not None:
            try:
                handled = self._forwarder.forward_wheel(e)
            except Exception:
                handled = False
        if handled:
            self._schedule_toolbar_navigation_cleanup()
            e.accept()
            return
        super().wheelEvent(e)

    def _reset_brush_tracking(self) -> None:
        self._stroke_points.clear()
        self._stroke_timestamps.clear()
        self._stroke_last_midpoint = None
        self._stroke_filter_point = None
        self._stroke_speed = 0.0

    def _start_paint_session(self, event) -> None:
        self._push_history()
        self.drawing = True
        origin = QPointF(event.position())
        self.last_point = QPointF(origin)
        self.prev_point = QPointF(origin)
        now = time.time()
        self.last_time = now
        self._reset_brush_tracking()
        self._stroke_points.append(QPointF(origin))
        self._stroke_timestamps.append(now)
        self._stroke_last_midpoint = QPointF(origin)
        self._stroke_filter_point = QPointF(origin)
        self.last_width = max(1.0, float(self.pen_size) * 0.4)
        self.shape_start_point = event.pos() if self.mode == "shape" else None
        if self.mode == "shape":
            self._last_preview_bounds = None
        self._eraser_last_point = event.pos() if self.mode == "eraser" else None
        if self.mode == "brush":
            self._ensure_brush_painter()
            self._update_brush_pen_appearance(max(1.0, float(self.pen_size)), 180)
        elif self.mode == "eraser":
            self._ensure_eraser_painter()

    def _finalize_paint_session(self, release_pos: QPoint) -> Optional[Union[QRect, QRectF]]:
        dirty_region: Optional[Union[QRect, QRectF]] = None
        if self.mode == "shape" and self.current_shape:
            dirty_region = self._draw_shape_final(release_pos)
        self.drawing = False
        self.shape_start_point = None
        if self.mode == "eraser":
            self._eraser_last_point = None
            self._release_eraser_painter()
        elif self.mode == "brush":
            self._reset_brush_tracking()
            self._release_brush_painter()
        return dirty_region

    def mousePressEvent(self, e) -> None:
        context = self._build_input_context(e)
        self._maybe_restore_tool_for_press(context)
        if e.button() == Qt.MouseButton.LeftButton and context.tool != ToolMode.CURSOR:
            self._ensure_keyboard_capture()
            self._start_paint_session(e)
            self.raise_toolbar()
            e.accept()
        super().mousePressEvent(e)

    def mouseMoveEvent(self, e) -> None:
        context = self._build_input_context(e)
        self._handle_move_restore(context)
        if self.drawing and context.tool != ToolMode.CURSOR:
            p = e.pos(); pf = e.position()
            dirty_region = None
            if self.mode == "brush":
                dirty_region = self._draw_brush_line(pf)
            elif self.mode == "eraser":
                dirty_region = self._erase_at(p)
            elif self.mode == "shape" and self.current_shape:
                dirty_region = self._draw_shape_preview(p)
            self._apply_dirty_region(dirty_region)
            self.raise_toolbar()
        super().mouseMoveEvent(e)

    def mouseReleaseEvent(self, e) -> None:
        if e.button() == Qt.MouseButton.LeftButton and self.drawing:
            dirty_region = self._finalize_paint_session(e.pos())
            if dirty_region is not None:
                self._apply_dirty_region(dirty_region)
            self.raise_toolbar()
            e.accept()
        super().mouseReleaseEvent(e)

    def keyPressEvent(self, e: QKeyEvent) -> None:
        key_code = int(e.key())
        nav_key = key_code in self._NAVIGATION_KEY_CODES
        handled = False
        if nav_key:
            self._nav_direct_keys.discard(key_code)
            vk_code = self._NAVIGATION_QT_TO_VK.get(key_code, 0)
            if vk_code:
                handled = self._send_slide_virtual_key(vk_code)
                if handled:
                    self._nav_direct_keys.add(key_code)
        if not handled and self._forwarder and self._forwarder.forward_key(e, is_press=True):
            if nav_key:
                self._after_navigation_key(e)
            handled = True
        if handled:
            e.accept()
            return
        if e.key() == Qt.Key.Key_Escape:
            self.set_mode("cursor"); return
        super().keyPressEvent(e)

    def keyReleaseEvent(self, e: QKeyEvent) -> None:
        key_code = int(e.key())
        nav_key = key_code in self._NAVIGATION_KEY_CODES
        if nav_key:
            if key_code in self._nav_direct_keys:
                self._nav_direct_keys.discard(key_code)
                if self.mode == "cursor" and self._interaction_mode == "navigation":
                    QTimer.singleShot(0, self._clear_navigation_passthrough)
                e.accept()
                return
            forwarded = bool(self._forwarder and self._forwarder.forward_key(e, is_press=False))
            if forwarded:
                if self.mode == "cursor" and self._interaction_mode == "navigation":
                    QTimer.singleShot(0, self._clear_navigation_passthrough)
                e.accept()
                return
            if self.mode == "cursor" and self._interaction_mode == "navigation":
                QTimer.singleShot(0, self._clear_navigation_passthrough)
            e.accept()
            return
        if self._forwarder and self._forwarder.forward_key(e, is_press=False):
            e.accept()
            return
        super().keyReleaseEvent(e)

    def _draw_brush_line(self, cur: QPointF) -> Optional[QRectF]:
        now = time.time()
        cur_point = QPointF(cur)
        if not self._stroke_points:
            self._stroke_points.clear()
            self._stroke_timestamps.clear()
            self._stroke_points.append(QPointF(cur_point))
            self._stroke_timestamps.append(now)
            self.prev_point = QPointF(cur_point)
            self.last_point = QPointF(cur_point)
            self._stroke_last_midpoint = QPointF(cur_point)
            self._stroke_filter_point = QPointF(cur_point)
            self.last_time = now
            self._ensure_brush_painter()
            return None

        painter = self._ensure_brush_painter()
        last_point = QPointF(self._stroke_points[-1])
        filter_point = QPointF(self._stroke_filter_point) if self._stroke_filter_point else QPointF(last_point)
        smoothing = 0.68
        smoothed_x = filter_point.x() + (cur_point.x() - filter_point.x()) * smoothing
        smoothed_y = filter_point.y() + (cur_point.y() - filter_point.y()) * smoothing
        cur_point = QPointF(smoothed_x, smoothed_y)
        self._stroke_filter_point = QPointF(cur_point)

        self._stroke_points.append(cur_point)
        self._stroke_timestamps.append(now)
        if len(self._stroke_timestamps) < 2:
            return None

        elapsed = max(1e-4, now - self._stroke_timestamps[-2])
        distance = math.hypot(cur_point.x() - last_point.x(), cur_point.y() - last_point.y())
        if distance < 0.08 and elapsed < 0.012:
            return None
        speed = distance / elapsed
        self._stroke_speed = self._stroke_speed * 0.6 + speed * 0.4

        curvature = 0.0
        if len(self._stroke_points) >= 3:
            p0 = self._stroke_points[-3]
            p1 = self._stroke_points[-2]
            p2 = self._stroke_points[-1]
            v1x, v1y = p1.x() - p0.x(), p1.y() - p0.y()
            v2x, v2y = p2.x() - p1.x(), p2.y() - p1.y()
            denom = math.hypot(v1x, v1y) * math.hypot(v2x, v2y)
            if denom > 1e-5:
                curvature = abs(v1x * v2y - v1y * v2x) / denom

        base_size = float(max(1, self.pen_size))
        travel = distance / max(1.0, base_size)
        pressure = min(1.0, (now - self.last_time) * 3.5 + travel * 0.25)
        self.last_time = now

        speed_scale = 1.0 / (1.0 + self._stroke_speed / (base_size * 16.0 + 24.0))
        curve_scale = min(1.0, curvature * base_size * 0.7)
        target_w = base_size * (0.45 + 0.4 * speed_scale + 0.25 * curve_scale)
        target_w *= 1.0 + 0.28 * pressure
        cur_w = self.last_width * 0.52 + target_w * 0.48

        last_mid = QPointF(self._stroke_last_midpoint) if self._stroke_last_midpoint else QPointF(last_point)
        current_mid = (last_point + cur_point) / 2.0

        path = QPainterPath(last_mid)
        path.quadTo(last_point, current_mid)

        fade_alpha = int(max(40, min(230, 220 * speed_scale + 60 * curve_scale)))
        self._update_brush_pen_appearance(cur_w, fade_alpha)
        painter.setPen(self._brush_shadow_pen)
        painter.drawPath(path)
        painter.setPen(self._brush_pen)
        painter.drawPath(path)

        self.prev_point = QPointF(last_point)
        self.last_point = QPointF(cur_point)
        self._stroke_last_midpoint = QPointF(current_mid)
        self.last_width = cur_w

        dirty = path.boundingRect()
        margin = cur_w * 0.9 + 4.0
        return dirty.adjusted(-margin, -margin, margin, margin)

    def _erase_at(self, pos) -> Optional[QRectF]:
        current = QPointF(pos) if isinstance(pos, QPointF) else QPointF(QPoint(pos))
        if isinstance(self._eraser_last_point, QPoint):
            start_point = QPointF(self._eraser_last_point)
            distance = math.hypot(current.x() - start_point.x(), current.y() - start_point.y())
        else:
            start_point = QPointF(current)
            distance = 0.0

        radius = max(8.0, float(self.pen_size) * 1.6)
        target_width = max(12.0, radius * 2.0)
        if abs(target_width - self._eraser_stroker_width) > 0.5:
            self._eraser_stroker.setWidth(target_width)
            self._eraser_stroker_width = target_width

        path = QPainterPath(start_point)
        if distance >= 0.35:
            path.lineTo(current)

        erase_path = QPainterPath()
        if distance >= 0.35:
            erase_path = self._eraser_stroker.createStroke(path)
        erase_path.addEllipse(current, radius, radius)
        if distance >= 0.35:
            erase_path.addEllipse(start_point, radius, radius)

        painter = self._ensure_eraser_painter()
        painter.fillPath(erase_path, QColor(0, 0, 0, 0))

        self._eraser_last_point = current.toPoint()

        dirty = erase_path.boundingRect()
        if dirty.isNull():
            return None
        margin = max(radius * 0.5, 6.0)
        return dirty.adjusted(-margin, -margin, margin, margin)

    def _draw_shape_preview(self, end_point) -> Optional[QRect]:
        if not self.shape_start_point:
            return None
        self.temp_canvas.fill(Qt.GlobalColor.transparent)
        p = QPainter(self.temp_canvas); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        pen = QPen(self.pen_color, self.pen_size)
        if self.current_shape and "dashed" in self.current_shape: pen.setStyle(Qt.PenStyle.DashLine)
        p.setPen(pen); self._draw_shape(p, self.shape_start_point, end_point); p.end()
        self.raise_toolbar()
        bounds = self._shape_dirty_bounds(self.shape_start_point, end_point, self.pen_size)
        if bounds is not None and self._last_preview_bounds is not None:
            bounds = bounds.united(self._last_preview_bounds)
        self._last_preview_bounds = bounds
        return bounds

    def _draw_shape_final(self, end_point) -> Optional[QRect]:
        if not self.shape_start_point:
            return None
        bounds = self._shape_dirty_bounds(self.shape_start_point, end_point, self.pen_size)
        p = QPainter(self.canvas); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        pen = QPen(self.pen_color, self.pen_size)
        if self.current_shape and "dashed" in self.current_shape: pen.setStyle(Qt.PenStyle.DashLine)
        p.setPen(pen); self._draw_shape(p, self.shape_start_point, end_point); p.end()
        self.temp_canvas.fill(Qt.GlobalColor.transparent)
        self.raise_toolbar()
        last_bounds = self._last_preview_bounds
        self._last_preview_bounds = None
        if bounds is not None and last_bounds is not None:
            bounds = bounds.united(last_bounds)
        return bounds

    def _draw_shape(self, painter: QPainter, start_point, end_point) -> None:
        rect = QRect(start_point, end_point)
        shape = (self.current_shape or "line").replace("dashed_", "")
        if shape == "rect": painter.drawRect(rect.normalized())
        elif shape == "circle": painter.drawEllipse(rect.normalized())
        else: painter.drawLine(start_point, end_point)

    def paintEvent(self, e) -> None:
        p = QPainter(self)
        canvas_visible = self._mode_canvas_visible and not self._canvas_hidden
        if canvas_visible and self.whiteboard_active:
            p.fillRect(self.rect(), self.whiteboard_color)
        elif canvas_visible:
            p.fillRect(self.rect(), QColor(0, 0, 0, 1))
        else:
            p.fillRect(self.rect(), QColor(0, 0, 0, 0))
        if canvas_visible:
            p.drawPixmap(0, 0, self.canvas)
            if self.drawing and self.mode == "shape":
                p.drawPixmap(0, 0, self.temp_canvas)
        p.end()

    def showEvent(self, e) -> None:
        super().showEvent(e)
        self.raise_toolbar()

    def closeEvent(self, e) -> None:
        self._release_canvas_painters()
        self.save_settings()
        self.save_window_position()
        super().closeEvent(e)


# ---------- 语音 ----------
class TTSManager(QObject):
    """简单封装语音播报，优先使用 pyttsx3，必要时回退到 PowerShell。"""

    def __init__(self, preferred_voice_id: str = "", parent: Optional[QObject] = None) -> None:
        super().__init__(parent)
        self.engine = None
        self.voice_ids: List[str] = []
        self.default_voice_id = ""
        self.current_voice_id = ""
        self.failure_reason = ""
        self.failure_suggestions: List[str] = []
        self.supports_voice_selection = False
        self._mode: str = "none"
        self._powershell_path = ""
        self._powershell_busy = False
        self._queue: Queue[str] = Queue()
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._pump)
        missing_reason = ""
        if pyttsx3 is not None:
            try:
                init_kwargs = {"driverName": "sapi5"} if sys.platform == "win32" else {}
                self.engine = pyttsx3.init(**init_kwargs)
                voices = self.engine.getProperty("voices") or []
                self.voice_ids = [v.id for v in voices if getattr(v, "id", None)]
                if not self.voice_ids:
                    self._record_failure("未检测到任何可用的发音人")
                    self.engine = None
                else:
                    self.default_voice_id = self.voice_ids[0]
                    self.current_voice_id = (
                        preferred_voice_id if preferred_voice_id in self.voice_ids else self.default_voice_id
                    )
                    if self.current_voice_id:
                        try:
                            self.engine.setProperty("voice", self.current_voice_id)
                        except Exception as exc:
                            self._record_failure("无法设置默认发音人", exc)
                            self.engine = None
                    if self.engine is not None:
                        self.supports_voice_selection = True
                        self._mode = "pyttsx3"
                        self.engine.startLoop(False)
                        self._timer.start(100)
                        return
            except Exception as exc:
                self._record_failure("初始化语音引擎失败", exc)
                self.engine = None
        else:
            missing_reason = "未检测到 pyttsx3 模块"
        self._init_powershell_fallback()
        if self.available:
            return
        if missing_reason:
            if self.failure_reason:
                if missing_reason not in self.failure_reason:
                    self.failure_reason = f"{self.failure_reason}；{missing_reason}"
            else:
                self.failure_reason = missing_reason
        if not self.failure_reason:
            self.failure_reason = "未检测到可用的语音播报方式"
        env_reason, env_suggestions = detect_speech_environment_issues(force_refresh=True)
        if env_reason:
            if env_reason not in self.failure_reason:
                self.failure_reason = f"{self.failure_reason}；{env_reason}" if self.failure_reason else env_reason
        if env_suggestions:
            combined = list(self.failure_suggestions)
            combined.extend(env_suggestions)
            self.failure_suggestions = dedupe_strings(combined)

    @property
    def available(self) -> bool:
        return self._mode in {"pyttsx3", "powershell"}

    def diagnostics(self) -> tuple[str, List[str]]:
        reason = self.failure_reason
        suggestions = list(self.failure_suggestions)
        env_reason, env_suggestions = detect_speech_environment_issues()
        if env_reason:
            if reason:
                if env_reason not in reason:
                    reason = f"{reason}；{env_reason}"
            else:
                reason = env_reason
        suggestions.extend(env_suggestions)
        return reason, dedupe_strings(suggestions)

    def _init_powershell_fallback(self) -> None:
        if sys.platform != "win32":
            return
        path = _find_powershell_executable()
        if not path:
            if not self.failure_reason:
                self._record_failure("未检测到 PowerShell，可用的语音播报方式受限")
            return
        self._powershell_path = os.path.abspath(path)
        ps_ok, ps_reason = _probe_powershell_speech_runtime(self._powershell_path)
        if not ps_ok:
            message = ps_reason or "PowerShell 语音环境检测失败"
            self._record_failure(message)
            return
        self.engine = object()
        self.voice_ids = []
        self.default_voice_id = ""
        self.current_voice_id = ""
        self.supports_voice_selection = False
        self.failure_reason = ""
        self.failure_suggestions = []
        self._mode = "powershell"
        self._timer.start(120)

    def _record_failure(self, fallback: str, exc: Optional[Exception] = None) -> None:
        message = ""
        if exc is not None:
            message = str(exc).strip()
        if message and message not in fallback:
            reason = f"{fallback}：{message}"
        else:
            reason = fallback
        self.failure_reason = reason
        suggestions: List[str] = []
        lower = message.lower()
        if "comtypes" in lower:
            suggestions.append("请安装 comtypes（pip install comtypes）后重新启动程序。")
        if "pywin32" in lower or "win32" in lower:
            suggestions.append("请安装 pywin32（pip install pywin32）后重新启动程序。")
        platform_hint = []
        if sys.platform == "win32":
            platform_hint.append("请确认 Windows 已启用 SAPI5 中文语音包。")
        elif sys.platform == "darwin":
            platform_hint.append("请在系统“辅助功能 -> 语音”中启用所需的语音包。")
        else:
            platform_hint.append("请确保系统已安装可用的语音引擎（如 espeak）并重新启动程序。")
        platform_hint.append("可尝试重新安装 pyttsx3 或检查语音服务状态后重启软件。")
        for hint in platform_hint:
            if hint not in suggestions:
                suggestions.append(hint)
        self.failure_suggestions = suggestions

    def set_voice(self, voice_id: str) -> None:
        if not self.supports_voice_selection:
            return
        if voice_id in self.voice_ids:
            self.current_voice_id = voice_id
            if self.engine:
                try:
                    self.engine.setProperty("voice", voice_id)
                except Exception:
                    pass

    def speak(self, text: str) -> None:
        if not self.available:
            return
        while not self._queue.empty():
            try:
                self._queue.get_nowait()
            except Empty:
                break
        self._queue.put(text)

    def _pump(self) -> None:
        if self._mode == "pyttsx3":
            if not self.engine:
                return
            try:
                text = self._queue.get_nowait()
                self.engine.stop()
                if self.current_voice_id:
                    self.engine.setProperty("voice", self.current_voice_id)
                self.engine.say(text)
            except Empty:
                pass
            try:
                self.engine.iterate()
            except Exception as exc:
                self._record_failure("语音引擎运行异常", exc)
                self.shutdown()
        elif self._mode == "powershell":
            if self._powershell_busy:
                return
            try:
                text = self._queue.get_nowait()
            except Empty:
                return
            self._powershell_busy = True
            worker = threading.Thread(target=self._run_powershell_speech, args=(text,), daemon=True)
            worker.start()

    def _run_powershell_speech(self, text: str) -> None:
        try:
            if not text or not self._powershell_path:
                return
            payload = base64.b64encode(text.encode("utf-8")).decode("ascii")
            script = (
                "$msg = [System.Text.Encoding]::UTF8.GetString([System.Convert]::FromBase64String('" + payload + "'));"
                "Add-Type -AssemblyName System.Speech;"
                "$sp = New-Object System.Speech.Synthesis.SpeechSynthesizer;"
                "$sp.Speak($msg);"
            )
            startupinfo = None
            if os.name == "nt":
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            subprocess.run(
                [self._powershell_path, "-NoLogo", "-NonInteractive", "-NoProfile", "-Command", script],
                check=True,
                timeout=30,
                startupinfo=startupinfo,
            )
        except Exception as exc:
            self._record_failure("PowerShell 语音播报失败", exc)
            QTimer.singleShot(0, self.shutdown)
        finally:
            self._powershell_busy = False

    def shutdown(self) -> None:
        if self._mode == "pyttsx3" and self.engine:
            try:
                self.engine.endLoop()
            except Exception:
                pass
            try:
                self.engine.stop()
            except Exception:
                pass
        self.engine = None
        self._mode = "none"
        self._powershell_busy = False
        self._timer.stop()


# ---------- 点名/计时 ----------
class CountdownSettingsDialog(QDialog):
    """设置倒计时分钟和秒数的小窗口。"""

    def __init__(self, parent: Optional[QWidget], minutes: int, seconds: int) -> None:
        super().__init__(parent)
        self.setWindowTitle("设置倒计时")
        self.setWindowFlag(Qt.WindowType.WindowStaysOnTopHint, True)
        self.result: Optional[tuple[int, int]] = None

        layout = QVBoxLayout(self); layout.setContentsMargins(10, 10, 10, 10); layout.setSpacing(6)

        ml = QHBoxLayout(); ml.addWidget(QLabel("分钟 (0-25):"))
        self.minutes_spin = QSpinBox(); self.minutes_spin.setRange(0, 300); self.minutes_spin.setValue(max(0, min(300, minutes)))
        minute_slider = QSlider(Qt.Orientation.Horizontal); minute_slider.setRange(0, 25)
        minute_slider.setValue(min(self.minutes_spin.value(), minute_slider.maximum()))
        minute_slider.valueChanged.connect(self.minutes_spin.setValue)

        def sync(v: int, slider=minute_slider):
            if v <= slider.maximum():
                prev = slider.blockSignals(True); slider.setValue(v); slider.blockSignals(prev)
        self.minutes_spin.valueChanged.connect(sync)
        ml.addWidget(self.minutes_spin); layout.addLayout(ml); layout.addWidget(minute_slider)

        sl = QHBoxLayout(); sl.addWidget(QLabel("秒 (0-59):"))
        self.seconds_spin = QSpinBox(); self.seconds_spin.setRange(0, 59); self.seconds_spin.setValue(max(0, min(59, seconds)))
        second_slider = QSlider(Qt.Orientation.Horizontal); second_slider.setRange(0, 59)
        second_slider.setValue(self.seconds_spin.value())
        second_slider.valueChanged.connect(self.seconds_spin.setValue); self.seconds_spin.valueChanged.connect(second_slider.setValue)
        sl.addWidget(self.seconds_spin); layout.addLayout(sl); layout.addWidget(second_slider)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._accept); buttons.rejected.connect(self.reject)
        style_dialog_buttons(
            buttons,
            {
                QDialogButtonBox.StandardButton.Ok: ButtonStyles.PRIMARY,
                QDialogButtonBox.StandardButton.Cancel: ButtonStyles.TOOLBAR,
            },
            extra_padding=12,
            minimum_height=34,
        )
        layout.addWidget(buttons)
        self.setFixedSize(self.sizeHint())

    def _accept(self) -> None:
        self.result = (self.minutes_spin.value(), self.seconds_spin.value()); self.accept()

    def showEvent(self, event) -> None:  # type: ignore[override]
        super().showEvent(event)
        ensure_widget_within_screen(self)


class ClickableFrame(QFrame):
    clicked = pyqtSignal()
    def mousePressEvent(self, e) -> None:
        if e.button() == Qt.MouseButton.LeftButton: self.clicked.emit()
        super().mousePressEvent(e)


def preferred_calligraphy_font(default: str = "Microsoft YaHei UI") -> str:
    """返回系统中更具书法风格的字体，若不可用则回退到默认字体。"""

    try:
        families = set(QFontDatabase().families())
    except Exception:
        return default
    for candidate in ("楷体", "KaiTi", "Kaiti SC", "STKaiti", "DFKai-SB", "FZKai-Z03S"):
        if candidate in families:
            return candidate
    return default


class StudentListDialog(QDialog):
    def __init__(self, parent: Optional[QWidget], students: List[tuple[str, str, int]]) -> None:
        super().__init__(parent)
        self.setWindowTitle("学生名单")
        self.setModal(True)
        self._selected_index: Optional[int] = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)

        grid = QGridLayout()
        grid.setContentsMargins(0, 0, 0, 0)
        grid.setHorizontalSpacing(6)
        grid.setVerticalSpacing(6)
        grid.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignHCenter)

        button_font = QFont("Microsoft YaHei UI", 10, QFont.Weight.Medium)
        metrics = QFontMetrics(button_font)
        max_text = max((metrics.horizontalAdvance(f"{sid} {name}") for sid, name, _ in students), default=120)
        min_button_width = max(120, max_text + 24)
        button_height = recommended_control_height(button_font, extra=16, minimum=38)

        screen = QApplication.primaryScreen()
        available_width = screen.availableGeometry().width() if screen else 1280
        max_width_per_button = max(96, int((available_width * 0.9 - 40) / 10))
        button_width = min(min_button_width, max_width_per_button)
        button_size = QSize(button_width, button_height)

        total_rows = max(1, math.ceil(len(students) / 10))

        for column in range(10):
            grid.setColumnStretch(column, 0)
            grid.setColumnMinimumWidth(column, button_width)

        for row in range(total_rows):
            grid.setRowStretch(row, 0)
            grid.setRowMinimumHeight(row, button_height)

        for position, (sid, name, data_index) in enumerate(students):
            row = position // 10
            column = position % 10
            button = QPushButton(f"{sid} {name}")
            button.setFont(button_font)
            button.setFixedSize(button_size)
            button.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
            apply_button_style(button, ButtonStyles.GRID, height=button_height)
            button.clicked.connect(lambda _checked=False, value=data_index: self._select_student(value))
            grid.addWidget(button, row, column, Qt.AlignmentFlag.AlignCenter)

        layout.addLayout(grid)

        box = QDialogButtonBox(QDialogButtonBox.StandardButton.Close, parent=self)
        box.rejected.connect(self.reject)
        close_button = box.button(QDialogButtonBox.StandardButton.Close)
        if close_button is not None:
            close_button.setText("关闭")
            apply_button_style(
                close_button,
                ButtonStyles.PRIMARY,
                height=recommended_control_height(close_button.font(), extra=14, minimum=36),
            )
        layout.addWidget(box)

        if screen is not None:
            available = screen.availableGeometry()
            rows = total_rows
            h_spacing = grid.horizontalSpacing() if grid.horizontalSpacing() is not None else 6
            v_spacing = grid.verticalSpacing() if grid.verticalSpacing() is not None else 6
            preferred_width = min(int(available.width() * 0.9), button_width * 10 + h_spacing * 9 + 40)
            preferred_height = min(
                int(available.height() * 0.85),
                rows * button_height + max(0, rows - 1) * v_spacing + box.sizeHint().height() + 48,
            )
            self.resize(preferred_width, preferred_height)

    def _select_student(self, index: int) -> None:
        self._selected_index = index
        self.accept()

    @property
    def selected_index(self) -> Optional[int]:
        return self._selected_index


class ScoreboardDialog(QDialog):
    ORDER_RANK = "rank"
    ORDER_ID = "id"

    @dataclass
    class _CardMetrics:
        count: int
        columns: int
        rows: int
        card_width: int
        card_height: int
        padding_h: int
        padding_v: int
        inner_spacing: int
        font_size: int
        horizontal_spacing: int
        vertical_spacing: int

    def __init__(
        self,
        parent: Optional[QWidget],
        students: List[tuple[str, str, int]],
        order: str = "rank",
        order_changed: Optional[Callable[[str], None]] = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("成绩展示")
        self.setModal(True)
        self.setObjectName("ScoreboardDialog")
        self._pending_maximize = True

        self.students = list(students)
        self._order_changed_callback = order_changed
        self._order = order if order in {self.ORDER_RANK, self.ORDER_ID} else self.ORDER_RANK
        self._grid_row_count = 0
        self._grid_column_count = 0
        self._card_metrics: Optional[ScoreboardDialog._CardMetrics] = None
        self._card_metrics_key: Optional[tuple[int, int, int]] = None

        calligraphy_font = preferred_calligraphy_font()
        self._calligraphy_font = calligraphy_font

        layout = QVBoxLayout(self)
        layout.setContentsMargins(32, 32, 32, 32)
        layout.setSpacing(24)

        title = QLabel("成绩展示")
        title.setObjectName("ScoreboardHeader")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setFont(QFont(calligraphy_font, 44, QFont.Weight.Bold))
        layout.addWidget(title)

        order_layout = QHBoxLayout()
        order_layout.setContentsMargins(0, 0, 0, 0)
        order_layout.setSpacing(12)
        order_label = QLabel("排序方式：")
        order_label.setFont(QFont(calligraphy_font, 28, QFont.Weight.Bold))
        order_layout.addWidget(order_label, 0, Qt.AlignmentFlag.AlignLeft)

        button_font = QFont(calligraphy_font, 22, QFont.Weight.Bold)
        self.order_button_group = QButtonGroup(self)
        self.order_button_group.setExclusive(True)
        self.order_buttons: Dict[str, QPushButton] = {}
        for key, text in ((self.ORDER_RANK, "按排名"), (self.ORDER_ID, "按学号")):
            button = QPushButton(text)
            button.setCheckable(True)
            button.setFont(button_font)
            height = recommended_control_height(button_font, extra=16, minimum=44)
            apply_button_style(button, ButtonStyles.ORDER_TOGGLE, height=height)
            button.setMinimumWidth(140)
            button.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
            self.order_button_group.addButton(button)
            button.clicked.connect(lambda _checked=False, value=key: self._on_order_button_clicked(value))
            order_layout.addWidget(button, 0, Qt.AlignmentFlag.AlignLeft)
            self.order_buttons[key] = button
        order_layout.addStretch(1)
        layout.addLayout(order_layout)

        grid_container = QWidget()
        grid_container.setObjectName("ScoreboardGridContainer")
        grid_container.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.grid_layout = QGridLayout(grid_container)
        self.grid_layout.setContentsMargins(18, 18, 18, 18)
        self.grid_layout.setHorizontalSpacing(20)
        self.grid_layout.setVerticalSpacing(20)
        layout.addWidget(grid_container, 1)

        box = QDialogButtonBox(QDialogButtonBox.StandardButton.Close, parent=self)
        box.setFont(QFont(calligraphy_font, 22, QFont.Weight.Bold))
        close_button = box.button(QDialogButtonBox.StandardButton.Close)
        if close_button is not None:
            close_button.setText("关闭")
            close_button.setFont(QFont(calligraphy_font, 22, QFont.Weight.Bold))
            apply_button_style(
                close_button,
                ButtonStyles.PRIMARY,
                height=recommended_control_height(close_button.font(), extra=18, minimum=46),
            )
        box.rejected.connect(self.reject)
        layout.addWidget(box)

        self.setStyleSheet(
            "#ScoreboardDialog {"
            "    background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,"
            "        stop:0 #f7f9fc, stop:1 #e3edff);"
            "}"
            "#ScoreboardGridContainer {"
            "    background-color: rgba(255, 255, 255, 0.85);"
            "    border-radius: 24px;"
            "}"
            "QLabel#ScoreboardHeader {"
            "    color: #0b3d91;"
            "}"
            "QLabel[class=\"scoreboardName\"] {"
            "    color: #103d73;"
            "}"
            "QLabel[class=\"scoreboardScore\"] {"
            "    color: #103d73;"
            "}"
            "QWidget[class=\"scoreboardWrapper\"] {"
            "    background-color: rgba(255, 255, 255, 0.95);"
            "    border-radius: 18px;"
            "    border: 1px solid rgba(16, 61, 115, 0.12);"
            "}"
            "QPushButton[class=\"orderButton\"] {"
            "    background-color: rgba(255, 255, 255, 0.88);"
            "    border-radius: 22px;"
            "    border: 1px solid rgba(16, 61, 115, 0.24);"
            "    padding: 4px 18px;"
            "    color: #0b3d91;"
            "}"
            "QPushButton[class=\"orderButton\"]:hover {"
            "    border-color: #1a73e8;"
            "    background-color: rgba(26, 115, 232, 0.12);"
            "}"
            "QPushButton[class=\"orderButton\"]:checked {"
            "    background-color: #1a73e8;"
            "    border-color: #1a73e8;"
            "    color: #ffffff;"
            "}"
        )

        screen = QApplication.primaryScreen()
        self._available_geometry = screen.availableGeometry() if screen is not None else QRect(0, 0, 1920, 1080)

        self._update_order_buttons()
        self._populate_grid()

        if screen is not None:
            self.setGeometry(self._available_geometry)

    def _update_order_buttons(self) -> None:
        for key, button in self.order_buttons.items():
            block = button.blockSignals(True)
            button.setChecked(key == self._order)
            button.blockSignals(block)

    def _on_order_button_clicked(self, order: str) -> None:
        if order not in {self.ORDER_RANK, self.ORDER_ID}:
            self._update_order_buttons()
            return
        if order == self._order:
            self._update_order_buttons()
            return
        self._order = order
        if callable(self._order_changed_callback):
            try:
                self._order_changed_callback(order)
            except Exception:
                pass
        self._update_order_buttons()
        self._populate_grid()

    def _clear_grid(self) -> None:
        while self.grid_layout.count():
            item = self.grid_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
        for row in range(self._grid_row_count):
            self.grid_layout.setRowStretch(row, 0)
            self.grid_layout.setRowMinimumHeight(row, 0)
        for column in range(self._grid_column_count):
            self.grid_layout.setColumnStretch(column, 0)
            self.grid_layout.setColumnMinimumWidth(column, 0)
        self._grid_row_count = 0
        self._grid_column_count = 0

    def _collect_display_candidates(self) -> tuple[List[tuple[int, str, str]], List[str], List[str]]:
        sorted_students = self._sort_students()
        alternate_order = (
            self.ORDER_ID if self._order == self.ORDER_RANK else self.ORDER_RANK
        )
        alternate_students = self._sort_students(alternate_order)

        display_entries: List[tuple[int, str, str]] = []
        display_candidates: List[str] = []
        score_candidates: List[str] = []

        for idx, (sid, name, score) in enumerate(sorted_students):
            display_text = self._format_display_text(idx, sid, name)
            score_text = self._format_score_text(score)
            display_entries.append((idx, display_text, score_text))
            display_candidates.append(display_text)
            score_candidates.append(score_text)

        for idx, (sid, name, score) in enumerate(alternate_students):
            display_candidates.append(
                self._format_display_text_for_order(
                    alternate_order, idx, sid, name
                )
            )
            score_candidates.append(self._format_score_text(score))

        return display_entries, display_candidates, score_candidates

    def _compute_card_metrics(self) -> Optional[_CardMetrics]:
        count = len(self.students)
        if count == 0:
            return None

        available = self._available_geometry
        key = (count, available.width(), available.height())
        if self._card_metrics is not None and self._card_metrics_key == key:
            return self._card_metrics

        columns = 10
        rows = max(1, math.ceil(count / columns))

        usable_width = max(available.width() - 160, 640)
        usable_height = max(available.height() - 240, 520)

        margins = self.grid_layout.contentsMargins()
        horizontal_spacing = max(14, int(usable_width * 0.01))
        vertical_spacing = max(18, int(usable_height * 0.035 / rows))

        spacing_total_x = horizontal_spacing * max(0, columns - 1)
        spacing_total_y = vertical_spacing * max(0, rows - 1)

        available_width_for_cards = (
            usable_width - margins.left() - margins.right() - spacing_total_x
        )
        available_height_for_cards = (
            usable_height - margins.top() - margins.bottom() - spacing_total_y
        )

        per_card_width = max(1.0, available_width_for_cards / columns)
        per_card_height = max(1.0, available_height_for_cards / rows)

        card_width = int(math.floor(per_card_width))
        card_height = int(math.floor(per_card_height))

        if per_card_width >= 120:
            card_width = max(card_width, 120)
        if per_card_height >= 180:
            card_height = max(card_height, 180)

        padding_h = max(12, int(card_width * 0.08))
        padding_v = max(14, int(card_height * 0.1))
        inner_spacing = max(6, int(card_height * 0.045))

        display_entries, display_candidates, score_candidates = self._collect_display_candidates()

        if not display_candidates:
            return None

        calligraphy_font = self._calligraphy_font or QApplication.font().family()
        if not calligraphy_font:
            calligraphy_font = QFont().family()

        try:
            probe_font = QFont(calligraphy_font, 64, QFont.Weight.Bold)
            metrics = QFontMetrics(probe_font)
        except Exception:
            probe_font = QFont()
            metrics = QFontMetrics(probe_font)

        widest_display = max(
            display_candidates,
            key=lambda text: metrics.tightBoundingRect(text).width(),
        )
        widest_score = max(
            score_candidates,
            key=lambda text: metrics.tightBoundingRect(text).width(),
        )

        usable_name_width = max(60, card_width - 2 * padding_h)
        content_height = max(80, card_height - 2 * padding_v - inner_spacing)
        name_height = int(content_height * 0.58)
        score_height = max(32, content_height - name_height)

        font_upper_bound = int(min(card_width * 0.28, card_height * 0.36))
        font_upper_bound = max(20, font_upper_bound)
        fit_minimum = 14

        name_fit = self._fit_font_size(
            widest_display,
            calligraphy_font,
            QFont.Weight.Bold,
            usable_name_width,
            name_height,
            fit_minimum,
            font_upper_bound,
        )
        score_fit = self._fit_font_size(
            widest_score,
            calligraphy_font,
            QFont.Weight.Bold,
            usable_name_width,
            score_height,
            fit_minimum,
            font_upper_bound,
        )

        final_font_size = max(fit_minimum, min(name_fit, score_fit, font_upper_bound))

        self._card_metrics = ScoreboardDialog._CardMetrics(
            count=count,
            columns=columns,
            rows=rows,
            card_width=card_width,
            card_height=card_height,
            padding_h=padding_h,
            padding_v=padding_v,
            inner_spacing=inner_spacing,
            font_size=final_font_size,
            horizontal_spacing=horizontal_spacing,
            vertical_spacing=vertical_spacing,
        )
        self._card_metrics_key = key
        return self._card_metrics

    def _ensure_metrics(self) -> Optional[_CardMetrics]:
        metrics = self._compute_card_metrics()
        if metrics is not None:
            return metrics
        self._card_metrics = None
        self._card_metrics_key = None
        return None

    def _sort_students(self, order: Optional[str] = None) -> List[tuple[str, str, int]]:
        data = list(self.students)
        current_order = self._order if order is None else order
        if current_order == self.ORDER_ID:
            def _id_key(item: tuple[str, str, int]) -> tuple[int, str, str]:
                sid_text = str(item[0]).strip()
                try:
                    sid_value = int(sid_text)
                except (TypeError, ValueError):
                    sid_value = sys.maxsize
                return (sid_value, sid_text, item[1])

            data.sort(key=_id_key)
        else:
            def _rank_key(item: tuple[str, str, int]) -> tuple[int, int, str]:
                sid_text = str(item[0]).strip()
                try:
                    sid_value = int(sid_text)
                except (TypeError, ValueError):
                    sid_value = sys.maxsize
                return (-item[2], sid_value, item[1])

            data.sort(key=_rank_key)
        return data

    def _create_card(
        self,
        display_text: str,
        score_text: str,
        card_width: int,
        card_height: int,
        font_size: int,
        padding_h: int,
        padding_v: int,
        inner_spacing: int,
    ) -> QWidget:
        calligraphy_font = self._calligraphy_font
        wrapper = QWidget()
        wrapper.setProperty("class", "scoreboardWrapper")
        wrapper.setFixedSize(card_width, card_height)
        wrapper.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)

        layout = QVBoxLayout(wrapper)
        layout.setContentsMargins(padding_h, padding_v, padding_h, padding_v)
        layout.setSpacing(inner_spacing)

        name_label = QLabel(display_text or "未命名")
        name_label.setProperty("class", "scoreboardName")
        name_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        name_label.setWordWrap(False)
        name_label.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed
        )
        name_label.setFont(QFont(calligraphy_font, font_size, QFont.Weight.Bold))
        name_label.setStyleSheet("margin: 0px; padding: 0px;")
        layout.addWidget(name_label)

        score_label = QLabel(score_text)
        score_label.setProperty("class", "scoreboardScore")
        score_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        score_label.setWordWrap(False)
        score_label.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed
        )
        score_label.setFont(QFont(calligraphy_font, font_size, QFont.Weight.Bold))
        score_label.setStyleSheet(f"margin-top: {max(6, inner_spacing // 2)}px;")
        layout.addWidget(score_label)

        layout.addStretch(1)
        return wrapper

    def _format_display_text(self, index: int, sid: str, name: str) -> str:
        return self._format_display_text_for_order(self._order, index, sid, name)

    def _format_display_text_for_order(
        self, order: str, index: int, sid: str, name: str
    ) -> str:
        clean_name = (name or "").strip() or "未命名"
        if order == self.ORDER_ID:
            sid_display = str(sid).strip() or "—"
            return f"{sid_display}.{clean_name}"
        return f"{index + 1}.{clean_name}"

    @staticmethod
    def _format_score_text(score: int | float | str) -> str:
        text = "—"
        try:
            value = float(score)
        except (TypeError, ValueError):
            score_str = str(score).strip()
            if score_str and score_str.lower() != "none":
                text = score_str
        else:
            if math.isfinite(value):
                if abs(value - int(value)) < 1e-6:
                    text = str(int(round(value)))
                else:
                    text = f"{value:.2f}".rstrip("0").rstrip(".")
        return f"{text} 分"

    @staticmethod
    def _fit_font_size(
        text: str,
        family: str,
        weight: QFont.Weight,
        max_width: int,
        max_height: int,
        minimum: int,
        maximum: int,
    ) -> int:
        if not text:
            return max(6, min(minimum, maximum))
        if max_width <= 0 or max_height <= 0:
            return max(6, min(minimum, maximum))
        lower = max(6, min(minimum, maximum))
        upper = max(6, max(minimum, maximum))
        for size in range(upper, lower - 1, -1):
            font = QFont(family, size, weight)
            metrics = QFontMetrics(font)
            rect = metrics.tightBoundingRect(text)
            if rect.width() <= max_width and rect.height() <= max_height:
                return size
        return lower

    def _populate_grid(self) -> None:
        self._clear_grid()
        count = len(self.students)
        layout = self.grid_layout
        calligraphy_font = self._calligraphy_font

        screen = QApplication.primaryScreen()
        self._available_geometry = screen.availableGeometry() if screen is not None else QRect(0, 0, 1920, 1080)

        self._update_order_buttons()
        self._populate_grid()

        if screen is not None:
            self.setGeometry(self._available_geometry)

    def _update_order_buttons(self) -> None:
        for key, button in self.order_buttons.items():
            block = button.blockSignals(True)
            button.setChecked(key == self._order)
            button.blockSignals(block)

    def _on_order_button_clicked(self, order: str) -> None:
        if order not in {self.ORDER_RANK, self.ORDER_ID}:
            self._update_order_buttons()
            return
        if order == self._order:
            self._update_order_buttons()
            return
        self._order = order
        if callable(self._order_changed_callback):
            try:
                self._order_changed_callback(order)
            except Exception:
                pass
        self._update_order_buttons()
        self._populate_grid()

    def _clear_grid(self) -> None:
        while self.grid_layout.count():
            item = self.grid_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
        for row in range(self._grid_row_count):
            self.grid_layout.setRowStretch(row, 0)
            self.grid_layout.setRowMinimumHeight(row, 0)
        for column in range(self._grid_column_count):
            self.grid_layout.setColumnStretch(column, 0)
            self.grid_layout.setColumnMinimumWidth(column, 0)
        self._grid_row_count = 0
        self._grid_column_count = 0

    def _collect_display_candidates(self) -> tuple[List[tuple[int, str, str]], List[str], List[str]]:
        sorted_students = self._sort_students()
        alternate_order = (
            self.ORDER_ID if self._order == self.ORDER_RANK else self.ORDER_RANK
        )
        alternate_students = self._sort_students(alternate_order)

        display_entries: List[tuple[int, str, str]] = []
        display_candidates: List[str] = []
        score_candidates: List[str] = []

        for idx, (sid, name, score) in enumerate(sorted_students):
            display_text = self._format_display_text(idx, sid, name)
            score_text = self._format_score_text(score)
            display_entries.append((idx, display_text, score_text))
            display_candidates.append(display_text)
            score_candidates.append(score_text)

        for idx, (sid, name, score) in enumerate(alternate_students):
            display_candidates.append(
                self._format_display_text_for_order(
                    alternate_order, idx, sid, name
                )
            )
            score_candidates.append(self._format_score_text(score))

        return display_entries, display_candidates, score_candidates

    def _compute_card_metrics(self) -> Optional[_CardMetrics]:
        count = len(self.students)
        if count == 0:
            return None

        available = self._available_geometry
        key = (count, available.width(), available.height())
        if self._card_metrics is not None and self._card_metrics_key == key:
            return self._card_metrics

        columns = 10
        rows = max(1, math.ceil(count / columns))

        usable_width = max(available.width() - 160, 640)
        usable_height = max(available.height() - 240, 520)

        margins = self.grid_layout.contentsMargins()
        horizontal_spacing = max(14, int(usable_width * 0.01))
        vertical_spacing = max(18, int(usable_height * 0.035 / rows))

        spacing_total_x = horizontal_spacing * max(0, columns - 1)
        spacing_total_y = vertical_spacing * max(0, rows - 1)

        available_width_for_cards = (
            usable_width - margins.left() - margins.right() - spacing_total_x
        )
        available_height_for_cards = (
            usable_height - margins.top() - margins.bottom() - spacing_total_y
        )

        per_card_width = max(1.0, available_width_for_cards / columns)
        per_card_height = max(1.0, available_height_for_cards / rows)

        card_width = int(math.floor(per_card_width))
        card_height = int(math.floor(per_card_height))

        if per_card_width >= 120:
            card_width = max(card_width, 120)
        if per_card_height >= 180:
            card_height = max(card_height, 180)

        padding_h = max(12, int(card_width * 0.08))
        padding_v = max(14, int(card_height * 0.1))
        inner_spacing = max(6, int(card_height * 0.045))

        display_entries, display_candidates, score_candidates = self._collect_display_candidates()

        if not display_candidates:
            return None

        calligraphy_font = self._calligraphy_font or QApplication.font().family()
        if not calligraphy_font:
            calligraphy_font = QFont().family()

        try:
            probe_font = QFont(calligraphy_font, 64, QFont.Weight.Bold)
            metrics = QFontMetrics(probe_font)
        except Exception:
            probe_font = QFont()
            metrics = QFontMetrics(probe_font)

        widest_display = max(
            display_candidates,
            key=lambda text: metrics.tightBoundingRect(text).width(),
        )
        widest_score = max(
            score_candidates,
            key=lambda text: metrics.tightBoundingRect(text).width(),
        )

        usable_name_width = max(60, card_width - 2 * padding_h)
        content_height = max(80, card_height - 2 * padding_v - inner_spacing)
        name_height = int(content_height * 0.58)
        score_height = max(32, content_height - name_height)

        font_upper_bound = int(min(card_width * 0.28, card_height * 0.36))
        font_upper_bound = max(20, font_upper_bound)
        fit_minimum = 14

        name_fit = self._fit_font_size(
            widest_display,
            calligraphy_font,
            QFont.Weight.Bold,
            usable_name_width,
            name_height,
            fit_minimum,
            font_upper_bound,
        )
        score_fit = self._fit_font_size(
            widest_score,
            calligraphy_font,
            QFont.Weight.Bold,
            usable_name_width,
            score_height,
            fit_minimum,
            font_upper_bound,
        )

        final_font_size = max(fit_minimum, min(name_fit, score_fit, font_upper_bound))

        self._card_metrics = ScoreboardDialog._CardMetrics(
            count=count,
            columns=columns,
            rows=rows,
            card_width=card_width,
            card_height=card_height,
            padding_h=padding_h,
            padding_v=padding_v,
            inner_spacing=inner_spacing,
            font_size=final_font_size,
            horizontal_spacing=horizontal_spacing,
            vertical_spacing=vertical_spacing,
        )
        self._card_metrics_key = key
        return self._card_metrics

    def _ensure_metrics(self) -> Optional[_CardMetrics]:
        metrics = self._compute_card_metrics()
        if metrics is not None:
            return metrics
        self._card_metrics = None
        self._card_metrics_key = None
        return None

    def _sort_students(self, order: Optional[str] = None) -> List[tuple[str, str, int]]:
        data = list(self.students)
        current_order = self._order if order is None else order
        if current_order == self.ORDER_ID:
            def _id_key(item: tuple[str, str, int]) -> tuple[int, str, str]:
                sid_text = str(item[0]).strip()
                try:
                    sid_value = int(sid_text)
                except (TypeError, ValueError):
                    sid_value = sys.maxsize
                return (sid_value, sid_text, item[1])

            data.sort(key=_id_key)
        else:
            def _rank_key(item: tuple[str, str, int]) -> tuple[int, int, str]:
                sid_text = str(item[0]).strip()
                try:
                    sid_value = int(sid_text)
                except (TypeError, ValueError):
                    sid_value = sys.maxsize
                return (-item[2], sid_value, item[1])

            data.sort(key=_rank_key)
        return data

    def _create_card(
        self,
        display_text: str,
        score_text: str,
        card_width: int,
        card_height: int,
        font_size: int,
        padding_h: int,
        padding_v: int,
        inner_spacing: int,
    ) -> QWidget:
        calligraphy_font = self._calligraphy_font
        wrapper = QWidget()
        wrapper.setProperty("class", "scoreboardWrapper")
        wrapper.setFixedSize(card_width, card_height)
        wrapper.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)

        layout = QVBoxLayout(wrapper)
        layout.setContentsMargins(padding_h, padding_v, padding_h, padding_v)
        layout.setSpacing(inner_spacing)

        name_label = QLabel(display_text or "未命名")
        name_label.setProperty("class", "scoreboardName")
        name_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        name_label.setWordWrap(False)
        name_label.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed
        )
        name_label.setFont(QFont(calligraphy_font, font_size, QFont.Weight.Bold))
        name_label.setStyleSheet("margin: 0px; padding: 0px;")
        layout.addWidget(name_label)

        score_label = QLabel(score_text)
        score_label.setProperty("class", "scoreboardScore")
        score_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        score_label.setWordWrap(False)
        score_label.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed
        )
        score_label.setFont(QFont(calligraphy_font, font_size, QFont.Weight.Bold))
        score_label.setStyleSheet(f"margin-top: {max(6, inner_spacing // 2)}px;")
        layout.addWidget(score_label)

        layout.addStretch(1)
        return wrapper

    def _format_display_text(self, index: int, sid: str, name: str) -> str:
        return self._format_display_text_for_order(self._order, index, sid, name)

    def _format_display_text_for_order(
        self, order: str, index: int, sid: str, name: str
    ) -> str:
        clean_name = (name or "").strip() or "未命名"
        if order == self.ORDER_ID:
            sid_display = str(sid).strip() or "—"
            return f"{sid_display}.{clean_name}"
        return f"{index + 1}.{clean_name}"

    @staticmethod
    def _format_score_text(score: int | float | str) -> str:
        text = "—"
        try:
            value = float(score)
        except (TypeError, ValueError):
            score_str = str(score).strip()
            if score_str and score_str.lower() != "none":
                text = score_str
        else:
            if math.isfinite(value):
                if abs(value - int(value)) < 1e-6:
                    text = str(int(round(value)))
                else:
                    text = f"{value:.2f}".rstrip("0").rstrip(".")
        return f"{text} 分"

    @staticmethod
    def _fit_font_size(
        text: str,
        family: str,
        weight: QFont.Weight,
        max_width: int,
        max_height: int,
        minimum: int,
        maximum: int,
    ) -> int:
        if not text:
            return max(6, min(minimum, maximum))
        if max_width <= 0 or max_height <= 0:
            return max(6, min(minimum, maximum))
        lower = max(6, min(minimum, maximum))
        upper = max(6, max(minimum, maximum))
        for size in range(upper, lower - 1, -1):
            font = QFont(family, size, weight)
            metrics = QFontMetrics(font)
            rect = metrics.tightBoundingRect(text)
            if rect.width() <= max_width and rect.height() <= max_height:
                return size
        return lower

    def _populate_grid(self) -> None:
        self._clear_grid()
        count = len(self.students)
        layout = self.grid_layout
        calligraphy_font = self._calligraphy_font

        if count == 0:
            empty = QLabel("暂无成绩数据")
            empty.setAlignment(Qt.AlignmentFlag.AlignCenter)
            empty.setFont(QFont(calligraphy_font, 28, QFont.Weight.Bold))
            empty.setStyleSheet("color: #103d73;")
            layout.addWidget(empty, 0, 0, Qt.AlignmentFlag.AlignCenter)
            layout.setRowStretch(0, 1)
            layout.setColumnStretch(0, 1)
            self._grid_row_count = 1
            self._grid_column_count = 1
            return

        metrics = self._ensure_metrics()
        if metrics is None:
            return

        layout.setHorizontalSpacing(metrics.horizontal_spacing)
        layout.setVerticalSpacing(metrics.vertical_spacing)

        for column in range(metrics.columns):
            layout.setColumnStretch(column, 1)
            layout.setColumnMinimumWidth(column, metrics.card_width)
        for row in range(metrics.rows):
            layout.setRowStretch(row, 1)
            layout.setRowMinimumHeight(row, metrics.card_height)

        self._grid_row_count = metrics.rows
        self._grid_column_count = metrics.columns

        container = layout.parentWidget()
        if container is not None:
            margins = layout.contentsMargins()
            total_width = (
                metrics.columns * metrics.card_width
                + metrics.horizontal_spacing * max(0, metrics.columns - 1)
                + margins.left()
                + margins.right()
            )
            total_height = (
                metrics.rows * metrics.card_height
                + metrics.vertical_spacing * max(0, metrics.rows - 1)
                + margins.top()
                + margins.bottom()
            )
            container.setMinimumSize(total_width, total_height)

        display_entries, _, _ = self._collect_display_candidates()

        for idx, display_text, score_text in display_entries:
            row = idx // metrics.columns
            column = idx % metrics.columns
            card = self._create_card(
                display_text,
                score_text,
                metrics.card_width,
                metrics.card_height,
                metrics.font_size,
                metrics.padding_h,
                metrics.padding_v,
                metrics.inner_spacing,
            )
            layout.addWidget(card, row, column, Qt.AlignmentFlag.AlignCenter)

    def showEvent(self, event) -> None:  # type: ignore[override]
        super().showEvent(event)
        if self._pending_maximize:
            self._pending_maximize = False
            self.setWindowState(self.windowState() | Qt.WindowState.WindowMaximized)


@dataclass
class ClassRollState:
    current_group: str
    group_remaining: Dict[str, List[int]]
    group_last: Dict[str, Optional[int]]
    global_drawn: List[int]
    current_student: Optional[int] = None
    pending_student: Optional[int] = None

    def to_json(self) -> Dict[str, Any]:
        payload: Dict[str, Any] = {
            "current_group": self.current_group,
            "group_remaining": {group: list(values) for group, values in self.group_remaining.items()},
            "group_last": {group: value for group, value in self.group_last.items()},
            "global_drawn": list(self.global_drawn),
            "current_student": self.current_student,
            "pending_student": self.pending_student,
        }
        return payload

    @classmethod
    def from_mapping(cls, data: Mapping[str, Any]) -> Optional["ClassRollState"]:
        if not isinstance(data, Mapping):
            return None

        current_group = str(data.get("current_group", "") or "")

        remaining_raw = data.get("group_remaining", {})
        remaining: Dict[str, List[int]] = {}
        if isinstance(remaining_raw, Mapping):
            for key, values in remaining_raw.items():
                if not isinstance(key, str):
                    continue
                if isinstance(values, Iterable) and not isinstance(values, (str, bytes)):
                    cleaned: List[int] = []
                    for value in values:
                        try:
                            cleaned.append(int(value))
                        except (TypeError, ValueError):
                            continue
                    remaining[key] = cleaned

        last_raw = data.get("group_last", {})
        last: Dict[str, Optional[int]] = {}
        if isinstance(last_raw, Mapping):
            for key, value in last_raw.items():
                if not isinstance(key, str):
                    continue
                if value is None:
                    last[key] = None
                    continue
                try:
                    last[key] = int(value)
                except (TypeError, ValueError):
                    continue

        global_raw = data.get("global_drawn", [])
        global_drawn: List[int] = []
        if isinstance(global_raw, Iterable) and not isinstance(global_raw, (str, bytes)):
            for value in global_raw:
                try:
                    global_drawn.append(int(value))
                except (TypeError, ValueError):
                    continue

        def _parse_optional_int(value: Any) -> Optional[int]:
            if value is None:
                return None
            try:
                return int(value)
            except (TypeError, ValueError):
                return None

        current_student = _parse_optional_int(data.get("current_student"))
        pending_student = _parse_optional_int(data.get("pending_student"))

        return cls(
            current_group=current_group,
            group_remaining=remaining,
            group_last=last,
            global_drawn=global_drawn,
            current_student=current_student,
            pending_student=pending_student,
        )


class StudentPhotoOverlay(QWidget):
    closed_by_user = pyqtSignal()
    auto_closed = pyqtSignal()

    def __init__(self, owner: Optional[QWidget] = None) -> None:
        flags = (
            Qt.WindowType.Tool
            | Qt.WindowType.FramelessWindowHint
            | Qt.WindowType.WindowStaysOnTopHint
        )
        super().__init__(None, flags)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        self.setAttribute(Qt.WidgetAttribute.WA_ShowWithoutActivating, True)
        self.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._owner = owner
        self._current_pixmap = QPixmap()
        self._full_screen_active = False
        self._auto_close_duration_ms = 0
        self._auto_close_timer = QTimer(self)
        self._auto_close_timer.setSingleShot(True)
        self._auto_close_timer.timeout.connect(self._handle_auto_close)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        self._container = QFrame(self)
        self._container.setObjectName("studentPhotoContainer")
        self._container.setStyleSheet(
            """
            QFrame#studentPhotoContainer {
                background-color: rgba(0, 0, 0, 208);
                border-radius: 16px;
            }
            QToolButton {
                color: #f5f5f5;
                background-color: rgba(255, 255, 255, 48);
                border: 1px solid rgba(255, 255, 255, 96);
                border-radius: 11px;
                padding: 0;
            }
            QToolButton:hover {
                background-color: rgba(255, 255, 255, 88);
            }
            """
        )
        outer.addWidget(self._container)

        self._container_layout = QVBoxLayout(self._container)
        self._container_layout.setContentsMargins(28, 28, 28, 20)
        self._container_layout.setSpacing(12)

        self._photo_label = QLabel(self._container)
        self._photo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._photo_label.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum)
        self._container_layout.addWidget(self._photo_label, 1)

        button_row = QHBoxLayout()
        button_row.setContentsMargins(0, 0, 0, 0)
        button_row.setSpacing(0)
        self._left_close = self._make_close_button()
        self._right_close = self._make_close_button()
        button_row.addWidget(self._left_close, 0, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignBottom)
        button_row.addStretch(1)
        button_row.addWidget(self._right_close, 0, Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignBottom)
        self._container_layout.addLayout(button_row)

        self._left_close.clicked.connect(lambda: self._handle_close_request(manual=True))
        self._right_close.clicked.connect(lambda: self._handle_close_request(manual=True))

    def update_owner(self, owner: Optional[QWidget]) -> None:
        self._owner = owner

    def _make_close_button(self) -> QToolButton:
        button = QToolButton(self._container)
        button.setAutoRaise(True)
        button.setCursor(Qt.CursorShape.PointingHandCursor)
        button.setText("X")
        button.setToolTip("关闭照片")
        button.setFixedSize(24, 24)
        return button

    def _handle_close_request(self, *, manual: bool) -> None:
        self._auto_close_timer.stop()
        super().hide()
        if manual:
            self.closed_by_user.emit()
        else:
            self.auto_closed.emit()

    def _handle_auto_close(self) -> None:
        self._handle_close_request(manual=False)

    def cancel_auto_close(self) -> None:
        self._auto_close_timer.stop()

    def schedule_auto_close(self, duration_ms: int) -> None:
        self._auto_close_duration_ms = max(0, int(duration_ms))
        self._auto_close_timer.stop()
        if self._auto_close_duration_ms > 0 and self.isVisible():
            self._auto_close_timer.start(self._auto_close_duration_ms)

    def display_photo(self, pixmap: QPixmap, screen_rect: QRect, duration_ms: int) -> None:
        if pixmap.isNull():
            self.hide()
            return

        self._auto_close_duration_ms = max(0, int(duration_ms))
        self._auto_close_timer.stop()
        self._current_pixmap = pixmap
        max_size = screen_rect.size()
        original_size = pixmap.size()
        can_cover_screen = (
            original_size.width() >= screen_rect.width()
            and original_size.height() >= screen_rect.height()
        )
        full_screen = can_cover_screen
        if full_screen:
            scaled = pixmap.scaled(
                max_size,
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation,
            )
        else:
            if (
                original_size.width() <= screen_rect.width()
                and original_size.height() <= screen_rect.height()
            ):
                scaled = pixmap
            else:
                scaled = pixmap.scaled(
                    max_size,
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation,
                )
        self._apply_full_screen_style(full_screen)
        self._photo_label.setPixmap(scaled)
        self._photo_label.setFixedSize(scaled.size())
        if full_screen:
            self.resize(screen_rect.size())
            self.move(screen_rect.topLeft())
            self._container.resize(self.size())
        else:
            if self.layout() is not None:
                self.layout().activate()
            self.adjustSize()
            window_size = self.size()
            width = min(window_size.width(), screen_rect.width())
            height = min(window_size.height(), screen_rect.height())
            x = screen_rect.x() + max(0, (screen_rect.width() - width) // 2)
            y = screen_rect.y() + max(0, (screen_rect.height() - height) // 2)
            self.resize(width, height)
            self.move(int(x), int(y))
            self._container.adjustSize()
        self.show()
        self._stack_below_owner()
        self.schedule_auto_close(self._auto_close_duration_ms)

    def _apply_full_screen_style(self, full_screen: bool) -> None:
        if self._full_screen_active == full_screen:
            return
        self._full_screen_active = full_screen
        if full_screen:
            self._container_layout.setContentsMargins(24, 24, 24, 32)
            self._container.setStyleSheet(
                """
                QFrame#studentPhotoContainer {
                    background-color: rgba(0, 0, 0, 230);
                    border-radius: 0px;
                }
                QToolButton {
                    color: #f5f5f5;
                    background-color: rgba(255, 255, 255, 48);
                    border: 1px solid rgba(255, 255, 255, 96);
                    border-radius: 11px;
                    padding: 0;
                }
                QToolButton:hover {
                    background-color: rgba(255, 255, 255, 96);
                }
                """
            )
        else:
            self._container_layout.setContentsMargins(8, 8, 8, 12)
            self._container.setStyleSheet(
                """
                QFrame#studentPhotoContainer {
                    background-color: transparent;
                    border: none;
                }
                QToolButton {
                    color: #f5f5f5;
                    background-color: rgba(255, 255, 255, 48);
                    border: 1px solid rgba(255, 255, 255, 96);
                    border-radius: 11px;
                    padding: 0;
                }
                QToolButton:hover {
                    background-color: rgba(255, 255, 255, 88);
                }
                """
            )

    def _stack_below_owner(self) -> None:
        owner = self._owner
        if owner is None:
            return
        try:
            overlay_hwnd = int(self.winId()) if self.winId() else 0
            owner_hwnd = int(owner.winId()) if owner.winId() else 0
        except Exception:
            overlay_hwnd = 0
            owner_hwnd = 0
        if overlay_hwnd == 0 or owner_hwnd == 0:
            try:
                owner.raise_()
            except Exception:
                pass
            return
        if win32gui is not None and win32con is not None:
            flags = win32con.SWP_NOMOVE | win32con.SWP_NOSIZE | win32con.SWP_NOACTIVATE | win32con.SWP_NOOWNERZORDER
            try:
                win32gui.SetWindowPos(overlay_hwnd, win32con.HWND_TOPMOST, 0, 0, 0, 0, flags)
                win32gui.SetWindowPos(owner_hwnd, overlay_hwnd, 0, 0, 0, 0, flags)
            except Exception:
                logger.debug("win32gui.SetWindowPos failed for photo overlay", exc_info=True)
        else:
            try:
                self.lower()
                owner.raise_()
            except Exception:
                pass
        try:
            owner.raise_()
        except Exception:
            pass


class RollCallTimerWindow(QWidget):
    """集成点名与计时的主功能窗口。"""
    window_closed = pyqtSignal()
    visibility_changed = pyqtSignal(bool)

    STUDENT_FILE = "students.xlsx"
    ENCRYPTED_STUDENT_FILE = "students.xlsx.enc"
    MIN_FONT_SIZE = 5
    MAX_FONT_SIZE = 220

    def __init__(
        self,
        settings_manager: SettingsManager,
        student_workbook: Optional[StudentWorkbook],
        parent: Optional[QWidget] = None,
        *,
        defer_password_prompt: bool = False,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("点名 / 计时")
        flags = (
            Qt.WindowType.Window
            | Qt.WindowType.WindowTitleHint
            | Qt.WindowType.WindowCloseButtonHint
            | Qt.WindowType.WindowStaysOnTopHint
            | Qt.WindowType.CustomizeWindowHint
        )
        self.setWindowFlags(flags)
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose)
        self.settings_manager = settings_manager
        self._encrypted_file_path = self.ENCRYPTED_STUDENT_FILE
        self.student_workbook: Optional[StudentWorkbook] = student_workbook
        base_dataframe: Optional[PandasDataFrame] = None
        if PANDAS_READY and self.student_workbook is not None:
            try:
                base_dataframe = self.student_workbook.get_active_dataframe()
            except Exception:
                base_dataframe = pd.DataFrame(columns=["学号", "姓名", "分组", "成绩"])
        if base_dataframe is None and PANDAS_READY:
            base_dataframe = pd.DataFrame(columns=["学号", "姓名", "分组", "成绩"])
        self.student_data = base_dataframe
        self._student_data_pending_load = False
        encrypted_state, encrypted_password = _get_session_student_encryption()
        self._student_file_encrypted = bool(encrypted_state)
        self._student_password = encrypted_password
        if defer_password_prompt:
            base_empty = True
            if PANDAS_READY and isinstance(self.student_data, pd.DataFrame):
                base_empty = getattr(self.student_data, "empty", True)
            elif self.student_data is not None:
                base_empty = False
            if base_empty:
                has_plain = os.path.exists(self.STUDENT_FILE)
                has_encrypted = os.path.exists(self._encrypted_file_path)
                if has_plain or has_encrypted:
                    self._student_data_pending_load = True
                    if self.student_data is None and PANDAS_READY:
                        self.student_data = pd.DataFrame(columns=["学号", "姓名", "分组", "成绩"])
        try:
            self._rng = random.SystemRandom()
        except NotImplementedError:
            self._rng = random.Random()

        s = self.settings_manager.load_settings().get("RollCallTimer", {})
        def _get_int(key: str, default: int) -> int:
            try:
                return int(s.get(key, str(default)))
            except (TypeError, ValueError):
                return default
        apply_geometry_from_text(self, s.get("geometry", "420x240+180+180"))
        self.setMinimumSize(260, 160)
        # 记录初始最小宽高，供后续还原窗口尺寸时使用
        self._base_minimum_width = self.minimumWidth()
        self._base_minimum_height = self.minimumHeight()
        self._ensure_min_width = self._base_minimum_width
        self._ensure_min_height = self._base_minimum_height

        self.mode = s.get("mode", "roll_call") if s.get("mode", "roll_call") in {"roll_call", "timer"} else "roll_call"
        self.timer_modes = ["countdown", "stopwatch", "clock"]
        self.timer_mode_index = self.timer_modes.index(s.get("timer_mode", "countdown")) if s.get("timer_mode", "countdown") in self.timer_modes else 0
        self._active_timer_mode: Optional[str] = None

        self.timer_countdown_minutes = _get_int("timer_countdown_minutes", 5)
        self.timer_countdown_seconds = _get_int("timer_countdown_seconds", 0)
        self.timer_sound_enabled = str_to_bool(s.get("timer_sound_enabled", "True"), True)

        self.show_id = str_to_bool(s.get("show_id", "True"), True)
        self.show_name = str_to_bool(s.get("show_name", "True"), True)
        self.show_photo = str_to_bool(s.get("show_photo", "False"), False)
        self.photo_duration_seconds = max(0, _get_int("photo_duration_seconds", 0))
        if not (self.show_id or self.show_name or self.show_photo):
            self.show_id = True

        self.photo_root_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "student_photos")
        self._photo_extensions = [".png", ".jpg", ".jpeg", ".bmp", ".gif"]
        self._photo_overlay: Optional[StudentPhotoOverlay] = None
        self._last_photo_student_id: Optional[str] = None
        self._photo_manual_hidden = False
        self._ensure_photo_root_directory()

        self.current_class_name = str(s.get("current_class", "")).strip()
        self.current_group_name = s.get("current_group", "全部")
        self.groups = ["全部"]

        self.current_student_index: Optional[int] = None
        self._placeholder_on_show = True
        self._group_all_indices: Dict[str, List[int]] = {}
        self._group_remaining_indices: Dict[str, List[int]] = {}
        self._group_last_student: Dict[str, Optional[int]] = {}
        # 记录各分组初始的随机顺序，便于在界面切换时保持未点名名单不被重新洗牌
        self._group_initial_sequences: Dict[str, List[int]] = {}
        # 记录每个分组已点过名的学生索引，便于核对剩余名单
        self._group_drawn_history: Dict[str, set[int]] = {}
        # 统一维护一个全局已点名集合，确保“全部”分组与子分组状态一致
        self._global_drawn_students: set[int] = set()
        self._student_groups: Dict[int, set[str]] = {}
        self._class_roll_states: Dict[str, ClassRollState] = {}
        self.timer_seconds_left = max(0, _get_int("timer_seconds_left", self.timer_countdown_minutes * 60 + self.timer_countdown_seconds))
        self.timer_stopwatch_seconds = max(0, _get_int("timer_stopwatch_seconds", 0))
        self.timer_running = str_to_bool(s.get("timer_running", "False"), False)

        order_value = str(s.get("scoreboard_order", "rank")).strip().lower()
        self.scoreboard_order = order_value if order_value in {"rank", "id"} else "rank"
        saved_encrypted = str_to_bool(s.get("students_encrypted", bool_to_str(self._student_file_encrypted)), self._student_file_encrypted)
        disk_encrypted = os.path.exists(self._encrypted_file_path) and not os.path.exists(self.STUDENT_FILE)
        if disk_encrypted:
            self._student_file_encrypted = True
        elif not saved_encrypted:
            self._student_file_encrypted = False
            self._student_password = None

        self.last_id_font_size = max(self.MIN_FONT_SIZE, _get_int("id_font_size", 48))
        self.last_name_font_size = max(self.MIN_FONT_SIZE, _get_int("name_font_size", 60))
        self.last_timer_font_size = max(self.MIN_FONT_SIZE, _get_int("timer_font_size", 56))

        self.count_timer = QTimer(self); self.count_timer.setInterval(1000); self.count_timer.timeout.connect(self._on_count_timer)
        self.clock_timer = QTimer(self); self.clock_timer.setInterval(1000); self.clock_timer.timeout.connect(self._update_clock)

        self.tts_manager: Optional[TTSManager] = None
        self.speech_enabled = str_to_bool(s.get("speech_enabled", "False"), False)
        self.selected_voice_id = s.get("speech_voice_id", "")
        manager = TTSManager(self.selected_voice_id, parent=self)
        self.tts_manager = manager
        if not manager.available:
            self.speech_enabled = False
        self._speech_issue_reported = False
        self._speech_check_scheduled = False
        self._pending_passive_student: Optional[int] = None
        self._score_persist_failed = False
        self._score_write_lock = threading.Lock()

        # QFontDatabase 在 Qt 6 中以静态方法为主，这里直接调用类方法避免实例化失败
        families_list = []
        get_families = getattr(QFontDatabase, "families", None)
        if callable(get_families):
            try:
                families_list = list(get_families())
            except TypeError:
                # 个别绑定版本可能要求显式写入枚举参数
                try:
                    families_list = list(get_families(QFontDatabase.WritingSystem.Any))  # type: ignore[arg-type]
                except Exception:
                    families_list = []
        families = set(families_list)
        self.name_font_family = "楷体" if "楷体" in families else ("KaiTi" if "KaiTi" in families else "Microsoft YaHei UI")

        # 使用轻量级的延迟写入机制，避免频繁操作磁盘。
        self._save_timer = QTimer(self)
        self._save_timer.setSingleShot(True)
        self._save_timer.setInterval(250)
        self._save_timer.timeout.connect(self.save_settings)

        self._build_ui()
        if self.student_workbook is not None and not self._student_data_pending_load:
            self._apply_student_workbook(self.student_workbook, propagate=False)
        else:
            self._set_student_dataframe(self.student_data, propagate=False)
        self._apply_saved_fonts()
        self._update_menu_state()
        self._restore_group_state(s)
        self.update_mode_ui(force_timer_reset=self.mode == "timer")
        self.on_group_change(initial=True)
        self.display_current_student()
        self._update_encryption_button()

    def _build_ui(self) -> None:
        self.setStyleSheet("background-color: #f4f5f7;")
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        toolbar_layout = QVBoxLayout()
        toolbar_layout.setContentsMargins(0, 0, 0, 0)
        toolbar_layout.setSpacing(2)

        top = QHBoxLayout()
        top.setContentsMargins(0, 0, 0, 0)
        top.setSpacing(4)
        self.title_label = QLabel("点名"); f = QFont("Microsoft YaHei UI", 10, QFont.Weight.Bold)
        self.title_label.setFont(f); self.title_label.setStyleSheet("color: #202124;")
        self.title_label.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        top.addWidget(self.title_label, 0, Qt.AlignmentFlag.AlignLeft)

        self.mode_button = QPushButton("切换到计时")
        mode_font = QFont("Microsoft YaHei UI", 9, QFont.Weight.Medium)
        self.mode_button.setFont(mode_font)
        fm = self.mode_button.fontMetrics()
        max_text = max(("切换到计时", "切换到点名"), key=lambda t: fm.horizontalAdvance(t))
        target_width = fm.horizontalAdvance(max_text) + 28
        self.mode_button.setMinimumWidth(target_width)
        self.mode_button.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Fixed)
        control_height = recommended_control_height(mode_font, extra=14, minimum=36)
        apply_button_style(self.mode_button, ButtonStyles.TOOLBAR, height=control_height)
        self.mode_button.clicked.connect(self.toggle_mode)
        top.addWidget(self.mode_button, 0, Qt.AlignmentFlag.AlignLeft)

        compact_font = QFont("Microsoft YaHei UI", 9, QFont.Weight.Medium)
        toolbar_height = recommended_control_height(compact_font, extra=14, minimum=36)

        def _setup_secondary_button(button: QPushButton) -> None:
            apply_button_style(button, ButtonStyles.TOOLBAR, height=toolbar_height)
            button.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Fixed)
            button.setFont(compact_font)

        def _lock_button_width(button: QPushButton) -> None:
            """将按钮的宽度锁定在推荐值，避免随布局波动。"""

            hint = button.sizeHint()
            width = max(hint.width(), button.minimumSizeHint().width())
            button.setMinimumWidth(width)
            button.setMaximumWidth(width)
            button.setSizePolicy(QSizePolicy.Policy.Fixed, button.sizePolicy().verticalPolicy())

        control_bar = QWidget()
        control_bar.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        control_layout = QHBoxLayout(control_bar)
        control_layout.setContentsMargins(0, 0, 0, 0)
        control_layout.setSpacing(2)

        def _recycle_button(button: Optional[QPushButton]) -> None:
            if button is None:
                return
            parent = button.parentWidget()
            layout = parent.layout() if parent is not None else None
            if isinstance(layout, QHBoxLayout):
                layout.removeWidget(button)
            button.setParent(None)
            button.deleteLater()

        existing_class_button = getattr(self, "class_button", None)
        if isinstance(existing_class_button, QPushButton):
            _recycle_button(existing_class_button)
        # 仅保留一个班级切换按钮，并将其固定在“重置”按钮左侧。
        self.class_button = QPushButton("班级"); _setup_secondary_button(self.class_button)
        self.class_button.clicked.connect(self.show_class_selector)
        control_layout.addWidget(self.class_button)

        self.reset_button = QPushButton("重置"); _setup_secondary_button(self.reset_button)
        self.reset_button.clicked.connect(self.reset_roll_call_pools)
        _lock_button_width(self.reset_button)
        control_layout.addWidget(self.reset_button)

        self.showcase_button = QPushButton("展示"); _setup_secondary_button(self.showcase_button)
        self.showcase_button.clicked.connect(self.show_scoreboard)
        control_layout.addWidget(self.showcase_button)

        self.encrypt_button = QPushButton(""); _setup_secondary_button(self.encrypt_button)
        self.encrypt_button.clicked.connect(self._on_encrypt_button_clicked)
        control_layout.addWidget(self.encrypt_button)

        top.addWidget(control_bar, 0, Qt.AlignmentFlag.AlignLeft)
        top.addStretch(1)

        self.menu_button = QToolButton(); self.menu_button.setText("..."); self.menu_button.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        self.menu_button.setFixedSize(toolbar_height, toolbar_height)
        self.menu_button.setStyleSheet("font-size: 18px; padding-bottom: 6px;")
        self.main_menu = self._build_menu(); self.menu_button.setMenu(self.main_menu)
        top.addWidget(self.menu_button, 0, Qt.AlignmentFlag.AlignRight)
        toolbar_layout.addLayout(top)

        group_row = QHBoxLayout()
        group_row.setContentsMargins(0, 0, 0, 0)
        group_row.setSpacing(2)

        self.group_label = QLabel("分组")
        self.group_label.setFont(QFont("Microsoft YaHei UI", 9, QFont.Weight.Medium))
        self.group_label.setStyleSheet("color: #3c4043;")
        self.group_label.setFixedHeight(toolbar_height)
        self.group_label.setAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
        self.group_label.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        group_row.addWidget(self.group_label, 0, Qt.AlignmentFlag.AlignLeft)

        group_container = QWidget()
        group_container.setFixedHeight(toolbar_height)
        group_container.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        group_container_layout = QHBoxLayout(group_container)
        group_container_layout.setContentsMargins(0, 0, 0, 0)
        group_container_layout.setSpacing(4)

        self.group_container = group_container

        self.group_bar = QWidget(group_container)
        self.group_bar.setFixedHeight(toolbar_height)
        self.group_bar.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.group_bar_layout = QHBoxLayout(self.group_bar)
        self.group_bar_layout.setContentsMargins(0, 0, 0, 0)
        self.group_bar_layout.setSpacing(1)
        self.group_button_group = QButtonGroup(self)
        self.group_button_group.setExclusive(True)
        self.group_buttons: Dict[str, QPushButton] = {}
        self._rebuild_group_buttons_ui()
        group_container_layout.addWidget(self.group_bar, 1, Qt.AlignmentFlag.AlignLeft)

        self.list_button = QPushButton("名单"); _setup_secondary_button(self.list_button)
        self.list_button.clicked.connect(self.show_student_selector)
        _lock_button_width(self.list_button)
        group_container_layout.addWidget(self.list_button, 0, Qt.AlignmentFlag.AlignLeft)

        self.add_score_button = QPushButton("加分"); _setup_secondary_button(self.add_score_button)
        self.add_score_button.setEnabled(False)
        self.add_score_button.clicked.connect(self.increment_current_score)
        self.add_score_button.setMinimumWidth(self.add_score_button.sizeHint().width())
        group_container_layout.addWidget(self.add_score_button, 0, Qt.AlignmentFlag.AlignLeft)

        group_row.addWidget(group_container, 1, Qt.AlignmentFlag.AlignLeft)
        group_row.addStretch(1)
        toolbar_layout.addLayout(group_row)
        layout.addLayout(toolbar_layout)

        self.stack = QStackedWidget(); layout.addWidget(self.stack, 1)

        self.roll_call_frame = ClickableFrame(); self.roll_call_frame.setFrameShape(QFrame.Shape.NoFrame)
        rl = QGridLayout(self.roll_call_frame); rl.setContentsMargins(6, 6, 6, 6); rl.setSpacing(6)
        self.id_label = QLabel(""); self.name_label = QLabel("")
        for lab in (self.id_label, self.name_label):
            lab.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lab.setStyleSheet("color: #ffffff; background-color: #1a73e8; border-radius: 8px; padding: 8px;")
            lab.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.score_label = QLabel("成绩：--")
        self.score_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.score_label.setFont(QFont("Microsoft YaHei UI", 24, QFont.Weight.DemiBold))
        self.score_label.setStyleSheet(
            "color: #0b57d0;"
            " background-color: #e8f0fe;"
            " border-radius: 12px;"
            " padding: 2px 16px;"
            " margin: 0px;"
        )

        rl.addWidget(self.id_label, 0, 0); rl.addWidget(self.name_label, 0, 1)
        rl.addWidget(self.score_label, 1, 0, 1, 2)
        self.stack.addWidget(self.roll_call_frame)

        self.timer_frame = QWidget(); tl = QVBoxLayout(self.timer_frame); tl.setContentsMargins(6, 6, 6, 6); tl.setSpacing(8)
        self.time_display_label = QLabel("00:00"); self.time_display_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.time_display_label.setStyleSheet("color: #ffffff; background-color: #202124; border-radius: 8px; padding: 8px;")
        self.time_display_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        tl.addWidget(self.time_display_label, 1)

        ctrl = QHBoxLayout(); ctrl.setSpacing(4)
        self.timer_mode_button = QPushButton("倒计时"); self.timer_mode_button.clicked.connect(self.toggle_timer_mode)
        self.timer_start_pause_button = QPushButton("开始"); self.timer_start_pause_button.clicked.connect(self.start_pause_timer)
        self.timer_reset_button = QPushButton("重置"); self.timer_reset_button.clicked.connect(self.reset_timer)
        self.timer_set_button = QPushButton("设定"); self.timer_set_button.clicked.connect(self.set_countdown_time)
        for b in (self.timer_mode_button, self.timer_start_pause_button, self.timer_reset_button, self.timer_set_button):
            b.setFont(compact_font)
        timer_height = recommended_control_height(compact_font, extra=14, minimum=36)
        for b in (self.timer_mode_button, self.timer_start_pause_button, self.timer_reset_button, self.timer_set_button):
            apply_button_style(b, ButtonStyles.TOOLBAR, height=timer_height)
            b.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            ctrl.addWidget(b)
        tl.addLayout(ctrl); self.stack.addWidget(self.timer_frame)

        self.roll_call_frame.clicked.connect(self.roll_student)
        self.id_label.installEventFilter(self); self.name_label.installEventFilter(self)

    def _update_encryption_button(self) -> None:
        if not hasattr(self, "encrypt_button"):
            return
        disk_encrypted = os.path.exists(self._encrypted_file_path) and not os.path.exists(self.STUDENT_FILE)
        if disk_encrypted and not self._student_file_encrypted:
            self._student_file_encrypted = True
        elif not disk_encrypted and self._student_file_encrypted and not os.path.exists(self._encrypted_file_path):
            self._student_file_encrypted = False
            self._student_password = None
        self.encrypt_button.setText("解密" if self._student_file_encrypted else "加密")
        if self._student_file_encrypted:
            self.encrypt_button.setToolTip("当前学生数据已加密，点击输入密码以解密或更新。")
        else:
            self.encrypt_button.setToolTip("点击为 students.xlsx 设置密码并生成加密文件。")

    def _on_encrypt_button_clicked(self) -> None:
        if self._student_file_encrypted:
            self._handle_decrypt_student_file()
        else:
            self._handle_encrypt_student_file()

    def _prompt_new_encryption_password(self) -> Optional[str]:
        password, ok = PasswordSetupDialog.get_new_password(
            self,
            "设置加密密码",
            "请输入新的加密密码：",
            "请再次输入密码以确认：",
        )
        if not ok or not password:
            show_quiet_information(self, "未能成功设置密码，已取消加密操作。")
            return None
        return password

    def _prompt_existing_encryption_password(self, title: str) -> Optional[str]:
        attempts = 0
        while attempts < 3:
            password, ok = PasswordPromptDialog.get_password(
                self,
                title,
                "请输入当前的加密密码：",
                allow_empty=False,
            )
            if not ok:
                return None
            password = password.strip()
            if not password:
                show_quiet_information(self, "密码不能为空，请重新输入。")
                attempts += 1
                continue
            return password
        show_quiet_information(self, "密码输入次数过多，操作已取消。")
        return None

    def _set_student_dataframe(self, df: Optional[PandasDataFrame], *, propagate: bool = True) -> None:
        if not PANDAS_READY:
            self.student_data = df
            return
        if df is None:
            df = pd.DataFrame(columns=["学号", "姓名", "分组", "成绩"])
        try:
            working = df.copy()
        except Exception:
            working = pd.DataFrame(df)
        self.student_data = working
        self.groups = ["全部"]
        if not working.empty:
            group_values = {
                str(g).strip().upper()
                for g in working.get("分组", pd.Series([], dtype="object")).dropna()
                if str(g).strip()
            }
            self.groups.extend(sorted(group_values))
        if self.current_group_name not in self.groups:
            self.current_group_name = "全部"
        self._group_all_indices = {}
        self._group_remaining_indices = {}
        self._group_last_student = {}
        self._group_initial_sequences = {}
        self._group_drawn_history = {}
        self._global_drawn_students = set()
        self._student_groups = {}
        self._rebuild_group_buttons_ui()
        self._rebuild_group_indices()
        self._ensure_group_pool(self.current_group_name, force_reset=True)
        self.current_student_index = None
        self._pending_passive_student = None
        self._restore_active_class_state()
        self._store_active_class_state()
        self._update_class_button_label()
        if propagate:
            self._propagate_student_dataframe()
        self.display_current_student()

    def _apply_student_workbook(self, workbook: StudentWorkbook, *, propagate: bool) -> None:
        self.student_workbook = workbook
        self._prune_orphan_class_states()
        if not PANDAS_READY:
            self.current_class_name = workbook.active_class
            self.student_data = None
            return
        if self.current_class_name:
            workbook.set_active_class(self.current_class_name)
        self.current_class_name = workbook.active_class
        df = workbook.get_active_dataframe()
        self._set_student_dataframe(df, propagate=propagate)

    def _snapshot_current_class(self) -> None:
        if not PANDAS_READY:
            return
        if self.student_workbook is None:
            return
        if self.student_data is None or not isinstance(self.student_data, pd.DataFrame):
            return
        class_name = (self.current_class_name or self.student_workbook.active_class or "").strip()
        if not class_name:
            available = self.student_workbook.class_names()
            class_name = available[0] if available else self.student_workbook.active_class
        if class_name not in self.student_workbook.class_names():
            class_name = self.student_workbook.active_class
        try:
            snapshot = self.student_data.copy()
        except Exception:
            snapshot = pd.DataFrame(self.student_data)
        self.student_workbook.update_class(class_name, snapshot)
        self.student_workbook.set_active_class(class_name)
        self.current_class_name = class_name
        self._store_active_class_state(class_name)

    def _resolve_active_class_name(self) -> str:
        base = self.current_class_name
        if not base and self.student_workbook is not None:
            base = self.student_workbook.active_class
        return str(base or "").strip()

    def _capture_roll_state(self) -> Optional[ClassRollState]:
        if not PANDAS_READY:
            return None
        if not isinstance(self.student_data, pd.DataFrame):
            return None

        base_sets: Dict[str, Set[int]] = {}
        for group, indices in self._group_all_indices.items():
            base_list = self._collect_base_indices(indices)
            base_sets[group] = set(base_list)

        if "全部" not in base_sets:
            try:
                base_sets["全部"] = set(self._collect_base_indices(list(self.student_data.index)))
            except Exception:
                base_sets["全部"] = set()

        all_set = base_sets.get("全部", set())

        remaining_payload: Dict[str, List[int]] = {}
        for group, indices in self._group_remaining_indices.items():
            base_set = base_sets.get(group, all_set)
            if base_set:
                restored = self._normalize_indices(indices, allowed=base_set)
            else:
                restored = []
            remaining_payload[group] = restored

        last_payload: Dict[str, Optional[int]] = {}
        for group, value in self._group_last_student.items():
            base_set = base_sets.get(group, all_set)
            if value is None:
                last_payload[group] = None
                continue
            try:
                idx = int(value)
            except (TypeError, ValueError):
                last_payload[group] = None
                continue
            if base_set and idx not in base_set:
                last_payload[group] = None
            else:
                last_payload[group] = idx

        global_drawn_payload: List[int] = []
        for value in sorted(self._global_drawn_students):
            try:
                idx = int(value)
            except (TypeError, ValueError):
                continue
            if not all_set or idx in all_set:
                global_drawn_payload.append(idx)

        if self.groups:
            if self.current_group_name in self.groups:
                target_group = self.current_group_name
            elif "全部" in self.groups:
                target_group = "全部"
            else:
                target_group = self.groups[0]
        else:
            target_group = ""

        def _sanitize_index(value: Any) -> Optional[int]:
            if value is None:
                return None
            try:
                idx = int(value)
            except (TypeError, ValueError):
                return None
            if all_set and idx not in all_set:
                return None
            return idx

        current_student = _sanitize_index(self.current_student_index)
        pending_student = _sanitize_index(self._pending_passive_student)

        return ClassRollState(
            current_group=target_group,
            group_remaining=remaining_payload,
            group_last=last_payload,
            global_drawn=global_drawn_payload,
            current_student=current_student,
            pending_student=pending_student,
        )

    def _store_active_class_state(self, class_name: Optional[str] = None) -> None:
        if not PANDAS_READY:
            return
        self._prune_orphan_class_states()
        target = (class_name or self._resolve_active_class_name()).strip()
        if not target:
            return
        snapshot = self._capture_roll_state()
        if snapshot is None:
            return
        self._class_roll_states[target] = snapshot

    def _prune_orphan_class_states(self) -> None:
        if not self._class_roll_states:
            return
        workbook = self.student_workbook
        if workbook is None:
            return
        try:
            valid = {str(name).strip() for name in workbook.class_names() if str(name).strip()}
        except Exception:
            valid = set()
        if not valid:
            self._class_roll_states.clear()
            return
        for stored_name in list(self._class_roll_states.keys()):
            trimmed = str(stored_name).strip()
            if not trimmed or trimmed not in valid:
                self._class_roll_states.pop(stored_name, None)

    def _encode_class_states(self) -> str:
        payload = {name: state.to_json() for name, state in self._class_roll_states.items()}
        return json.dumps(payload, ensure_ascii=False)

    def _parse_legacy_roll_state(self, section: Mapping[str, str]) -> Optional[ClassRollState]:
        def _load_dict(key: str) -> Dict[str, Any]:
            raw = section.get(key, "")
            if not raw:
                return {}
            try:
                data = json.loads(raw)
            except Exception:
                return {}
            return data if isinstance(data, dict) else {}

        remaining = _load_dict("group_remaining")
        last = _load_dict("group_last")

        global_drawn_raw = section.get("global_drawn", "")
        global_payload: List[int] = []
        if global_drawn_raw:
            try:
                payload = json.loads(global_drawn_raw)
            except Exception:
                payload = []
            if isinstance(payload, list):
                for value in payload:
                    try:
                        global_payload.append(int(value))
                    except (TypeError, ValueError):
                        continue

        payload_map: Dict[str, Any] = {
            "current_group": section.get("current_group", self.current_group_name),
            "group_remaining": remaining,
            "group_last": last,
            "global_drawn": global_payload,
        }
        return ClassRollState.from_mapping(payload_map)

    def _restore_active_class_state(self) -> None:
        if not PANDAS_READY:
            return
        class_name = self._resolve_active_class_name()
        if not class_name:
            return
        snapshot = self._class_roll_states.get(class_name)
        if snapshot is None:
            return
        self._apply_roll_state(snapshot)

    def _can_apply_roll_state(self) -> bool:
        """检查当前是否具备恢复点名状态所需的数据上下文。"""

        if not PANDAS_READY:
            return False
        if self._student_data_pending_load:
            return False
        return isinstance(self.student_data, pd.DataFrame)

    def _apply_roll_state(self, snapshot: ClassRollState) -> None:
        if not self._can_apply_roll_state():
            return

        remaining_data = snapshot.group_remaining or {}
        last_data = snapshot.group_last or {}
        restored_global: Set[int] = set()
        for value in snapshot.global_drawn:
            try:
                restored_global.add(int(value))
            except (TypeError, ValueError):
                continue

        existing_global = set(restored_global)
        self._global_drawn_students = set()
        self._group_drawn_history["全部"] = self._global_drawn_students

        for group, indices in remaining_data.items():
            if group not in self._group_all_indices:
                continue
            base_list = self._collect_base_indices(self._group_all_indices[group])
            base_set = set(base_list)
            if base_set:
                restored_list = self._normalize_indices(indices, allowed=base_set)
            else:
                restored_list = []
            self._group_remaining_indices[group] = restored_list

        for group, value in last_data.items():
            if group not in self._group_all_indices:
                continue
            if value is None:
                self._group_last_student[group] = None
                continue
            try:
                idx = int(value)
            except (TypeError, ValueError):
                continue
            base_indices = self._collect_base_indices(self._group_all_indices[group])
            base_set = set(base_indices)
            if base_set and idx not in base_set:
                continue
            self._group_last_student[group] = idx

        for group, base_indices in self._group_all_indices.items():
            normalized_base = self._collect_base_indices(base_indices)
            remaining_set = set(self._normalize_indices(self._group_remaining_indices.get(group, [])))
            drawn = {idx for idx in normalized_base if idx not in remaining_set}
            if group != "全部" and existing_global:
                drawn.update(idx for idx in existing_global if idx in normalized_base)
            seq = list(self._group_remaining_indices.get(group, []))
            seq.extend(idx for idx in normalized_base if idx not in seq)
            self._group_initial_sequences[group] = seq
            if group == "全部":
                self._global_drawn_students.update(drawn)
            else:
                self._group_drawn_history[group] = drawn
                self._global_drawn_students.update(drawn)

        if existing_global:
            self._global_drawn_students.update(existing_global)

        self._group_drawn_history["全部"] = self._global_drawn_students
        self._refresh_all_group_pool()

        target_group = snapshot.current_group.strip() if snapshot.current_group else ""
        if target_group not in self.groups:
            target_group = "全部" if "全部" in self.groups else (self.groups[0] if self.groups else "全部")
        self.current_group_name = target_group
        self._update_group_button_state(target_group)

        base_all = self._collect_base_indices(self._group_all_indices.get("全部", []))
        base_all_set = set(base_all)

        def _valid_index(value: Optional[int]) -> Optional[int]:
            if value is None:
                return None
            try:
                idx = int(value)
            except (TypeError, ValueError):
                return None
            if base_all_set and idx not in base_all_set:
                return None
            return idx

        self.current_student_index = _valid_index(snapshot.current_student)
        self._pending_passive_student = _valid_index(snapshot.pending_student)

        self._store_active_class_state(self._resolve_active_class_name())

    def _update_class_button_label(self) -> None:
        if not hasattr(self, "class_button"):
            return
        name = ""
        if self.student_workbook is not None:
            base_name = self.current_class_name or self.student_workbook.active_class
            name = base_name.strip()
        text = name or "班级"
        self.class_button.setText(text)
        metrics = self.class_button.fontMetrics()
        baseline = metrics.horizontalAdvance("班级")
        active_width = metrics.horizontalAdvance(text)
        minimum = max(baseline, active_width) + 24
        if self.class_button.minimumWidth() != minimum:
            self.class_button.setMinimumWidth(minimum)
        has_data = self.student_workbook is not None and not self.student_workbook.is_empty()
        can_select = self.mode == "roll_call" and (has_data or self._student_data_pending_load)
        self.class_button.setEnabled(can_select)
        if has_data:
            self.class_button.setToolTip("选择班级")
        else:
            self.class_button.setToolTip("暂无学生数据，无法选择班级")

    def _ensure_student_data_ready(self) -> bool:
        """确保在需要访问学生数据前已经完成懒加载。"""

        if not self._student_data_pending_load:
            return True
        return self._load_student_data_if_needed()

    def show_class_selector(self) -> None:
        if self.mode != "roll_call":
            return
        if not self._ensure_student_data_ready():
            return
        workbook = self.student_workbook
        if workbook is None:
            show_quiet_information(self, "暂无学生数据，无法选择班级。")
            return
        class_names = workbook.class_names()
        if not class_names:
            show_quiet_information(self, "暂无班级可供选择。")
            return
        menu = QMenu(self)
        current = self.current_class_name or workbook.active_class
        for name in class_names:
            action = menu.addAction(name)
            action.setCheckable(True)
            action.setChecked(name == current)
            action.triggered.connect(lambda _checked=False, n=name: self._switch_class(n))
        pos = self.class_button.mapToGlobal(self.class_button.rect().bottomLeft())
        menu.exec(pos)

    def _switch_class(self, class_name: str) -> None:
        if self.student_workbook is None:
            return
        if class_name not in self.student_workbook.class_names():
            return
        target = class_name.strip()
        current = self.current_class_name or self.student_workbook.active_class
        if target == current:
            return
        if not self._ensure_student_data_ready():
            return
        self._snapshot_current_class()
        self.student_workbook.set_active_class(target)
        self.current_class_name = target
        if PANDAS_READY:
            df = self.student_workbook.get_active_dataframe()
        else:
            df = None
        self._set_student_dataframe(df, propagate=True)
        self._schedule_save()

    def _create_new_class(self) -> None:
        if not self._ensure_student_data_ready():
            return
        if self.student_workbook is None:
            self.student_workbook = StudentWorkbook(OrderedDict(), active_class="")
        if not PANDAS_READY:
            show_quiet_information(self, "当前环境缺少 pandas，无法创建班级。")
            return
        self._snapshot_current_class()
        suggested = f"班级{len(self.student_workbook.class_names()) + 1}" if self.student_workbook.class_names() else "班级1"
        name, ok = QInputDialog.getText(
            self,
            "新建班级",
            "请输入班级名称：",
            QLineEdit.EchoMode.Normal,
            suggested,
        )
        if not ok:
            return
        new_name = self.student_workbook.add_class(name)
        self.current_class_name = new_name
        self._apply_student_workbook(self.student_workbook, propagate=True)
        self._schedule_save()
        self._update_class_button_label()

    def _load_student_data_if_needed(self) -> bool:
        if not self._student_data_pending_load:
            return True
        if not (PANDAS_AVAILABLE and OPENPYXL_AVAILABLE):
            return False
        workbook = load_student_data(self)
        if workbook is None:
            return False
        self._student_data_pending_load = False
        self._apply_student_workbook(workbook, propagate=True)
        encrypted_state, encrypted_password = _get_session_student_encryption()
        self._student_file_encrypted = bool(encrypted_state)
        self._student_password = encrypted_password
        saved = self.settings_manager.load_settings().get("RollCallTimer", {})
        self._restore_group_state(saved)
        self._update_encryption_button()
        self._update_class_button_label()
        self.display_current_student()
        self._schedule_save()
        return True

    def _handle_encrypt_student_file(self) -> None:
        if not PANDAS_READY:
            show_quiet_information(self, "当前环境缺少 pandas，无法执行加密。")
            return
        password = self._prompt_new_encryption_password()
        if not password:
            return
        if not self._ensure_student_data_ready():
            return
        if self.student_workbook is None:
            if self.student_data is None or not isinstance(self.student_data, pd.DataFrame):
                show_quiet_information(self, "没有可加密的学生数据。")
                return
            try:
                snapshot = self.student_data.copy()
            except Exception:
                snapshot = pd.DataFrame(self.student_data)
            class_name = self.current_class_name or "班级1"
            self.current_class_name = class_name
            self.student_workbook = StudentWorkbook(
                OrderedDict({class_name: snapshot}),
                active_class=class_name,
            )
        else:
            self._snapshot_current_class()
        try:
            data = self.student_workbook.as_dict()
            _save_student_workbook(
                data,
                self.STUDENT_FILE,
                self._encrypted_file_path,
                encrypted=True,
                password=password,
            )
            _set_session_student_encryption(True, password)
            self._student_file_encrypted = True
            self._student_password = password
            self._update_encryption_button()
            self._propagate_student_dataframe()
            self._update_class_button_label()
            show_quiet_information(self, "已生成加密文件 students.xlsx.enc，并移除明文数据。")
            self._schedule_save()
        except Exception as exc:
            show_quiet_information(self, f"加密失败：{exc}")

    def _handle_decrypt_student_file(self) -> None:
        encrypted_path = self._encrypted_file_path
        if not os.path.exists(encrypted_path):
            show_quiet_information(self, "未找到加密文件，无法解密。")
            self._student_file_encrypted = False
            self._update_encryption_button()
            return
        password = self._prompt_existing_encryption_password("解密学生数据")
        if not password:
            return
        try:
            with open(encrypted_path, "rb") as fh:
                payload = fh.read()
            plain_bytes = _decrypt_student_bytes(password, payload)
        except Exception as exc:
            show_quiet_information(self, f"解密失败：{exc}")
            return
        try:
            buffer = io.BytesIO(plain_bytes)
            raw_data = pd.read_excel(buffer, sheet_name=None)
            workbook = StudentWorkbook(OrderedDict(raw_data), active_class="")
        except Exception as exc:
            show_quiet_information(self, f"读取解密后的学生数据失败：{exc}")
            return
        try:
            _save_student_workbook(
                workbook.as_dict(),
                self.STUDENT_FILE,
                self._encrypted_file_path,
                encrypted=False,
                password=None,
            )
        except Exception as exc:
            show_quiet_information(self, f"写入学生数据失败：{exc}")
            return
        self._student_file_encrypted = False
        self._student_password = None
        _set_session_student_encryption(False, None)
        self._apply_decrypted_student_data(workbook)
        self._update_encryption_button()
        show_quiet_information(self, "已成功解密学生数据并恢复 students.xlsx。")
        self._schedule_save()

    def _apply_decrypted_student_data(self, workbook: StudentWorkbook) -> None:
        if not PANDAS_READY:
            return
        self._apply_student_workbook(workbook, propagate=True)
        self.display_current_student()

    def _propagate_student_dataframe(self) -> None:
        parent = self.parent()
        if parent is None:
            return
        if hasattr(parent, "student_data"):
            try:
                setattr(parent, "student_data", self.student_data)
            except Exception:
                pass
        if hasattr(parent, "student_workbook"):
            try:
                setattr(parent, "student_workbook", self.student_workbook)
            except Exception:
                pass

    def _apply_saved_fonts(self) -> None:
        id_font = QFont("Microsoft YaHei UI", self.last_id_font_size, QFont.Weight.Bold)
        name_weight = QFont.Weight.Normal if self.name_font_family in {"楷体", "KaiTi"} else QFont.Weight.Bold
        name_font = QFont(self.name_font_family, self.last_name_font_size, name_weight)
        timer_font = QFont("Consolas", self.last_timer_font_size, QFont.Weight.Bold)
        self.id_label.setFont(id_font)
        self.name_label.setFont(name_font)
        self.time_display_label.setFont(timer_font)

    def _build_menu(self) -> QMenu:
        menu = QMenu(self)
        disp = menu.addMenu("显示选项")
        self.show_id_action = disp.addAction("显示学号"); self.show_id_action.setCheckable(True); self.show_id_action.setChecked(self.show_id)
        self.show_id_action.toggled.connect(self._on_display_option_changed)
        self.show_name_action = disp.addAction("显示姓名"); self.show_name_action.setCheckable(True); self.show_name_action.setChecked(self.show_name)
        self.show_name_action.toggled.connect(self._on_display_option_changed)
        self.show_photo_action = disp.addAction("显示照片"); self.show_photo_action.setCheckable(True); self.show_photo_action.setChecked(self.show_photo)
        self.show_photo_action.toggled.connect(self._on_display_option_changed)
        self.photo_duration_menu = disp.addMenu("照片显示时间")
        self.photo_duration_actions: List[QAction] = []
        duration_choices: List[Tuple[int, str]] = [
            (0, "不自动关闭"),
            (3, "3 秒"),
            (5, "5 秒"),
            (10, "10 秒"),
            (15, "15 秒"),
            (30, "30 秒"),
        ]
        current_seconds = int(self.photo_duration_seconds)
        for seconds, label in duration_choices:
            action = self.photo_duration_menu.addAction(label)
            action.setCheckable(True)
            action.setData(seconds)
            action.setChecked(seconds == current_seconds)
            action.triggered.connect(lambda _checked, s=seconds: self._set_photo_duration(s))
            self.photo_duration_actions.append(action)
        self._sync_photo_duration_actions()

        menu.addSeparator()
        speech = menu.addMenu("语音播报")
        manager = self.tts_manager
        checked = bool(self.speech_enabled and manager and manager.available)
        self.speech_enabled = checked
        self.speech_enabled_action = speech.addAction("启用语音播报"); self.speech_enabled_action.setCheckable(True)
        self.speech_enabled_action.setChecked(checked); self.speech_enabled_action.toggled.connect(self._toggle_speech)
        self.voice_menu = speech.addMenu("选择发音人"); self.voice_actions = []
        if manager and manager.available:
            self.speech_enabled_action.setEnabled(True)
            self.speech_enabled_action.setToolTip("点名时自动朗读当前学生姓名。")
            if manager.supports_voice_selection and manager.voice_ids:
                for vid in manager.voice_ids:
                    act = self.voice_menu.addAction(vid); act.setCheckable(True); act.setChecked(vid == manager.current_voice_id)
                    act.triggered.connect(lambda _c, v=vid: self._set_voice(v)); self.voice_actions.append(act)
            else:
                self.voice_menu.setEnabled(False)
                self.voice_menu.setToolTip("当前语音引擎不支持切换发音人。")
        else:
            self.voice_menu.setEnabled(False)
            self.speech_enabled_action.setEnabled(False)
            reason, suggestions = self._collect_speech_issue_details()
            tooltip_lines = [reason] if reason else []
            tooltip_lines.extend(suggestions)
            if tooltip_lines:
                self.speech_enabled_action.setToolTip("\n".join(tooltip_lines))

        menu.addSeparator()
        self.timer_sound_action = menu.addAction("倒计时结束提示音"); self.timer_sound_action.setCheckable(True)
        self.timer_sound_action.setChecked(self.timer_sound_enabled); self.timer_sound_action.toggled.connect(self._toggle_timer_sound)
        return menu

    def _update_menu_state(self) -> None:
        if self.show_id_action.isChecked() != self.show_id: self.show_id_action.setChecked(self.show_id)
        if self.show_name_action.isChecked() != self.show_name: self.show_name_action.setChecked(self.show_name)
        self.timer_sound_action.setChecked(self.timer_sound_enabled)
        manager = self.tts_manager
        if manager and manager.available:
            self.speech_enabled_action.setEnabled(True)
            self.speech_enabled_action.setChecked(self.speech_enabled)
            self.speech_enabled_action.setToolTip("点名时自动朗读当前学生姓名。")
            if manager.supports_voice_selection and manager.voice_ids:
                self.voice_menu.setEnabled(True)
                for act in self.voice_actions:
                    act.setChecked(act.text() == manager.current_voice_id)
                self.voice_menu.setToolTip("")
            else:
                self.voice_menu.setEnabled(False)
                self.voice_menu.setToolTip("当前语音引擎不支持切换发音人。")
        else:
            self.voice_menu.setEnabled(False)
            self.speech_enabled_action.setEnabled(False)
            self.speech_enabled_action.setChecked(False)
            reason, suggestions = self._collect_speech_issue_details()
            tooltip_lines = [reason] if reason else []
            tooltip_lines.extend(suggestions)
            if tooltip_lines:
                self.speech_enabled_action.setToolTip("\n".join(tooltip_lines))
            if not self._speech_issue_reported:
                self._diagnose_speech_engine()

    def _default_speech_suggestions(self) -> List[str]:
        hints: List[str] = []
        if sys.platform == "win32":
            hints.append("请确认 Windows 已启用 SAPI5 中文语音包。")
        elif sys.platform == "darwin":
            hints.append("请在系统“辅助功能 -> 语音”中启用所需的语音包。")
        else:
            hints.append("请确保系统已安装可用的语音引擎（如 espeak）并重新启动程序。")
        hints.append("可尝试重新安装 pyttsx3 或检查语音服务状态后重启软件。")
        return hints

    def _dedupe_suggestions(self, values: List[str]) -> List[str]:
        return dedupe_strings(values)

    def _collect_speech_issue_details(self) -> tuple[str, List[str]]:
        manager = self.tts_manager
        reason = ""
        suggestions: List[str] = []
        if manager is None:
            reason = "无法初始化系统语音引擎"
            suggestions = self._default_speech_suggestions()
        elif not manager.available:
            reason, suggestions = manager.diagnostics()
            reason = reason or "无法初始化系统语音引擎"
            if not suggestions:
                suggestions = self._default_speech_suggestions()
        elif manager.supports_voice_selection and not getattr(manager, "voice_ids", []):
            reason = "未检测到任何可用的发音人"
            suggestions = self._default_speech_suggestions()
            suggestions.append("请在操作系统语音设置中添加语音包后重新启动程序。")

        env_reason, env_suggestions = detect_speech_environment_issues(force_refresh=True)
        if env_reason:
            if not reason:
                reason = env_reason
            elif env_reason not in reason:
                reason = f"{reason}；{env_reason}"
        suggestions.extend(env_suggestions)
        if not suggestions and reason:
            suggestions = self._default_speech_suggestions()
        return reason, self._dedupe_suggestions(suggestions)

    def _ensure_speech_manager(self) -> Optional[TTSManager]:
        manager = self.tts_manager
        if manager and manager.available:
            return manager
        if manager is not None:
            try:
                manager.shutdown()
            except Exception:
                pass
        manager = TTSManager(self.selected_voice_id, parent=self)
        self.tts_manager = manager
        if manager.available:
            self._speech_issue_reported = False
            QTimer.singleShot(0, self._update_menu_state)
        return manager

    def _diagnose_speech_engine(self) -> None:
        if self._speech_issue_reported:
            return
        action = getattr(self, "speech_enabled_action", None)
        if action is None or action.isEnabled():
            return
        if not self.isVisible():
            if not self._speech_check_scheduled:
                self._speech_check_scheduled = True
                QTimer.singleShot(200, self._diagnose_speech_engine)
            return
        reason, suggestions = self._collect_speech_issue_details()
        if not reason:
            return
        advice = "\n".join(f"· {line}" for line in suggestions)
        message = f"语音播报功能当前不可用：{reason}"
        if advice:
            message = f"{message}\n{advice}"
        show_quiet_information(self, message, "语音播报提示")
        self._speech_issue_reported = True
        self._speech_check_scheduled = False

    def eventFilter(self, obj, e):
        if obj in (self.id_label, self.name_label) and e.type() == QEvent.Type.MouseButtonPress:
            if e.button() == Qt.MouseButton.LeftButton:
                self.roll_student(); return True
        return super().eventFilter(obj, e)

    def _on_display_option_changed(self) -> None:
        photo_checked = getattr(self, "show_photo_action", None)
        if (
            not self.show_id_action.isChecked()
            and not self.show_name_action.isChecked()
            and not (photo_checked.isChecked() if photo_checked else False)
        ):
            self.show_id_action.setChecked(True)
        self.show_id = self.show_id_action.isChecked()
        self.show_name = self.show_name_action.isChecked()
        if photo_checked is not None:
            self.show_photo = photo_checked.isChecked()
        if not self.show_photo:
            self._hide_student_photo(force=True)
        self.update_display_layout()
        self.display_current_student()
        self._schedule_save()

    def _set_photo_duration(self, seconds: int) -> None:
        seconds = max(0, int(seconds))
        if self.photo_duration_seconds == seconds:
            self._sync_photo_duration_actions()
            return
        self.photo_duration_seconds = seconds
        self._sync_photo_duration_actions()
        overlay = getattr(self, "_photo_overlay", None)
        if overlay is not None and overlay.isVisible():
            overlay.schedule_auto_close(int(self.photo_duration_seconds * 1000))
        self._schedule_save()

    def _sync_photo_duration_actions(self) -> None:
        current = int(self.photo_duration_seconds)
        for action in getattr(self, "photo_duration_actions", []):
            if not isinstance(action, QAction):
                continue
            data = action.data()
            try:
                value = int(data) if data is not None else 0
            except (TypeError, ValueError):
                value = 0
            block = action.blockSignals(True)
            action.setChecked(value == current)
            action.blockSignals(block)

    def _toggle_speech(self, enabled: bool) -> None:
        if not enabled:
            self.speech_enabled = False
            self._schedule_save()
            return
        manager = self._ensure_speech_manager()
        if not manager or not manager.available:
            reason, suggestions = self._collect_speech_issue_details()
            message = reason or "未检测到语音引擎，无法开启语音播报。"
            advice = "\n".join(f"· {line}" for line in suggestions)
            if advice:
                message = f"{message}\n{advice}"
            show_quiet_information(self, message, "语音播报提示")
            self.speech_enabled_action.setChecked(False)
            self._speech_issue_reported = True
            return
        if manager.supports_voice_selection and not getattr(manager, "voice_ids", []):
            reason, suggestions = self._collect_speech_issue_details()
            message = reason or "未检测到可用的发音人。"
            advice = "\n".join(f"· {line}" for line in suggestions)
            if advice:
                message = f"{message}\n{advice}"
            show_quiet_information(self, message, "语音播报提示")
            self.speech_enabled_action.setChecked(False)
            self._speech_issue_reported = True
            return
        self.speech_enabled = enabled
        self._schedule_save()

    def _set_voice(self, voice_id: str) -> None:
        manager = self._ensure_speech_manager()
        if not manager or not manager.supports_voice_selection:
            return
        manager.set_voice(voice_id); self.selected_voice_id = voice_id
        for a in self.voice_actions: a.setChecked(a.text() == voice_id)
        self._schedule_save()

    def _toggle_timer_sound(self, enabled: bool) -> None:
        self.timer_sound_enabled = enabled
        self._schedule_save()

    def _set_scoreboard_order(self, order: str) -> None:
        normalized = str(order).strip().lower()
        if normalized not in {"rank", "id"}:
            return
        if self.scoreboard_order == normalized:
            return
        self.scoreboard_order = normalized
        self._schedule_save()

    def _speak_text(self, text: str) -> None:
        if not text:
            return
        manager = self.tts_manager
        if not (self.speech_enabled and manager and manager.available):
            return
        manager.speak(text)

    def _announce_current_student(self) -> None:
        if (
            not self.speech_enabled
            or self.tts_manager is None
            or not self.tts_manager.available
            or self.current_student_index is None
            or self.student_data is None
            or self.student_data.empty
        ):
            return
        try:
            stu = self.student_data.loc[self.current_student_index]
        except Exception:
            return
        name_value = stu.get("姓名", "")
        if isinstance(name_value, str):
            name = name_value.strip()
        else:
            name = str(name_value).strip() if pd.notna(name_value) else ""
        if name:
            self._speak_text(name)

    def show_student_selector(self) -> None:
        if self.mode != "roll_call":
            return
        if self.student_data is None or self.student_data.empty:
            show_quiet_information(self, "暂无学生数据，无法显示名单。")
            return
        records: List[tuple[int, str, str, int]] = []
        for idx, row in self.student_data.iterrows():
            sid_value = row.get("学号", "")
            sid_display = re.sub(r"\s+", "", _normalize_text(sid_value))
            name = re.sub(r"\s+", "", _normalize_text(row.get("姓名", "")))
            try:
                sort_key = int(sid_display) if sid_display else sys.maxsize
            except (TypeError, ValueError):
                sort_key = sys.maxsize
            records.append((sort_key, sid_display, name, idx))
        if not records:
            show_quiet_information(self, "当前没有可显示的学生名单。")
            return
        records.sort(key=lambda item: (item[0], item[1]))
        dialog_data = []
        for _, sid, name, data_idx in records:
            display_sid = sid if sid else "无学号"
            display_name = name or "未命名"
            dialog_data.append((display_sid, display_name, data_idx))
        dialog = StudentListDialog(self, dialog_data)
        if dialog.exec() == QDialog.DialogCode.Accepted and dialog.selected_index is not None:
            selected = dialog.selected_index
            if selected in self.student_data.index:
                self.current_student_index = selected
                self._pending_passive_student = None
                self.display_current_student()
                self._announce_current_student()

    def increment_current_score(self) -> None:
        if self.mode != "roll_call":
            return
        if self.current_student_index is None:
            show_quiet_information(self, "请先选择需要加分的学生。")
            return
        if self.student_data is None:
            return
        if "成绩" not in self.student_data.columns:
            self.student_data["成绩"] = 0
        value = self.student_data.at[self.current_student_index, "成绩"]
        try:
            base = int(value)
        except (TypeError, ValueError):
            base = 0
        new_score = base + 1
        self.student_data.at[self.current_student_index, "成绩"] = new_score
        self._pending_passive_student = None
        self._update_score_display()
        self._persist_student_scores()
        self._speak_text("加分")

    def show_scoreboard(self) -> None:
        if self.mode != "roll_call":
            return
        if self.student_data is None or self.student_data.empty:
            show_quiet_information(self, "暂无学生数据，无法展示成绩。")
            return
        if "成绩" not in self.student_data.columns:
            self.student_data["成绩"] = 0
        records: List[tuple[str, str, int]] = []
        for _, row in self.student_data.iterrows():
            sid_value = row.get("学号", "")
            sid_display = re.sub(r"\s+", "", _normalize_text(sid_value))
            name = re.sub(r"\s+", "", _normalize_text(row.get("姓名", ""))) or "未命名"
            value = row.get("成绩", 0)
            try:
                score = int(value)
            except (TypeError, ValueError):
                score = 0
            records.append((sid_display, name, score))
        try:
            dialog = ScoreboardDialog(
                self,
                records,
                order=self.scoreboard_order,
                order_changed=self._set_scoreboard_order,
            )
        except Exception as exc:
            traceback.print_exc()
            show_quiet_information(self, f"打开成绩展示窗口时出错：{exc}")
            return
        dialog.exec()

    def _persist_student_scores(self) -> None:
        if not (PANDAS_AVAILABLE and OPENPYXL_AVAILABLE):
            return
        if not self._ensure_student_data_ready():
            return
        if self.student_workbook is None:
            if self.student_data is None or not isinstance(self.student_data, pd.DataFrame):
                return
            try:
                snapshot = self.student_data.copy()
            except Exception:
                snapshot = pd.DataFrame(self.student_data)
            class_name = self.current_class_name or "班级1"
            self.current_class_name = class_name
            self.student_workbook = StudentWorkbook(
                OrderedDict({class_name: snapshot}),
                active_class=class_name,
            )
        else:
            self._snapshot_current_class()
        if self.student_workbook is None:
            return
        try:
            with self._score_write_lock:
                data = self.student_workbook.as_dict()
                _save_student_workbook(
                    data,
                    self.STUDENT_FILE,
                    self._encrypted_file_path,
                    encrypted=self._student_file_encrypted,
                    password=self._student_password,
                )
            self._score_persist_failed = False
            self._update_class_button_label()
        except Exception as exc:
            if not self._score_persist_failed:
                show_quiet_information(self, f"保存成绩失败：{exc}")
                self._score_persist_failed = True

    def toggle_mode(self) -> None:
        self.mode = "timer" if self.mode == "roll_call" else "roll_call"
        if self.mode == "roll_call":
            self._placeholder_on_show = True
        self.update_mode_ui(force_timer_reset=self.mode == "timer")
        self._schedule_save()

    def update_mode_ui(self, force_timer_reset: bool = False) -> None:
        is_roll = self.mode == "roll_call"
        timer_reset_required = force_timer_reset
        if is_roll and not self._ensure_student_data_ready():
            self.mode = "timer"
            is_roll = False
            timer_reset_required = True
        self.title_label.setText("点名" if is_roll else "计时")
        self.mode_button.setText("切换到计时" if is_roll else "切换到点名")
        self.group_label.setVisible(is_roll)
        if hasattr(self, "group_container"):
            self.group_container.setVisible(is_roll)
        if hasattr(self, "group_bar"):
            self.group_bar.setVisible(is_roll)
        if hasattr(self, "add_score_button"):
            self.add_score_button.setVisible(is_roll)
        self._update_roll_call_controls()
        self._update_class_button_label()
        if is_roll:
            if self._placeholder_on_show:
                self.current_student_index = None
            self.stack.setCurrentWidget(self.roll_call_frame)
            self.count_timer.stop(); self.clock_timer.stop(); self.timer_running = False; self.timer_start_pause_button.setText("开始")
            self.update_display_layout(); self.display_current_student()
            self.schedule_font_update()
            self._placeholder_on_show = False
        else:
            self.stack.setCurrentWidget(self.timer_frame)
            changed = False
            if timer_reset_required:
                changed = self.reset_timer(persist=False)
            self.update_timer_mode_ui()
            if changed:
                self._schedule_save()
            self.schedule_font_update()
            self._hide_student_photo(force=True)
        if hasattr(self, "encrypt_button"):
            self.encrypt_button.setVisible(is_roll)
        if hasattr(self, "reset_button"):
            self.reset_button.setVisible(is_roll)
        self.updateGeometry()

    def _handle_timer_mode_transition(self, previous_mode: Optional[str], new_mode: str) -> None:
        if previous_mode == new_mode:
            return
        if new_mode in {"countdown", "stopwatch"}:
            self.timer_running = False
            self.count_timer.stop()
            self.timer_start_pause_button.setText("开始")
            if new_mode == "countdown":
                total = max(0, self.timer_countdown_minutes * 60 + self.timer_countdown_seconds)
                self.timer_seconds_left = total
            else:
                self.timer_stopwatch_seconds = 0

    def update_timer_mode_ui(self) -> None:
        mode = self.timer_modes[self.timer_mode_index]
        previous_mode = self._active_timer_mode
        if previous_mode is not None:
            self._handle_timer_mode_transition(previous_mode, mode)
        self._active_timer_mode = mode
        self.clock_timer.stop()
        if mode == "countdown":
            self.timer_mode_button.setText("倒计时")
            self.timer_start_pause_button.setEnabled(True); self.timer_reset_button.setEnabled(True); self.timer_set_button.setEnabled(True)
            self.timer_start_pause_button.setText("暂停" if self.timer_running else "开始")
            if self.timer_running and not self.count_timer.isActive(): self.count_timer.start()
            self.update_timer_display()
        elif mode == "stopwatch":
            self.timer_mode_button.setText("秒表")
            self.timer_start_pause_button.setEnabled(True); self.timer_reset_button.setEnabled(True); self.timer_set_button.setEnabled(False)
            self.timer_start_pause_button.setText("暂停" if self.timer_running else "开始")
            if self.timer_running and not self.count_timer.isActive(): self.count_timer.start()
            self.update_timer_display()
        else:
            self.timer_mode_button.setText("时钟")
            self.timer_start_pause_button.setEnabled(False); self.timer_reset_button.setEnabled(False); self.timer_set_button.setEnabled(False)
            self.timer_running = False; self.count_timer.stop(); self._update_clock(); self.clock_timer.start()
            self.timer_start_pause_button.setText("开始")
        self.schedule_font_update()

    def toggle_timer_mode(self) -> None:
        if self.timer_running: return
        self.timer_mode_index = (self.timer_mode_index + 1) % len(self.timer_modes); self.update_timer_mode_ui()
        self._schedule_save()

    def start_pause_timer(self) -> None:
        if self.timer_modes[self.timer_mode_index] == "clock": return
        self.timer_running = not self.timer_running
        if self.timer_running:
            self.timer_start_pause_button.setText("暂停")
            if not self.count_timer.isActive(): self.count_timer.start()
        else:
            self.timer_start_pause_button.setText("开始")
            self.count_timer.stop()
        self._schedule_save()

    def reset_timer(self, persist: bool = True) -> bool:
        changed = self.timer_running
        self.timer_running = False; self.count_timer.stop(); self.timer_start_pause_button.setText("开始")
        m = self.timer_modes[self.timer_mode_index]
        if m == "countdown":
            baseline = max(0, self.timer_countdown_minutes * 60 + self.timer_countdown_seconds)
            if self.timer_seconds_left != baseline:
                self.timer_seconds_left = baseline
                changed = True
        elif m == "stopwatch":
            if self.timer_stopwatch_seconds != 0:
                self.timer_stopwatch_seconds = 0
                changed = True
        self.update_timer_display()
        if persist and changed:
            self._schedule_save()
        return changed

    def set_countdown_time(self) -> None:
        d = CountdownSettingsDialog(self, self.timer_countdown_minutes, self.timer_countdown_seconds)
        if d.exec() and d.result:
            mi, se = d.result; self.timer_countdown_minutes = mi; self.timer_countdown_seconds = se
            changed = self.reset_timer()
            if not changed:
                self._schedule_save()

    def _on_count_timer(self) -> None:
        m = self.timer_modes[self.timer_mode_index]
        if m == "countdown":
            if self.timer_seconds_left > 0: self.timer_seconds_left -= 1
            else:
                self.count_timer.stop(); self.timer_running = False; self.timer_start_pause_button.setText("开始"); self.update_timer_display()
                if self.timer_sound_enabled: self.play_timer_sound()
                return
        elif m == "stopwatch":
            self.timer_stopwatch_seconds += 1
        self.update_timer_display()

    def update_timer_display(self) -> None:
        m = self.timer_modes[self.timer_mode_index]
        if m == "countdown": seconds = max(0, self.timer_seconds_left)
        elif m == "stopwatch": seconds = max(0, self.timer_stopwatch_seconds)
        else: seconds = 0
        if m in {"countdown", "stopwatch"}:
            mi, se = divmod(seconds, 60); self.time_display_label.setText(f"{int(mi):02d}:{int(se):02d}")
        else:
            self.time_display_label.setText(time.strftime("%H:%M:%S"))
        self.schedule_font_update()

    def _update_clock(self) -> None:
        self.time_display_label.setText(time.strftime("%H:%M:%S"))
        self.schedule_font_update()

    def play_timer_sound(self) -> None:
        if SOUNDDEVICE_AVAILABLE:
            def _play() -> None:
                try:
                    fs = 44100; duration = 0.5; frequency = 880
                    t = np.linspace(0, duration, int(fs * duration), endpoint=False)
                    data = 0.4 * np.sin(2 * np.pi * frequency * t)
                    sd.play(data.astype(np.float32), fs); sd.wait()
                except Exception:
                    pass
            threading.Thread(target=_play, daemon=True).start()

    def on_group_change(self, group_name: Optional[str] = None, initial: bool = False) -> None:
        if not self.groups:
            return
        if group_name is None:
            group_name = self.current_group_name
        if group_name not in self.groups:
            group_name = "全部" if "全部" in self.groups else self.groups[0]
        previous_group = self.current_group_name
        self.current_group_name = group_name
        self._update_group_button_state(group_name)
        if self.student_data.empty:
            self.current_student_index = None
            self.display_current_student()
            if not initial and previous_group != group_name:
                self._schedule_save()
            return
        self._pending_passive_student = None
        self._ensure_group_pool(group_name)
        self.current_student_index = None
        self.display_current_student()
        if not initial and previous_group != group_name:
            self._schedule_save()

    def roll_student(self, speak: bool = True) -> None:
        if self.mode != "roll_call": return
        group_name = self.current_group_name
        pool = self._group_remaining_indices.get(group_name)
        if pool is None:
            self._ensure_group_pool(group_name)
            pool = self._group_remaining_indices.get(group_name, [])
        if not pool:
            base_total = self._group_all_indices.get(group_name, [])
            if not base_total:
                show_quiet_information(self, f"'{group_name}' 分组当前没有可点名的学生。")
                self.current_student_index = None
                self.display_current_student()
                return
            if self._all_groups_completed():
                show_quiet_information(self, "所有学生都已完成点名，请点击“重置”按钮重新开始。")
            else:
                show_quiet_information(self, f"'{group_name}' 的同学已经全部点到，请切换其他分组或点击“重置”按钮。")
            return
        draw_index = self._rng.randrange(len(pool)) if len(pool) > 1 else 0
        self.current_student_index = pool.pop(draw_index)
        self._pending_passive_student = self.current_student_index
        self._group_last_student[group_name] = self.current_student_index
        self._mark_student_drawn(self.current_student_index)
        self.display_current_student()
        if speak:
            self._announce_current_student()
        # 即时同步保存配置，防止异常退出导致未点名名单丢失。
        self.save_settings()

    def _all_groups_completed(self) -> bool:
        """判断是否所有分组的学生都已点名完毕。"""

        total_students = len(self._group_all_indices.get("全部", []))
        if total_students == 0:
            return True
        if len(self._global_drawn_students) < total_students:
            return False
        for group, base in self._group_all_indices.items():
            if not base:
                continue
            remaining = self._group_remaining_indices.get(group, [])
            if remaining:
                return False
        return True

    def _reset_roll_call_state(self) -> None:
        """清空全部点名历史并重新洗牌。"""

        self.settings_manager.clear_roll_call_history()
        self._pending_passive_student = None
        self._rebuild_group_indices()
        self._ensure_group_pool(self.current_group_name)

    def _shuffle(self, values: List[int]) -> None:
        try:
            self._rng.shuffle(values)
        except Exception:
            random.shuffle(values)

    def _normalize_indices(self, values: Iterable[Any], *, allowed: Optional[Set[int]] = None) -> List[int]:
        """Convert an iterable of values to a deduplicated integer list."""

        normalized: List[int] = []
        seen: Set[int] = set()
        for value in values:
            try:
                idx = int(value)
            except (TypeError, ValueError):
                continue
            if allowed is not None and idx not in allowed:
                continue
            if idx in seen:
                continue
            normalized.append(idx)
            seen.add(idx)
        return normalized

    def _collect_base_indices(self, values: Optional[Iterable[Any]]) -> List[int]:
        """Normalize the raw index list preserved in each group."""

        if values is None:
            return []
        return self._normalize_indices(values)

    def reset_roll_call_pools(self) -> None:
        """根据当前分组执行重置：子分组独立重置，“全部”重置所有。"""

        if self.mode != "roll_call":
            return
        group_name = self.current_group_name
        if self.student_data is None or getattr(self.student_data, "empty", True):
            show_quiet_information(self, "暂无学生数据可供重置。")
            return
        if group_name == "全部":
            prompt = "确定要重置所有分组的点名状态并重新开始吗？"
        else:
            prompt = f"确定要重置“{group_name}”分组的点名状态并重新开始吗？"
        if not ask_quiet_confirmation(self, prompt, "确认重置"):
            return
        if group_name == "全部":
            self._reset_roll_call_state()
        else:
            self._reset_single_group(group_name)
        self.current_student_index = None
        self._pending_passive_student = None
        self.display_current_student()
        self.save_settings()

    def _reset_single_group(self, group_name: str) -> None:
        """仅重置指定分组，同时保持其它分组及全局状态不变。"""

        if group_name == "全部":
            return
        base_indices_raw = self._group_all_indices.get(group_name)
        if base_indices_raw is None:
            return
        base_indices = self._collect_base_indices(base_indices_raw)
        shuffled = list(base_indices)
        self._shuffle(shuffled)
        self._group_remaining_indices[group_name] = shuffled
        self._group_initial_sequences[group_name] = list(shuffled)
        self._group_last_student[group_name] = None
        if self._pending_passive_student in shuffled:
            self._pending_passive_student = None

        history = self._group_drawn_history.setdefault(group_name, set())
        if history:
            for idx in list(history):
                self._remove_from_global_history(idx, ignore_group=group_name)
            history.clear()

        last_all = self._group_last_student.get("全部")
        if last_all is not None:
            try:
                last_all_key = int(last_all)
            except (TypeError, ValueError):
                last_all_key = None
            if last_all_key is not None and last_all_key not in self._global_drawn_students:
                self._group_last_student["全部"] = None

        self._refresh_all_group_pool()

    def _rebuild_group_indices(self) -> None:
        """重新构建各分组的学生索引池。"""

        all_indices: Dict[str, List[int]] = {}
        remaining: Dict[str, List[int]] = {}
        last_student: Dict[str, Optional[int]] = {}
        student_groups: Dict[int, set[str]] = {}
        initial_sequences: Dict[str, List[int]] = {}

        if self.student_data.empty:
            all_indices["全部"] = []
        else:
            all_indices["全部"] = list(self.student_data.index)
            for idx in all_indices["全部"]:
                student_groups.setdefault(int(idx), set()).add("全部")
            group_series = self.student_data["分组"].astype(str).str.strip().str.upper()
            for group_name in self.groups:
                if group_name == "全部":
                    continue
                mask = group_series == group_name
                all_indices[group_name] = list(self.student_data[mask].index)
                for idx in all_indices[group_name]:
                    student_groups.setdefault(int(idx), set()).add(group_name)

        for group_name, indices in all_indices.items():
            pool = list(indices)
            self._shuffle(pool)
            remaining[group_name] = pool
            initial_sequences[group_name] = list(pool)
            last_student[group_name] = None

        self._group_all_indices = all_indices
        self._group_remaining_indices = remaining
        self._group_last_student = last_student
        self._group_initial_sequences = initial_sequences
        self._student_groups = student_groups
        self._group_drawn_history = {group: set() for group in all_indices}
        # “全部”分组直接引用全局集合，避免重复维护两份数据造成不一致
        if "全部" in self._group_drawn_history:
            self._global_drawn_students.clear()
            self._group_drawn_history["全部"] = self._global_drawn_students
        else:
            self._group_drawn_history["全部"] = self._global_drawn_students

        self._refresh_all_group_pool()

    def _remove_from_global_history(self, student_index: int, ignore_group: Optional[str] = None) -> None:
        """若学生未在其它分组被点名，则从全局记录中移除。"""

        try:
            student_key = int(student_index)
        except (TypeError, ValueError):
            return
        for group, history in self._group_drawn_history.items():
            if group == "全部" or group == ignore_group:
                continue
            if student_key in history:
                return
        self._global_drawn_students.discard(student_key)

    def _restore_group_state(self, section: Mapping[str, str]) -> None:
        """从配置中恢复各分组剩余学生池，保持未抽学生不重复。"""

        if not PANDAS_READY:
            return

        raw_states = section.get("class_states", "")
        restored_states: Dict[str, ClassRollState] = {}
        if raw_states:
            try:
                payload = json.loads(raw_states)
            except Exception:
                payload = {}
            if isinstance(payload, dict):
                for name, state_data in payload.items():
                    key = str(name).strip()
                    if not key:
                        continue
                    state = ClassRollState.from_mapping(state_data)
                    if state is not None:
                        restored_states[key] = state

        self._class_roll_states = restored_states
        self._prune_orphan_class_states()

        active_class = self._resolve_active_class_name()
        snapshot = self._class_roll_states.get(active_class)
        if snapshot is None:
            legacy = self._parse_legacy_roll_state(section)
            if legacy is not None and active_class:
                self._class_roll_states[active_class] = legacy
                snapshot = legacy

        if not self._can_apply_roll_state():
            return

        if snapshot is None:
            self._ensure_group_pool(self.current_group_name)
            return

        self._apply_roll_state(snapshot)
        sanitized = self._capture_roll_state()
        if sanitized is not None and active_class:
            self._class_roll_states[active_class] = sanitized

    def _ensure_group_pool(self, group_name: str, force_reset: bool = False) -> None:
        """确保指定分组仍有待抽取的学生，必要时重新洗牌。"""

        if group_name not in self._group_all_indices:
            if self.student_data.empty:
                base_list: List[int] = []
            elif group_name == "全部":
                base_list = list(self.student_data.index)
            else:
                group_series = self.student_data["分组"].astype(str).str.strip().str.upper()
                base_list = list(self.student_data[group_series == group_name].index)
            self._group_all_indices[group_name] = base_list
            self._group_remaining_indices[group_name] = []
            self._group_last_student.setdefault(group_name, None)
            if group_name == "全部":
                self._group_drawn_history[group_name] = self._global_drawn_students
            else:
                self._group_drawn_history.setdefault(group_name, set())
            for idx in base_list:
                entry = self._student_groups.setdefault(int(idx), set())
                entry.add(group_name)
                entry.add("全部")
            # 新增分组时同步生成初始顺序
            shuffled = list(base_list)
            self._shuffle(shuffled)
            self._group_initial_sequences[group_name] = shuffled

        base_indices = self._collect_base_indices(self._group_all_indices.get(group_name, []))
        if group_name == "全部":
            drawn_history = self._group_drawn_history.setdefault("全部", self._global_drawn_students)
            reference_drawn = self._global_drawn_students
        else:
            drawn_history = self._group_drawn_history.setdefault(group_name, set())
            reference_drawn = drawn_history

        if group_name == "全部":
            # “全部”分组直接依据全局集合生成剩余名单，避免与各子分组脱节
            if group_name not in self._group_initial_sequences:
                shuffled = list(base_indices)
                self._shuffle(shuffled)
                self._group_initial_sequences[group_name] = shuffled
            self._refresh_all_group_pool()
            self._group_last_student.setdefault(group_name, None)
            return

        if force_reset or group_name not in self._group_remaining_indices:
            drawn_history.clear()
            pool = list(base_indices)
            self._shuffle(pool)
            self._group_remaining_indices[group_name] = pool
            self._group_last_student.setdefault(group_name, None)
            self._group_initial_sequences[group_name] = list(pool)
            self._refresh_all_group_pool()
            return

        raw_pool = self._group_remaining_indices.get(group_name, [])
        normalized_pool: List[int] = []
        seen: set[int] = set()
        for value in raw_pool:
            try:
                idx = int(value)
            except (TypeError, ValueError):
                continue
            if idx in base_indices and idx not in seen and idx not in reference_drawn:
                normalized_pool.append(idx)
                seen.add(idx)

        source_order = self._group_initial_sequences.get(group_name)
        if source_order is None:
            # 如果未记录初始顺序，则退回数据原有顺序
            source_order = list(base_indices)
            self._group_initial_sequences[group_name] = list(source_order)

        additional: List[int] = []
        for idx in source_order:
            if idx in reference_drawn or idx in seen or idx not in base_indices:
                continue
            normalized_pool.append(idx)
            seen.add(idx)
        for idx in base_indices:
            if idx in reference_drawn or idx in seen:
                continue
            additional.append(idx)
            seen.add(idx)
        if additional:
            self._shuffle(additional)
            for value in additional:
                insert_at = self._rng.randrange(len(normalized_pool) + 1) if normalized_pool else 0
                normalized_pool.insert(insert_at, value)

        self._group_remaining_indices[group_name] = normalized_pool
        self._group_initial_sequences[group_name] = list(normalized_pool)
        self._group_last_student.setdefault(group_name, None)
        self._refresh_all_group_pool()

    def _mark_student_drawn(self, student_index: int) -> None:
        """抽中学生后，从所有关联分组的候选列表中移除该学生。"""

        student_key = None
        try:
            student_key = int(student_index)
        except (TypeError, ValueError):
            return

        groups = self._student_groups.get(student_key)
        if not groups:
            return
        self._global_drawn_students.add(student_key)
        for group in groups:
            if group == "全部":
                history = self._group_drawn_history.setdefault("全部", self._global_drawn_students)
            else:
                history = self._group_drawn_history.setdefault(group, set())
            history.add(student_key)
            pool = self._group_remaining_indices.get(group)
            if not pool:
                continue
            cleaned: List[int] = []
            for value in pool:
                try:
                    idx = int(value)
                except (TypeError, ValueError):
                    continue
                if idx != student_key:
                    cleaned.append(idx)
            self._group_remaining_indices[group] = cleaned

        self._refresh_all_group_pool()

    def _refresh_all_group_pool(self) -> None:
        """同步“全部”分组的剩余名单，使其与各子分组保持一致。"""

        base_all_list = self._collect_base_indices(self._group_all_indices.get("全部", []))
        base_all_set = set(base_all_list)

        subgroup_base: Dict[str, Tuple[List[int], Set[int]]] = {}
        subgroup_remaining: Dict[str, List[int]] = {}
        subgroup_remaining_union: Set[int] = set()
        drawn_from_subgroups: Set[int] = set()

        for group, raw_indices in self._group_all_indices.items():
            if group == "全部":
                continue
            base_list = self._collect_base_indices(raw_indices)
            base_set = set(base_list)
            subgroup_base[group] = (base_list, base_set)
            pool = self._group_remaining_indices.get(group, [])
            sanitized = self._normalize_indices(pool, allowed=base_set)
            if sanitized != pool:
                self._group_remaining_indices[group] = sanitized
            subgroup_remaining[group] = sanitized
            subgroup_remaining_union.update(sanitized)
            drawn_from_subgroups.update(idx for idx in base_set if idx not in sanitized)
            initial = self._group_initial_sequences.get(group)
            if initial is None:
                self._group_initial_sequences[group] = list(base_list)
            else:
                cleaned_initial = self._normalize_indices(initial, allowed=base_set)
                if cleaned_initial != list(initial):
                    self._group_initial_sequences[group] = cleaned_initial
                for idx in base_list:
                    if idx not in self._group_initial_sequences[group]:
                        self._group_initial_sequences[group].append(idx)

        valid_global = {
            idx
            for idx in self._global_drawn_students
            if idx in base_all_set and idx not in subgroup_remaining_union
        }
        new_global = {idx for idx in drawn_from_subgroups if idx in base_all_set}
        new_global.update(valid_global)

        self._global_drawn_students = set(new_global)
        self._group_drawn_history["全部"] = self._global_drawn_students

        for group, (base_list, base_set) in subgroup_base.items():
            pool = subgroup_remaining.get(group, [])
            filtered = [idx for idx in pool if idx in base_set and idx not in self._global_drawn_students]
            if filtered != pool:
                self._group_remaining_indices[group] = filtered
                subgroup_remaining[group] = filtered
            drawn_set = {idx for idx in base_set if idx not in filtered}
            self._group_drawn_history[group] = drawn_set

        order_hint = self._group_initial_sequences.get("全部")
        if order_hint is None:
            shuffled = list(base_all_list)
            self._shuffle(shuffled)
            order_hint = shuffled
        else:
            cleaned_all = self._normalize_indices(order_hint, allowed=base_all_set)
            if cleaned_all != list(order_hint):
                order_hint = cleaned_all
            else:
                order_hint = list(order_hint)
            for idx in base_all_list:
                if idx not in order_hint:
                    order_hint.append(idx)
        self._group_initial_sequences["全部"] = list(order_hint)

        normalized_all = [idx for idx in order_hint if idx not in self._global_drawn_students]
        seen_all: Set[int] = set(normalized_all)
        for idx in base_all_list:
            if idx in seen_all or idx in self._global_drawn_students:
                continue
            normalized_all.append(idx)
            seen_all.add(idx)
        self._group_remaining_indices["全部"] = normalized_all

    def display_current_student(self) -> None:
        photo_student_id: Optional[str] = None
        if self.current_student_index is None:
            self.id_label.setText("学号" if self.show_id else "")
            self.name_label.setText("学生" if self.show_name else "")
        else:
            stu = self.student_data.loc[self.current_student_index]
            raw_sid = stu.get("学号", "")
            raw_name = stu.get("姓名", "")
            sid = re.sub(r"\s+", "", _normalize_text(raw_sid))
            name = re.sub(r"\s+", "", _normalize_text(raw_name))
            self.id_label.setText(sid if self.show_id else ""); self.name_label.setText(name if self.show_name else "")
            if not self.show_id: self.id_label.setText("")
            if not self.show_name: self.name_label.setText("")
            photo_student_id = sid or None
        self.update_display_layout()
        self._update_score_display()
        self._update_roll_call_controls()
        self.schedule_font_update()
        self._maybe_show_student_photo(photo_student_id)

    def update_display_layout(self) -> None:
        self.id_label.setVisible(self.show_id); self.name_label.setVisible(self.show_name)
        layout: QGridLayout = self.roll_call_frame.layout()
        layout.setColumnStretch(0, 1); layout.setColumnStretch(1, 1)
        if not self.show_id: layout.setColumnStretch(0, 0)
        if not self.show_name: layout.setColumnStretch(1, 0)
        self.schedule_font_update()

    def _ensure_photo_root_directory(self) -> None:
        try:
            os.makedirs(self.photo_root_path, exist_ok=True)
        except Exception:
            logger.debug("Failed to create photo root directory at %s", self.photo_root_path, exc_info=True)

    def _maybe_show_student_photo(self, student_id: Optional[str]) -> None:
        if not self.show_photo or self.mode != "roll_call":
            self._hide_student_photo(force=True)
            return
        if not student_id:
            self._hide_student_photo()
            self._last_photo_student_id = None
            return
        normalized_id = student_id.strip()
        if not normalized_id:
            self._hide_student_photo()
            self._last_photo_student_id = None
            return
        if self._last_photo_student_id != normalized_id:
            self._photo_manual_hidden = False
        if self._photo_manual_hidden and self._last_photo_student_id == normalized_id:
            return
        path = self._resolve_student_photo_path(normalized_id)
        if not path:
            self._hide_student_photo()
            self._last_photo_student_id = normalized_id
            return
        pixmap = QPixmap(path)
        if pixmap.isNull():
            self._hide_student_photo()
            self._last_photo_student_id = normalized_id
            return
        overlay = self._ensure_photo_overlay()
        screen = (self.windowHandle().screen() if self.windowHandle() else None) or QApplication.primaryScreen()
        if screen is None:
            self._hide_student_photo()
            self._last_photo_student_id = normalized_id
            return
        screen_rect = screen.geometry()
        overlay.display_photo(pixmap, screen_rect, int(max(0, self.photo_duration_seconds) * 1000))
        self._last_photo_student_id = normalized_id
        self._photo_manual_hidden = False

    def _hide_student_photo(self, force: bool = False) -> None:
        overlay = getattr(self, "_photo_overlay", None)
        if overlay is not None:
            overlay.cancel_auto_close()
            if overlay.isVisible():
                overlay.hide()
        if force:
            self._photo_manual_hidden = False
            self._last_photo_student_id = None

    def _ensure_photo_overlay(self) -> StudentPhotoOverlay:
        if self._photo_overlay is None:
            self._photo_overlay = StudentPhotoOverlay(owner=self)
            self._photo_overlay.closed_by_user.connect(self._on_photo_overlay_closed)
            self._photo_overlay.auto_closed.connect(self._on_photo_overlay_auto_closed)
        else:
            self._photo_overlay.update_owner(self)
        return self._photo_overlay

    def _on_photo_overlay_closed(self) -> None:
        self._photo_manual_hidden = True
        overlay = getattr(self, "_photo_overlay", None)
        if overlay is not None:
            overlay.cancel_auto_close()

    def _on_photo_overlay_auto_closed(self) -> None:
        self._photo_manual_hidden = False

    def _resolve_student_photo_path(self, student_id: str) -> Optional[str]:
        class_name = self._sanitize_photo_segment(self._resolve_active_class_name())
        if not class_name:
            class_name = "default"
        base_dir = os.path.join(self.photo_root_path, class_name)
        try:
            os.makedirs(base_dir, exist_ok=True)
        except OSError:
            logger.debug("Unable to ensure photo directory %s", base_dir, exc_info=True)
        if not os.path.isdir(base_dir):
            return None
        for ext in self._photo_extensions:
            candidate = os.path.join(base_dir, f"{student_id}{ext}")
            if os.path.isfile(candidate):
                return candidate
            upper = os.path.join(base_dir, f"{student_id}{ext.upper()}")
            if os.path.isfile(upper):
                return upper
        lower_id = student_id.lower()
        try:
            for entry in os.listdir(base_dir):
                name, ext = os.path.splitext(entry)
                if not ext:
                    continue
                if ext.lower() not in self._photo_extensions:
                    continue
                if name.lower() == lower_id:
                    candidate = os.path.join(base_dir, entry)
                    if os.path.isfile(candidate):
                        return candidate
        except OSError:
            logger.debug("Failed to scan photo directory %s", base_dir, exc_info=True)
        return None

    @staticmethod
    def _sanitize_photo_segment(value: str) -> str:
        text = str(value or "").strip()
        if not text:
            return ""
        sanitized = re.sub(r"[^\w\u4e00-\u9fff-]+", "_", text)
        sanitized = sanitized.strip("_")
        return sanitized or "default"

    def _rebuild_group_buttons_ui(self) -> None:
        if not hasattr(self, "group_bar_layout"):
            return
        for button in list(self.group_button_group.buttons()):
            self.group_button_group.removeButton(button)
        while self.group_bar_layout.count():
            item = self.group_bar_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
        self.group_buttons = {}
        if not self.groups:
            return
        button_font = QFont("Microsoft YaHei UI", 9, QFont.Weight.Medium)
        button_height = recommended_control_height(button_font, extra=14, minimum=36)
        for group in self.groups:
            button = QPushButton(group)
            button.setCheckable(True)
            button.setFont(button_font)
            apply_button_style(button, ButtonStyles.TOOLBAR, height=button_height)
            button.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Fixed)
            button.setMinimumWidth(button.sizeHint().width())
            button.clicked.connect(lambda _checked=False, name=group: self.on_group_change(name))
            self.group_bar_layout.addWidget(button)
            self.group_button_group.addButton(button)
            self.group_buttons[group] = button
        self.group_bar_layout.addStretch(1)
        self._update_group_button_state(self.current_group_name)

    def _update_group_button_state(self, active_group: str) -> None:
        if not hasattr(self, "group_buttons"):
            return
        for name, button in self.group_buttons.items():
            block = button.blockSignals(True)
            button.setChecked(name == active_group)
            button.blockSignals(block)

    def _update_score_display(self) -> None:
        if not hasattr(self, "score_label"):
            return
        if (
            self.current_student_index is None
            or self.student_data is None
            or self.student_data.empty
            or "成绩" not in self.student_data.columns
        ):
            self.score_label.setText("成绩：--")
            return
        value = self.student_data.at[self.current_student_index, "成绩"]
        try:
            score = int(value)
        except (TypeError, ValueError):
            score = 0
        self.score_label.setText(f"成绩：{score}")

    def _update_roll_call_controls(self) -> None:
        if not all(hasattr(self, attr) for attr in ("list_button", "add_score_button", "showcase_button")):
            return
        is_roll = self.mode == "roll_call"
        self.list_button.setVisible(is_roll)
        self.add_score_button.setVisible(is_roll)
        self.showcase_button.setVisible(is_roll)
        self.score_label.setVisible(is_roll)
        has_student = self.current_student_index is not None
        has_data = self.student_data is not None and not getattr(self.student_data, "empty", True)
        self.add_score_button.setEnabled(is_roll and has_student)
        self.list_button.setEnabled(is_roll and has_data)
        self.showcase_button.setEnabled(is_roll and has_data)
        if hasattr(self, "class_button"):
            has_workbook = self.student_workbook is not None and not self.student_workbook.is_empty()
            can_select = is_roll and (has_workbook or self._student_data_pending_load)
            self.class_button.setVisible(is_roll)
            self.class_button.setEnabled(can_select)
        if hasattr(self, "encrypt_button"):
            self.encrypt_button.setVisible(is_roll)
            self.encrypt_button.setEnabled(is_roll and has_data)
        if hasattr(self, "reset_button"):
            self.reset_button.setEnabled(is_roll and has_data)

    def schedule_font_update(self) -> None:
        QTimer.singleShot(0, self.update_dynamic_fonts)

    def update_dynamic_fonts(self) -> None:
        name_font_size = self.last_name_font_size
        for lab in (self.id_label, self.name_label):
            if not lab.isVisible(): continue
            w = max(40, lab.width()); h = max(40, lab.height()); text = lab.text()
            size = self._calc_font_size(w, h, text)
            if lab is self.name_label:
                weight = QFont.Weight.Normal if self.name_font_family in {"楷体", "KaiTi"} else QFont.Weight.Bold
                lab.setFont(QFont(self.name_font_family, size, weight))
                self.last_name_font_size = size
                name_font_size = size
            else:
                lab.setFont(QFont("Microsoft YaHei UI", size, QFont.Weight.Bold))
                self.last_id_font_size = size
        if hasattr(self, "score_label") and self.score_label.isVisible():
            base = name_font_size if name_font_size > 0 else self.last_name_font_size
            if base <= 0:
                base = self.MIN_FONT_SIZE * 4
            score_size = max(1, int(round(base / 4)))
            self.score_label.setFont(QFont("Microsoft YaHei UI", score_size, QFont.Weight.Bold))
        if self.timer_frame.isVisible():
            text = self.time_display_label.text()
            w = max(60, self.time_display_label.width())
            h = max(60, self.time_display_label.height())
            size = self._calc_font_size(w, h, text, monospace=True)
            self.time_display_label.setFont(QFont("Consolas", size, QFont.Weight.Bold))
            self.last_timer_font_size = size

    def _calc_font_size(self, w: int, h: int, text: str, monospace: bool = False) -> int:
        if not text or w < 20 or h < 20:
            return self.MIN_FONT_SIZE
        w_eff = max(1, w - 16)
        h_eff = max(1, h - 16)
        is_cjk = any("\u4e00" <= c <= "\u9fff" for c in text)
        length = max(1, len(text))
        width_char_factor = 1.0 if is_cjk else (0.58 if monospace else 0.6)
        size_by_width = w_eff / (length * width_char_factor)
        size_by_height = h_eff * 0.70
        final_size = int(min(size_by_width, size_by_height))
        return max(self.MIN_FONT_SIZE, min(self.MAX_FONT_SIZE, final_size))

    def showEvent(self, e) -> None:
        super().showEvent(e)
        if self.mode == "roll_call" and self._placeholder_on_show:
            self.current_student_index = None
            self.display_current_student()
            self._placeholder_on_show = False
        elif self.mode == "timer":
            active_mode = self.timer_modes[self.timer_mode_index]
            if active_mode in {"countdown", "stopwatch"}:
                if self.reset_timer(persist=False):
                    self._schedule_save()
                self.update_timer_mode_ui()
        self.visibility_changed.emit(True)
        self.schedule_font_update()
        ensure_widget_within_screen(self)
        if not self._speech_issue_reported:
            self._diagnose_speech_engine()

    def resizeEvent(self, e: QResizeEvent) -> None:
        super().resizeEvent(e)
        self.schedule_font_update()

    def hideEvent(self, e) -> None:
        self._hide_student_photo(force=True)
        super().hideEvent(e)
        self._placeholder_on_show = True
        self.save_settings()
        self.visibility_changed.emit(False)

    def closeEvent(self, e) -> None:
        self._hide_student_photo(force=True)
        self.save_settings()
        self.count_timer.stop()
        self.clock_timer.stop()
        if self.tts_manager: self.tts_manager.shutdown()
        self.window_closed.emit()
        super().closeEvent(e)

    def _schedule_save(self) -> None:
        """延迟写入设置，避免频繁保存导致的磁盘抖动。"""

        if self._save_timer.isActive():
            self._save_timer.stop()
        self._save_timer.start()

    def save_settings(self) -> None:
        if self._save_timer.isActive():
            self._save_timer.stop()
        settings = self.settings_manager.load_settings()
        sec = settings.get("RollCallTimer", {})
        sec["geometry"] = geometry_to_text(self)
        sec["show_id"] = bool_to_str(self.show_id)
        sec["show_name"] = bool_to_str(self.show_name)
        sec["show_photo"] = bool_to_str(self.show_photo)
        sec["photo_duration_seconds"] = str(int(self.photo_duration_seconds))
        sec["speech_enabled"] = bool_to_str(self.speech_enabled)
        sec["speech_voice_id"] = self.selected_voice_id
        sec["current_class"] = self.current_class_name
        sec["current_group"] = self.current_group_name
        sec["timer_countdown_minutes"] = str(self.timer_countdown_minutes)
        sec["timer_countdown_seconds"] = str(self.timer_countdown_seconds)
        sec["timer_sound_enabled"] = bool_to_str(self.timer_sound_enabled)
        sec["mode"] = self.mode
        sec["timer_mode"] = self.timer_modes[self.timer_mode_index]
        sec["timer_seconds_left"] = str(self.timer_seconds_left)
        sec["timer_stopwatch_seconds"] = str(self.timer_stopwatch_seconds)
        sec["timer_running"] = bool_to_str(self.timer_running)
        sec["id_font_size"] = str(self.last_id_font_size)
        sec["name_font_size"] = str(self.last_name_font_size)
        sec["timer_font_size"] = str(self.last_timer_font_size)
        sec["scoreboard_order"] = self.scoreboard_order
        sec["students_encrypted"] = bool_to_str(self._student_file_encrypted)
        self._prune_orphan_class_states()
        if not self._student_data_pending_load:
            self._store_active_class_state()
        sec["class_states"] = self._encode_class_states()
        if self._student_data_pending_load:
            # 在尚未加载真实名单数据时，保留磁盘上已有的未点名状态，避免误把占位空列表写回
            # 此时直接返回，保持上一轮保存的名单信息不被覆盖。
            settings["RollCallTimer"] = sec
            self.settings_manager.save_settings(settings)
            return

        # 名单已经加载完成，正常序列化各分组的剩余名单及历史记录
        remaining_payload: Dict[str, List[int]] = {}
        for group, indices in self._group_remaining_indices.items():
            cleaned: List[int] = []
            for idx in indices:
                try:
                    cleaned.append(int(idx))
                except (TypeError, ValueError):
                    continue
            remaining_payload[group] = cleaned
        last_payload: Dict[str, Optional[int]] = {}
        all_groups = set(self._group_all_indices.keys()) | set(self._group_last_student.keys())
        for group in all_groups:
            value = self._group_last_student.get(group)
            if value is None:
                last_payload[group] = None
            else:
                try:
                    last_payload[group] = int(value)
                except (TypeError, ValueError):
                    last_payload[group] = None
        try:
            sec["group_remaining"] = json.dumps(remaining_payload, ensure_ascii=False)
        except TypeError:
            sec["group_remaining"] = json.dumps({}, ensure_ascii=False)
        try:
            sec["group_last"] = json.dumps(last_payload, ensure_ascii=False)
        except TypeError:
            sec["group_last"] = json.dumps({}, ensure_ascii=False)
        try:
            global_drawn_payload = sorted(int(idx) for idx in self._global_drawn_students)
            sec["global_drawn"] = json.dumps(global_drawn_payload, ensure_ascii=False)
        except TypeError:
            sec["global_drawn"] = json.dumps([], ensure_ascii=False)
        settings["RollCallTimer"] = sec
        self.settings_manager.save_settings(settings)


# ---------- 关于 ----------
class AboutDialog(QDialog):
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("关于")
        self.setWindowFlag(Qt.WindowType.WindowStaysOnTopHint, True)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(12)
        info = QLabel(
            "<b>ClassroomTools</b><br>"
            "作者：广州番禺王耀强<br>"
            "知乎主页：<a href='https://www.zhihu.com/people/sciman/columns'>sciman</a><br>"
            "公众号：sciman逸居<br>"
            "“初中物理教研”Q群：323728546"
        )
        info.setOpenExternalLinks(True)
        info.setTextFormat(Qt.TextFormat.RichText)
        layout.addWidget(info)
        btn = QPushButton("确定")
        btn.clicked.connect(self.accept)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        layout.addWidget(btn, alignment=Qt.AlignmentFlag.AlignCenter)
        self.setFixedSize(self.sizeHint())

    def showEvent(self, event) -> None:  # type: ignore[override]
        super().showEvent(event)
        ensure_widget_within_screen(self)


# ---------- 数据 ----------
_ENCRYPTED_MAGIC = b"CTS1"


def _derive_stream_keys(password: str, salt: bytes) -> tuple[bytes, bytes]:
    material = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 120000, dklen=64)
    return material[:32], material[32:]


def _generate_keystream(stream_key: bytes, length: int) -> bytes:
    output = bytearray()
    counter = 0
    while len(output) < length:
        block = hashlib.sha256(stream_key + counter.to_bytes(8, "big")).digest()
        output.extend(block)
        counter += 1
    return bytes(output[:length])


def _encrypt_student_bytes(password: str, data: bytes) -> bytes:
    if not password:
        raise ValueError("缺少加密密码")
    salt = os.urandom(16)
    stream_key, mac_key = _derive_stream_keys(password, salt)
    keystream = _generate_keystream(stream_key, len(data))
    cipher = bytes(b ^ k for b, k in zip(data, keystream))
    tag = hmac.new(mac_key, cipher, hashlib.sha256).digest()
    return _ENCRYPTED_MAGIC + salt + tag + cipher


def _decrypt_student_bytes(password: str, blob: bytes) -> bytes:
    if not blob.startswith(_ENCRYPTED_MAGIC) or len(blob) <= len(_ENCRYPTED_MAGIC) + 48:
        raise ValueError("文件格式无效")
    salt = blob[len(_ENCRYPTED_MAGIC): len(_ENCRYPTED_MAGIC) + 16]
    tag = blob[len(_ENCRYPTED_MAGIC) + 16: len(_ENCRYPTED_MAGIC) + 48]
    cipher = blob[len(_ENCRYPTED_MAGIC) + 48:]
    stream_key, mac_key = _derive_stream_keys(password, salt)
    expected_tag = hmac.new(mac_key, cipher, hashlib.sha256).digest()
    if not hmac.compare_digest(expected_tag, tag):
        raise ValueError("密码错误或文件已损坏")
    keystream = _generate_keystream(stream_key, len(cipher))
    return bytes(b ^ k for b, k in zip(cipher, keystream))


def _normalize_text(value: object) -> str:
    if PANDAS_READY:
        if pd.isna(value):
            return ""
    else:
        if value is None:
            return ""
        if isinstance(value, float) and math.isnan(value):
            return ""
        if isinstance(value, str) and not value:
            return ""
    text = str(value).strip()
    lowered = text.lower()
    if lowered in {"nan", "none", "nat"}:
        return ""
    return text


def _normalize_student_dataframe(
    df: PandasDataFrame,
    *,
    drop_incomplete: bool = True,
) -> PandasDataFrame:
    if not PANDAS_READY:
        return df.copy()

    normalized = df.copy()
    for column in ("学号", "姓名", "分组", "成绩"):
        if column not in normalized.columns:
            normalized[column] = pd.NA

    normalized["姓名"] = normalized["姓名"].apply(lambda v: re.sub(r"\s+", "", _normalize_text(v)))
    normalized["分组"] = normalized["分组"].apply(lambda v: re.sub(r"\s+", "", _normalize_text(v))).str.upper()

    id_series = normalized["学号"].apply(lambda v: re.sub(r"\s+", "", _normalize_text(v)))
    id_numeric = pd.to_numeric(id_series.replace("", pd.NA), errors="coerce")
    if not id_numeric.empty:
        fractional_mask = id_numeric.notna() & (id_numeric != id_numeric.round())
        id_numeric = id_numeric.where(~fractional_mask)
    normalized["学号"] = id_numeric.round().astype("Int64")

    score_series = normalized["成绩"].apply(lambda v: re.sub(r"\s+", "", _normalize_text(v)))
    score_numeric = pd.to_numeric(score_series.replace("", pd.NA), errors="coerce").fillna(0)
    score_numeric = score_numeric.round()
    normalized["成绩"] = score_numeric.astype("Int64")

    for column in normalized.select_dtypes(include=["object"]).columns:
        if column in {"姓名", "分组"}:
            continue
        normalized[column] = normalized[column].apply(_normalize_text)

    ordered_columns = [col for col in ["学号", "姓名", "分组", "成绩"] if col in normalized.columns]
    extra_columns = [col for col in normalized.columns if col not in ordered_columns]
    normalized = normalized[ordered_columns + extra_columns]

    if drop_incomplete:
        normalized = normalized[(normalized["学号"].notna()) & (normalized["姓名"] != "")].copy()
        normalized.reset_index(drop=True, inplace=True)

    return normalized


def _empty_student_dataframe() -> PandasDataFrame:
    if not PANDAS_READY:
        raise RuntimeError("Pandas support is required to create student data tables.")
    template = pd.DataFrame({"学号": [], "姓名": [], "分组": [], "成绩": []})
    return _normalize_student_dataframe(template, drop_incomplete=False)


def _sanitize_sheet_name(name: str, fallback: str) -> str:
    invalid = set("\\/:?*[]")
    cleaned = "".join(ch for ch in str(name) if ch not in invalid).strip()
    if not cleaned:
        cleaned = fallback
    if len(cleaned) > 31:
        cleaned = cleaned[:31]
    return cleaned


@dataclass
class StudentWorkbook:
    """封装多班级学生名单，允许按工作表划分班级。"""

    sheets: "OrderedDict[str, PandasDataFrame]"
    active_class: str = ""

    def __post_init__(self) -> None:
        ordered: "OrderedDict[str, PandasDataFrame]" = OrderedDict()
        if self.sheets:
            for idx, (name, df) in enumerate(self.sheets.items(), start=1):
                fallback = f"班级{idx}" if idx > 1 else "班级1"
                safe_name = _sanitize_sheet_name(name, fallback)
                try:
                    normalized = _normalize_student_dataframe(df, drop_incomplete=False)
                except Exception:
                    normalized = pd.DataFrame(df)
                ordered[safe_name] = normalized
        if not ordered:
            ordered["班级1"] = _empty_student_dataframe().copy()
        self.sheets = ordered
        if not self.active_class or self.active_class not in self.sheets:
            self.active_class = next(iter(self.sheets))

    def class_names(self) -> List[str]:
        return list(self.sheets.keys())

    def is_empty(self) -> bool:
        if not self.sheets:
            return True
        for df in self.sheets.values():
            try:
                if not df.empty:
                    return False
            except AttributeError:
                return False
        return True

    def get_active_dataframe(self) -> PandasDataFrame:
        if not self.sheets:
            return _empty_student_dataframe().copy()
        if self.active_class not in self.sheets:
            self.active_class = next(iter(self.sheets))
        df = self.sheets.get(self.active_class)
        if df is None:
            return _empty_student_dataframe().copy()
        try:
            return df.copy()
        except Exception:
            return pd.DataFrame(df)

    def set_active_class(self, class_name: str) -> None:
        name = str(class_name).strip()
        if name in self.sheets:
            self.active_class = name

    def update_class(self, class_name: str, df: PandasDataFrame) -> None:
        try:
            normalized = _normalize_student_dataframe(df, drop_incomplete=False)
        except Exception:
            normalized = pd.DataFrame(df)
        self.sheets[class_name] = normalized
        self.active_class = class_name

    def add_class(self, class_name: str) -> str:
        base_name = str(class_name).strip()
        if not base_name:
            base_name = f"班级{len(self.sheets) + 1}"
        safe_name = _sanitize_sheet_name(base_name, base_name)
        if safe_name in self.sheets:
            suffix = 2
            while f"{safe_name}_{suffix}" in self.sheets:
                suffix += 1
            safe_name = f"{safe_name}_{suffix}"
        self.sheets[safe_name] = _empty_student_dataframe().copy()
        self.active_class = safe_name
        return safe_name

    def as_dict(self) -> "OrderedDict[str, PandasDataFrame]":
        ordered: "OrderedDict[str, PandasDataFrame]" = OrderedDict()
        for name, df in self.sheets.items():
            try:
                ordered[name] = df.copy()
            except Exception:
                ordered[name] = pd.DataFrame(df)
        return ordered


def _write_student_workbook(file_path: str, data: Mapping[str, PandasDataFrame]) -> None:
    payload = _export_student_workbook_bytes(data)
    tmp_dir = os.path.dirname(os.path.abspath(file_path)) or "."
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=tmp_dir)
    try:
        with os.fdopen(fd, "wb") as tmp_file:
            tmp_file.write(payload)
        os.replace(tmp_path, file_path)
    finally:
        with contextlib.suppress(FileNotFoundError):
            os.remove(tmp_path)


def _write_encrypted_student_workbook(file_path: str, data: Mapping[str, PandasDataFrame], password: str) -> None:
    payload = _export_student_workbook_bytes(data)
    encrypted = _encrypt_student_bytes(password, payload)
    tmp_dir = os.path.dirname(os.path.abspath(file_path)) or "."
    fd, tmp_path = tempfile.mkstemp(suffix=".enc", dir=tmp_dir)
    try:
        with os.fdopen(fd, "wb") as tmp_file:
            tmp_file.write(encrypted)
        os.replace(tmp_path, file_path)
    finally:
        with contextlib.suppress(FileNotFoundError):
            os.remove(tmp_path)


def _export_student_workbook_bytes(data: Mapping[str, PandasDataFrame]) -> bytes:
    normalized: "OrderedDict[str, PandasDataFrame]" = OrderedDict()
    for idx, (name, df) in enumerate(data.items(), start=1):
        fallback = f"班级{idx}" if idx > 1 else "班级1"
        sheet_name = _sanitize_sheet_name(name, fallback)
        try:
            normalized_df = _normalize_student_dataframe(df, drop_incomplete=False)
        except Exception:
            normalized_df = pd.DataFrame(df)
        normalized[sheet_name] = normalized_df
    if not normalized:
        normalized["班级1"] = _empty_student_dataframe().copy()

    buffer = io.BytesIO()
    try:
        engine = "openpyxl" if OPENPYXL_AVAILABLE else None
        with pd.ExcelWriter(buffer, engine=engine) as writer:  # type: ignore[call-arg]
            for sheet_name, df in normalized.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        buffer.seek(0)
        return buffer.getvalue()
    except Exception:
        buffer.seek(0)
        first = next(iter(normalized.values()))
        first.to_excel(buffer, index=False)
        return buffer.getvalue()


def _save_student_workbook(
    data: Mapping[str, PandasDataFrame],
    file_path: str,
    encrypted_file_path: str,
    *,
    encrypted: bool,
    password: Optional[str],
) -> None:
    if encrypted:
        if not password:
            raise ValueError("缺少加密密码")
        _write_encrypted_student_workbook(encrypted_file_path, data, password)
        with contextlib.suppress(FileNotFoundError):
            os.remove(file_path)
    else:
        _write_student_workbook(file_path, data)
        with contextlib.suppress(FileNotFoundError):
            os.remove(encrypted_file_path)


def load_student_data(parent: Optional[QWidget]) -> Optional[StudentWorkbook]:
    """从 students.xlsx 读取点名所需的数据，不存在时自动生成模板。"""

    if not (PANDAS_AVAILABLE and OPENPYXL_AVAILABLE):
        QMessageBox.warning(parent, "提示", "未安装 pandas/openpyxl，点名功能不可用。")
        return None

    file_path = RollCallTimerWindow.STUDENT_FILE
    encrypted_path = getattr(RollCallTimerWindow, "ENCRYPTED_STUDENT_FILE", file_path + ".enc")

    if not os.path.exists(file_path) and os.path.exists(encrypted_path):
        attempts = 0
        while attempts < 3:
            password, ok = PasswordPromptDialog.get_password(
                parent,
                "解密学生数据",
                "检测到已加密的学生名单，请输入密码：",
                allow_empty=False,
            )
            if not ok:
                QMessageBox.information(parent, "提示", "已取消加载加密的学生名单。")
                return None
            password = password.strip()
            if not password:
                QMessageBox.warning(parent, "提示", "密码不能为空，请重新输入。")
                attempts += 1
                continue
            try:
                with open(encrypted_path, "rb") as fh:
                    payload = fh.read()
                plain_bytes = _decrypt_student_bytes(password, payload)
                buffer = io.BytesIO(plain_bytes)
                raw_data = pd.read_excel(buffer, sheet_name=None)
                workbook = StudentWorkbook(OrderedDict(raw_data), active_class="")
                _set_session_student_encryption(True, password)
                return workbook
            except Exception as exc:
                attempts += 1
                QMessageBox.warning(parent, "提示", f"解密失败：{exc}")
        QMessageBox.critical(parent, "错误", "多次输入密码失败，无法加载学生名单。")
        return None

    if not os.path.exists(file_path):
        try:
            template = pd.DataFrame(
                {"学号": [101, 102, 103], "姓名": ["张三", "李四", "王五"], "分组": ["A", "B", "A"], "成绩": [0, 0, 0]}
            )
            workbook = StudentWorkbook(OrderedDict({"班级1": template}), active_class="班级1")
            _write_student_workbook(file_path, workbook.as_dict())
            show_quiet_information(parent, f"未找到学生名单，已为您创建模板文件：{file_path}")
            _set_session_student_encryption(False, None)
        except Exception as exc:
            QMessageBox.critical(parent, "错误", f"创建模板文件失败：{exc}")
            return None

    try:
        raw_data = pd.read_excel(file_path, sheet_name=None)
        workbook = StudentWorkbook(OrderedDict(raw_data), active_class="")
        _write_student_workbook(file_path, workbook.as_dict())
        _set_session_student_encryption(False, None)
        if os.path.exists(encrypted_path):
            show_quiet_information(parent, "检测到同时存在加密文件，将优先使用明文 students.xlsx。")
        return workbook
    except Exception as exc:
        QMessageBox.critical(parent, "错误", f"无法加载学生名单，请检查文件格式。\n错误：{exc}")
        return None


# ---------- 启动器 ----------
class LauncherBubble(QWidget):
    """启动器缩小时显示的悬浮圆球，负责发出恢复指令。"""

    restore_requested = pyqtSignal()
    position_changed = pyqtSignal(QPoint)

    def __init__(self, diameter: int = 42) -> None:
        super().__init__(
            None,
            Qt.WindowType.Tool
            | Qt.WindowType.FramelessWindowHint
            | Qt.WindowType.WindowStaysOnTopHint,
        )
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        self.setAttribute(Qt.WidgetAttribute.WA_NoSystemBackground, True)
        self.setAttribute(Qt.WidgetAttribute.WA_ShowWithoutActivating, True)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setWindowTitle("ClassroomTools Bubble")
        self._diameter = max(32, diameter)
        self.setFixedSize(self._diameter, self._diameter)
        self.setWindowOpacity(0.74)
        self._ensure_min_width = self._diameter
        self._ensure_min_height = self._diameter
        self._dragging = False
        self._drag_offset = QPoint()
        self._moved = False

    def place_near(self, target: QPoint, screen: Optional[QScreen]) -> None:
        """将气泡吸附到距离 target 最近的屏幕边缘。"""

        if screen is None:
            screen = QApplication.screenAt(target) or QApplication.primaryScreen()
        if screen is None:
            return
        available = screen.availableGeometry()
        margin = 6
        bubble_size = self.size()
        center = QPoint(int(target.x()), int(target.y()))
        center.setX(max(available.left(), min(center.x(), available.right())))
        center.setY(max(available.top(), min(center.y(), available.bottom())))

        distances = {
            "left": abs(center.x() - available.left()),
            "right": abs(available.right() - center.x()),
            "top": abs(center.y() - available.top()),
            "bottom": abs(available.bottom() - center.y()),
        }
        nearest_edge = min(distances, key=distances.get)
        if nearest_edge == "left":
            x = available.left() + margin
            y = center.y() - bubble_size.height() // 2
        elif nearest_edge == "right":
            x = available.right() - bubble_size.width() - margin
            y = center.y() - bubble_size.height() // 2
        elif nearest_edge == "top":
            y = available.top() + margin
            x = center.x() - bubble_size.width() // 2
        else:
            y = available.bottom() - bubble_size.height() - margin
            x = center.x() - bubble_size.width() // 2

        x = max(available.left() + margin, min(x, available.right() - bubble_size.width() - margin))
        y = max(available.top() + margin, min(y, available.bottom() - bubble_size.height() - margin))
        self.move(int(x), int(y))
        self.position_changed.emit(self.pos())

    def snap_to_edge(self) -> None:
        screen = self.screen() or QApplication.screenAt(self.frameGeometry().center()) or QApplication.primaryScreen()
        if screen is None:
            return
        self.place_near(self.frameGeometry().center(), screen)

    def paintEvent(self, event) -> None:
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        rect = self.rect()
        color = QColor(32, 33, 36, 150)
        highlight = QColor(138, 180, 248, 160)
        painter.setBrush(QBrush(color))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawEllipse(rect)
        painter.setBrush(QBrush(highlight))
        painter.drawEllipse(rect.adjusted(rect.width() // 3, rect.height() // 3, -rect.width() // 6, -rect.height() // 6))

    def mousePressEvent(self, event) -> None:
        if event.button() == Qt.MouseButton.LeftButton:
            self._dragging = True
            self._drag_offset = event.position().toPoint()
            self._moved = False
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event) -> None:
        if self._dragging and event.buttons() & Qt.MouseButton.LeftButton:
            new_pos = event.globalPosition().toPoint() - self._drag_offset
            self.move(new_pos)
            self._moved = True
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event) -> None:
        if event.button() == Qt.MouseButton.LeftButton:
            if not self._moved:
                self.restore_requested.emit()
            else:
                self.snap_to_edge()
                self.position_changed.emit(self.pos())
            self._dragging = False
        super().mouseReleaseEvent(event)

    def showEvent(self, event) -> None:  # type: ignore[override]
        super().showEvent(event)
        ensure_widget_within_screen(self)
        self.raise_()
        QTimer.singleShot(0, self.raise_)


class LauncherWindow(QWidget):
    def __init__(self, settings_manager: SettingsManager, student_workbook: Optional[StudentWorkbook]) -> None:
        super().__init__(None, Qt.WindowType.Tool | Qt.WindowType.FramelessWindowHint | Qt.WindowType.WindowStaysOnTopHint)
        self.settings_manager = settings_manager
        self.student_workbook: Optional[StudentWorkbook] = student_workbook
        self.student_data: Optional[PandasDataFrame] = None
        if PANDAS_READY and student_workbook is not None:
            try:
                self.student_data = student_workbook.get_active_dataframe()
            except Exception:
                self.student_data = pd.DataFrame(columns=["学号", "姓名", "分组", "成绩"])
        self.overlay: Optional[OverlayWindow] = None
        self.roll_call_window: Optional[RollCallTimerWindow] = None
        self._dragging = False; self._drag_offset = QPoint()
        self.bubble: Optional[LauncherBubble] = None
        self._last_position = QPoint()
        self._bubble_position = QPoint()
        self._minimized = False
        self._minimized_on_start = False

        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.setStyleSheet(
            """
            QWidget#launcherContainer {
                background-color: rgba(28, 29, 32, 232);
                border-radius: 9px;
                border: 1px solid rgba(255, 255, 255, 40);
            }
            QPushButton {
                color: #f1f3f4;
                background-color: rgba(60, 64, 67, 230);
                border: 1px solid rgba(255, 255, 255, 35);
                border-radius: 6px;
                padding: 3px 9px;
                min-height: 26px;
            }
            QPushButton:hover {
                background-color: rgba(138, 180, 248, 240);
                border-color: rgba(138, 180, 248, 255);
                color: #202124;
            }
            QCheckBox {
                color: #f1f3f4;
                spacing: 4px;
            }
            QCheckBox::indicator {
                width: 14px;
                height: 14px;
                border-radius: 4px;
                border: 1px solid rgba(255, 255, 255, 60);
                background: rgba(32, 33, 36, 240);
            }
            QCheckBox::indicator:checked {
                background-color: #8ab4f8;
                border-color: transparent;
            }
            """
        )
        container = QWidget(self); container.setObjectName("launcherContainer")
        layout = QVBoxLayout(self); layout.setContentsMargins(0, 0, 0, 0); layout.addWidget(container)
        v = QVBoxLayout(container); v.setContentsMargins(8, 8, 8, 8); v.setSpacing(5)

        # 通过三段可伸缩空白保证“画笔”“点名/计时”两侧及中间留白一致
        row = QGridLayout(); row.setContentsMargins(0, 0, 0, 0); row.setHorizontalSpacing(0)
        for col in (0, 2, 4):
            row.setColumnMinimumWidth(col, 12)
            row.setColumnStretch(col, 1)
        self.paint_button = QPushButton("画笔")
        self.paint_button.clicked.connect(self.toggle_paint)
        row.addItem(QSpacerItem(0, 0, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum), 0, 0)
        row.addWidget(self.paint_button, 0, 1)
        row.addItem(QSpacerItem(0, 0, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum), 0, 2)
        self.roll_call_button = QPushButton("点名/计时")
        self.roll_call_button.clicked.connect(self.toggle_roll_call)
        row.addWidget(self.roll_call_button, 0, 3)
        row.addItem(QSpacerItem(0, 0, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum), 0, 4)
        unified_width = self._action_button_width()
        self.paint_button.setFixedWidth(unified_width)
        self.roll_call_button.setFixedWidth(unified_width)
        v.addLayout(row)

        bottom = QHBoxLayout(); bottom.setSpacing(3)
        self.minimize_button = QPushButton("缩小"); self.minimize_button.clicked.connect(self.minimize_launcher)
        bottom.addWidget(self.minimize_button)

        self.autostart_check = QCheckBox("开机启动"); self.autostart_check.stateChanged.connect(self.toggle_autostart); bottom.addWidget(self.autostart_check)
        bottom.addStretch(1)

        self.about_button = QPushButton("关于"); self.about_button.clicked.connect(self.show_about); bottom.addWidget(self.about_button)
        self.exit_button = QPushButton("退出")
        self.exit_button.clicked.connect(self.request_exit)
        bottom.addWidget(self.exit_button)
        v.addLayout(bottom)

        aux_width = max(self.minimize_button.sizeHint().width(), 52)
        self.minimize_button.setFixedWidth(aux_width)
        about_width = max(self.about_button.sizeHint().width(), 52)
        exit_width = max(self.exit_button.sizeHint().width(), 52)
        self.about_button.setFixedWidth(about_width)
        self.exit_button.setFixedWidth(exit_width)

        button_height = max(
            self.paint_button.sizeHint().height(),
            self.roll_call_button.sizeHint().height(),
            self.minimize_button.sizeHint().height(),
            self.about_button.sizeHint().height(),
            self.exit_button.sizeHint().height(),
        )
        for btn in (self.paint_button, self.roll_call_button, self.minimize_button, self.about_button, self.exit_button):
            btn.setFixedHeight(button_height)

        s = self.settings_manager.load_settings().get("Launcher", {})
        x = int(s.get("x", "120")); y = int(s.get("y", "120"))
        self.move(x, y)
        self._last_position = QPoint(x, y)
        bubble_x = int(s.get("bubble_x", str(x)))
        bubble_y = int(s.get("bubble_y", str(y)))
        self._bubble_position = QPoint(bubble_x, bubble_y)
        minimized = str_to_bool(s.get("minimized", "False"), False)
        self._minimized = minimized
        self._minimized_on_start = minimized

        startup = self.settings_manager.load_settings().get("Startup", {})
        autostart_enabled = str_to_bool(startup.get("autostart_enabled", "False"), False)
        self.autostart_check.setChecked(autostart_enabled and WINREG_AVAILABLE)
        self.autostart_check.setEnabled(WINREG_AVAILABLE)

        if not (PANDAS_AVAILABLE and OPENPYXL_AVAILABLE):
            self.roll_call_button.setEnabled(False)

        for w in (self, container, self.paint_button, self.roll_call_button, self.minimize_button, self.autostart_check):
            w.installEventFilter(self)

        # 锁定启动器的推荐尺寸，避免误拖拽造成遮挡
        self.adjustSize()
        self.setFixedSize(self.sizeHint())
        self._base_minimum_width = self.minimumWidth()
        self._base_minimum_height = self.minimumHeight()
        self._ensure_min_width = self.width()
        self._ensure_min_height = self.height()

    def _action_button_width(self) -> int:
        """计算“画笔”与“点名/计时”按钮的统一宽度，保证观感一致。"""

        paint_metrics = QFontMetrics(self.paint_button.font())
        roll_metrics = QFontMetrics(self.roll_call_button.font())
        paint_texts = ["画笔", "隐藏画笔"]
        roll_texts = ["点名/计时", "显示点名", "隐藏点名"]
        max_width = max(
            max(paint_metrics.horizontalAdvance(text) for text in paint_texts),
            max(roll_metrics.horizontalAdvance(text) for text in roll_texts),
        )
        return max_width + 28

    def showEvent(self, e) -> None:
        super().showEvent(e)
        ensure_widget_within_screen(self)
        self._last_position = self.pos()
        if self._minimized_on_start:
            QTimer.singleShot(0, self._restore_minimized_state)
        self.ensure_launcher_topmost()

    def ensure_launcher_topmost(self) -> None:
        if self._minimized:
            if self.bubble is not None and self.bubble.isVisible():
                self.bubble.raise_()
                QTimer.singleShot(0, self.bubble.raise_)
            return
        self.raise_()
        QTimer.singleShot(0, self.raise_)
        if self.bubble is not None and self.bubble.isVisible():
            self.bubble.raise_()
            QTimer.singleShot(0, self.bubble.raise_)

    def eventFilter(self, obj, e) -> bool:
        if e.type() == QEvent.Type.MouseButtonPress and e.button() == Qt.MouseButton.LeftButton:
            self._dragging = True; self._drag_offset = e.globalPosition().toPoint() - self.pos()
        elif e.type() == QEvent.Type.MouseMove and self._dragging and e.buttons() & Qt.MouseButton.LeftButton:
            self.move(e.globalPosition().toPoint() - self._drag_offset)
        elif e.type() == QEvent.Type.MouseButtonRelease and e.button() == Qt.MouseButton.LeftButton:
            self._dragging = False
            self._last_position = self.pos()
            self.save_position()
        return super().eventFilter(obj, e)

    def save_position(self) -> None:
        settings = self.settings_manager.load_settings()
        pos = self._last_position if (self._minimized and not self._last_position.isNull()) else self.pos()
        launcher = settings.get("Launcher", {})
        launcher["x"] = str(pos.x())
        launcher["y"] = str(pos.y())
        bubble_pos = self._bubble_position if not self._bubble_position.isNull() else pos
        launcher["bubble_x"] = str(bubble_pos.x())
        launcher["bubble_y"] = str(bubble_pos.y())
        launcher["minimized"] = bool_to_str(self._minimized)
        settings["Launcher"] = launcher
        startup = settings.get("Startup", {}); startup["autostart_enabled"] = bool_to_str(self.autostart_check.isChecked()); settings["Startup"] = startup
        self.settings_manager.save_settings(settings)

    def toggle_paint(self) -> None:
        """打开或隐藏屏幕画笔覆盖层。"""
        if self.overlay is None: self.overlay = OverlayWindow(self.settings_manager)
        if self.overlay.isVisible():
            self.overlay.hide_overlay(); self.paint_button.setText("画笔")
        else:
            self.overlay.show_overlay(); self.paint_button.setText("隐藏画笔")
        self.ensure_launcher_topmost()

    def toggle_roll_call(self) -> None:
        """切换点名/计时窗口的显示状态，必要时先创建窗口。"""
        if self.roll_call_window is None:
            if not PANDAS_AVAILABLE or not OPENPYXL_AVAILABLE:
                QMessageBox.warning(self, "提示", "未安装 pandas/openpyxl，点名功能不可用。")
                return
            settings = self.settings_manager.load_settings().get("RollCallTimer", {})
            initial_mode = settings.get("mode", "roll_call")
            defer_prompt = initial_mode == "timer"
            if self.student_workbook is None and not defer_prompt:
                workbook = load_student_data(self)
                if workbook is None:
                    QMessageBox.warning(self, "提示", "学生数据加载失败，无法打开点名器。")
                    return
                self.student_workbook = workbook
                if PANDAS_READY:
                    try:
                        self.student_data = workbook.get_active_dataframe()
                    except Exception:
                        self.student_data = pd.DataFrame(columns=["学号", "姓名", "分组", "成绩"])
            self.roll_call_window = RollCallTimerWindow(
                self.settings_manager,
                self.student_workbook,
                parent=self,
                defer_password_prompt=defer_prompt,
            )
            self.roll_call_window.window_closed.connect(self.on_roll_call_window_closed)
            self.roll_call_window.visibility_changed.connect(self.on_roll_call_visibility_changed)
            self.roll_call_window.show()
            self.roll_call_button.setText("隐藏点名")
        else:
            if self.roll_call_window.isVisible():
                self.roll_call_window.hide()
                self.roll_call_button.setText("显示点名")
            else:
                self.roll_call_window.show()
                self.roll_call_window.raise_()
                self.roll_call_button.setText("隐藏点名")

    def on_roll_call_window_closed(self) -> None:
        window = self.roll_call_window
        if window is not None:
            try:
                self.student_workbook = window.student_workbook
                self.student_data = window.student_data
            except Exception:
                pass
        self.roll_call_window = None
        self.roll_call_button.setText("点名/计时")

    def on_roll_call_visibility_changed(self, visible: bool) -> None:
        self.roll_call_button.setText("隐藏点名" if visible else "显示点名")

    def toggle_autostart(self) -> None:
        if not WINREG_AVAILABLE: return
        enabled = self.autostart_check.isChecked()
        self.set_autostart(enabled); self.save_position()

    def request_exit(self) -> None:
        """优雅地关闭应用程序，确保所有窗口在退出前持久化状态。"""

        app = QApplication.instance()
        if not self.close():
            return
        if app is not None:
            QTimer.singleShot(0, app.quit)

    def handle_about_to_quit(self) -> None:
        """在应用退出前的最后一道保险，保证关键状态已经写入配置。"""

        self.save_position()
        window = self.roll_call_window
        if window is not None:
            try:
                window.save_settings()
            except RuntimeError:
                pass

    def minimize_launcher(self, from_settings: bool = False) -> None:
        """将启动器收纳为悬浮圆球。"""

        if self._minimized and not from_settings:
            return
        if self.bubble is None:
            self.bubble = LauncherBubble()
            self.bubble.restore_requested.connect(self.restore_from_bubble)
            self.bubble.position_changed.connect(self._on_bubble_position_changed)
        target_center = self.frameGeometry().center()
        screen = self.screen() or QApplication.screenAt(target_center) or QApplication.primaryScreen()
        if not from_settings:
            self._last_position = self.pos()
        self.hide()
        if from_settings and not self._bubble_position.isNull():
            self.bubble.place_near(self._bubble_position, screen)
        else:
            self.bubble.place_near(target_center, screen)
        self.bubble.setWindowOpacity(0.74)
        self.bubble.show()
        self.bubble.raise_()
        QTimer.singleShot(0, self.bubble.raise_)
        self._minimized = True
        self.save_position()
        self.ensure_launcher_topmost()

    def _restore_minimized_state(self) -> None:
        if not self._minimized_on_start:
            return
        self._minimized_on_start = False
        self.minimize_launcher(from_settings=True)

    def restore_from_bubble(self) -> None:
        """从悬浮球恢复主启动器窗口。"""

        self._minimized = False
        target_pos: Optional[QPoint] = None
        screen = None
        if self.bubble:
            self._bubble_position = self.bubble.pos()
            bubble_geom = self.bubble.frameGeometry()
            bubble_center = bubble_geom.center()
            screen = self.bubble.screen() or QApplication.screenAt(bubble_center) or QApplication.primaryScreen()
            margin = 12
            width = self.width() or self.sizeHint().width()
            height = self.height() or self.sizeHint().height()
            if screen is not None:
                available = screen.availableGeometry()
                distances = {
                    "left": abs(bubble_center.x() - available.left()),
                    "right": abs(available.right() - bubble_center.x()),
                    "top": abs(bubble_center.y() - available.top()),
                    "bottom": abs(available.bottom() - bubble_center.y()),
                }
                nearest_edge = min(distances, key=distances.get)
                if nearest_edge == "left":
                    x = bubble_geom.right() + margin
                    y = bubble_center.y() - height // 2
                elif nearest_edge == "right":
                    x = bubble_geom.left() - width - margin
                    y = bubble_center.y() - height // 2
                elif nearest_edge == "top":
                    y = bubble_geom.bottom() + margin
                    x = bubble_center.x() - width // 2
                else:
                    y = bubble_geom.top() - height - margin
                    x = bubble_center.x() - width // 2
                x = max(available.left(), min(int(x), available.right() - width))
                y = max(available.top(), min(int(y), available.bottom() - height))
                target_pos = QPoint(x, y)
            self.bubble.hide()
        if target_pos is None and not self._last_position.isNull():
            target_pos = QPoint(self._last_position)
        if target_pos is not None:
            self.move(target_pos)
        self.show()
        self.raise_()
        self.activateWindow()
        ensure_widget_within_screen(self)
        self._last_position = self.pos()
        self.save_position()
        self.ensure_launcher_topmost()

    def _on_bubble_position_changed(self, pos: QPoint) -> None:
        self._bubble_position = QPoint(pos.x(), pos.y())
        if self._minimized:
            self.save_position()

    def set_autostart(self, enabled: bool) -> None:
        if not WINREG_AVAILABLE: return
        key_path = r"Software\\Microsoft\\Windows\\CurrentVersion\\Run"
        try:
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path, 0, winreg.KEY_SET_VALUE) as key:
                if enabled:
                    script_path = self.get_script_path()
                    if script_path.lower().endswith((".py", ".pyw")):
                        pythonw = os.path.join(sys.prefix, "pythonw.exe")
                        if not os.path.exists(pythonw): pythonw = sys.executable
                        value = f'"{pythonw}" "{script_path}"'
                    else:
                        value = f'"{script_path}"'
                    winreg.SetValueEx(key, "ClassroomTools", 0, winreg.REG_SZ, value)
                else:
                    try: winreg.DeleteValue(key, "ClassroomTools")
                    except FileNotFoundError: pass
        except PermissionError:
            QMessageBox.warning(self, "提示", "写入开机启动失败，请以管理员身份运行。")
        except Exception as exc:
            QMessageBox.warning(self, "提示", f"设置开机启动失败：{exc}")

    def get_script_path(self) -> str:
        if getattr(sys, "frozen", False): return sys.executable
        return os.path.abspath(sys.argv[0])

    def show_about(self) -> None:
        AboutDialog(self).exec()

    def closeEvent(self, e) -> None:
        self.save_position()
        if self.bubble is not None:
            self.bubble.close()
        if self.roll_call_window is not None: self.roll_call_window.close()
        if self.overlay is not None: self.overlay.close()
        super().closeEvent(e)


# ---------- 入口 ----------
def main() -> None:
    """应用程序入口：初始化 DPI、加载设置并启动启动器窗口。"""
    ensure_high_dpi_awareness()
    app = QApplication(sys.argv)
    app.setQuitOnLastWindowClosed(False)
    QToolTip.setFont(QFont("Microsoft YaHei UI", 9))

    settings_manager = SettingsManager()
    student_workbook = load_student_data(None) if PANDAS_AVAILABLE and not os.path.exists(
        RollCallTimerWindow.ENCRYPTED_STUDENT_FILE
    ) else None

    window = LauncherWindow(settings_manager, student_workbook)
    app.aboutToQuit.connect(window.handle_about_to_quit)
    window.show()
    sys.exit(app.exec())


# Nuitka 打包指令（根据当前依赖整理的推荐参数，保持在单行便于复制）：
# 单文件：python -m nuitka --onefile --enable-plugin=pyqt6 --include-qt-plugins=sensible --windows-disable-console --windows-icon-from-ico=icon.ico --include-data-file=students.xlsx=students.xlsx --include-data-file=settings.ini=settings.ini ClassroomTools.py
# 独立目录：python -m nuitka --standalone --enable-plugin=pyqt6 --include-qt-plugins=sensible --windows-disable-console --windows-icon-from-ico=icon.ico --output-dir=dist --include-data-file=students.xlsx=students.xlsx --include-data-file=settings.ini=settings.ini ClassroomTools.py
if __name__ == "__main__":
    main()
